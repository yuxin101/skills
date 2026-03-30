# template_engine/utils.py
import re
import logging
import pandas as pd
from copy import deepcopy
from typing import Any
from openpyxl.cell import MergedCell

logger = logging.getLogger(__name__)

class TemplateUtils:
    """模板工具函数"""
    
    @staticmethod
    def sanitize_value(value: Any) -> Any:
        """清理数据中的非法字符"""
        if value is None:
            return ""
        if isinstance(value, str):
            value = "".join(c for c in value if c.isprintable())
        return value
    
    @staticmethod
    def format_value(value: Any, data_type: str) -> Any:
        """格式化值"""
        if pd.isna(value) or value is None:
            return ""
        
        if data_type == "date" and hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        elif data_type == "number":
            return float(value) if value else 0
        else:
            return str(value)
    
    @staticmethod
    def copy_row_style(src_row, dst_row):
        """复制行样式"""
        dst_row.height = src_row.height
        for src_cell, dst_cell in zip(src_row, dst_row):
            if isinstance(dst_cell, MergedCell):
                continue
            if src_cell.has_style:
                try:
                    dst_cell.font = deepcopy(src_cell.font)
                    dst_cell.border = deepcopy(src_cell.border)
                    dst_cell.fill = deepcopy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment = deepcopy(src_cell.alignment)
                except Exception as e:
                    logger.error(f"复制样式时出错：{e}")
    
    @staticmethod
    def copy_merged_cells(ws, src_start_row, dst_start_row, row_count):
        """复制合并单元格范围"""
        for merged_range in list(ws.merged_cells.ranges):
            min_row, min_col, max_row, max_col = merged_range.bounds
            if min_row >= src_start_row and max_row < src_start_row + row_count:
                new_min_row = dst_start_row + (min_row - src_start_row)
                new_max_row = dst_start_row + (max_row - src_start_row)
                if new_max_row <= ws.max_row and max_col <= ws.max_column:
                    ws.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)
    
    @staticmethod
    def detect_placeholders(text: str) -> list:
        """检测文本中的占位符"""
        placeholder_pattern = r'\{\{(.*?)\}\}'
        return re.findall(placeholder_pattern, text)