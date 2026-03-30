import os
import re
import logging
import pandas as pd
import openpyxl
from typing import Dict, Any
from functools import lru_cache
logger = logging.getLogger(__name__)

class DataOperation:
    """数据操作服务"""
    
    @staticmethod
    def key_merge(df1, df2, on, how):
        result = pd.merge(df1, df2, on=on, how=how)
        return result

    @staticmethod
    def sheets_concat(input_path, data_key=None, sheet_re=None, axis=0, join='outer',
                      args=None, filter_criteria=None, ignore_index=True, fillna_zero=True):
        """
        合并单个Excel文件中的多个工作表
        
        参数:
        input_path: 输入文件路径
        data_key: 主键列名
        sheet_re: 工作表名称正则表达式
        axis: 合并方向 (0: 垂直, 1: 水平)
        join: 合并方式 ('inner' 或 'outer')
        args: 选择的列名（逗号分隔的字符串）
        filter_criteria: 筛选条件
        ignore_index: 是否忽略索引
        fillna_zero: 是否用0填充空值
        """
        # 处理 args 参数
        if args:
            # 将逗号分隔的字符串拆分为列表
            data_arg = [arg.strip() for arg in args.split(',')]
            index_column = [data_key] + data_arg if data_key else data_arg
        else:
            index_column = [data_key] if data_key else []

        # 读取输入 Excel 文件并获取 Sheet 列表
        excel_file = pd.ExcelFile(input_path)
        sheet_names = excel_file.sheet_names

        # 正则过滤工作表名
        if sheet_re:
            pattern = re.compile(sheet_re)
            sheet_names = [s for s in sheet_names if pattern.fullmatch(s)]

        # 初始化一个空列表，用于存储所有 Sheet 的数据
        dfs = []
        for sheet in sheet_names:
            # 读取数据并填充空值为 0
            df = pd.read_excel(input_path, sheet_name=sheet, dtype=str)
            if fillna_zero:
                df = df.fillna(0)
            df = df.convert_dtypes()

            # 如果指定了主键，检查主键列是否存在
            if data_key and data_key not in df.columns:
                raise ValueError(f"Sheet '{sheet}' 中缺少关键列 {data_key}")

            # 应用筛选条件
            if filter_criteria:
                df = df.query(filter_criteria)

            # 列选择
            if index_column:
                # 检查列名是否存在
                missing_columns = [col for col in index_column if col not in df.columns]
                if missing_columns:
                    raise KeyError(f"Sheet '{sheet}' 中缺少以下列：{missing_columns}")
                df = df[index_column]

            # 如果指定了主键，检查主键列是否有重复值
            if data_key and df[data_key].duplicated().any():
                logger.warning(f"Sheet '{sheet}' 中 '{data_key}' 列存在重复值，已自动去重")
                df = df.drop_duplicates(subset=data_key, keep='first')  # 保留第一个重复值

            # 如果指定了主键，设置主键为索引
            if data_key:
                df = df.set_index(data_key)

            # 横向合并时添加列名前缀
            if axis == 1:
                df = df.add_prefix(f"{sheet}_")

            # 将处理后的数据添加到列表中
            dfs.append(df)

        # 合并所有 Sheet 的数据
        concat_df = pd.concat(dfs, axis=axis, join=join, ignore_index=ignore_index)

        # 如果指定了主键，重置索引
        if data_key and data_key in concat_df.index.names:
            concat_df = concat_df.reset_index().rename(columns={'index': data_key})

        return concat_df

    @staticmethod
    def xlsxs_sheets_concat(input_dir, data_key=None, sheet_re=None, axis=0, join='outer',
                            args=None, filter_criteria=None, ignore_index=True, fillna_zero=True):
        """
        合并多个Excel文件中的所有工作表
        
        参数:
        input_dir: 输入目录路径
        data_key: 主键列名
        sheet_re: 工作表名称正则表达式
        axis: 合并方向 (0: 垂直, 1: 水平)
        join: 合并方式 ('inner' 或 'outer')
        args: 选择的列名（逗号分隔的字符串）
        filter_criteria: 筛选条件
        ignore_index: 是否忽略索引
        fillna_zero: 是否用0填充空值
        """
        xlsx_names = [xlsx_name for xlsx_name in os.listdir(input_dir)
                      if xlsx_name.endswith('.xlsx')]
        dfs = []

        # 处理 args 参数
        if args:
            # 将逗号分隔的字符串拆分为列表
            data_arg = [arg.strip() for arg in args.split(',')]
            index_column = [data_key] + data_arg if data_key else data_arg
        else:
            index_column = [data_key] if data_key else []
            
        for xlsx_name in xlsx_names:
            xlsx_path = os.path.join(input_dir, xlsx_name)
            xlsx = pd.ExcelFile(xlsx_path)
            sheet_names = xlsx.sheet_names

            # 正则过滤工作表名
            if sheet_re:
                pattern = re.compile(sheet_re)
                sheet_names = [s for s in sheet_names if pattern.fullmatch(s)]

            for sheet in sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet)
                if fillna_zero:
                    df = df.fillna(0)
                df = df.convert_dtypes()

                # 检查关键列
                if data_key and data_key not in df.columns:
                    raise ValueError(f"Sheet '{sheet}' 中缺少关键列 {data_key}")

                # 应用筛选条件
                if filter_criteria:
                    df = df.query(filter_criteria)

                # 列选择
                if index_column:
                    # 检查列名是否存在
                    missing_columns = [col for col in index_column if col not in df.columns]
                    if missing_columns:
                        raise KeyError(f"Sheet '{sheet}' 中缺少以下列：{missing_columns}")
                    df = df[index_column]

                # 主键去重与索引设置
                if data_key:
                    if df[data_key].duplicated().any():
                        logger.warning(f"Sheet '{sheet}' 中 '{data_key}' 列存在重复值，已自动去重")
                        df = df.drop_duplicates(subset=data_key, keep='first')
                    df = df.set_index(data_key)

                # 横向合并时添加列名前缀
                if axis == 1:
                    df = df.add_prefix(f"{xlsx_name}_{sheet}_")

                dfs.append(df)

        # 执行合并
        concat_df = pd.concat(dfs, axis=axis, join=join, ignore_index=ignore_index)

        # 重置主键索引
        if data_key and data_key in concat_df.index.names:
            concat_df = concat_df.reset_index().rename(columns={'index': data_key})

        return concat_df

    @staticmethod
    def data_pd_write(save_name, save_sheet_name, df, startrow=0, startcol=0):
        with pd.ExcelWriter(save_name, engine='openpyxl') as writer:
            pd.DataFrame(df).to_excel(writer, sheet_name=save_sheet_name,
                                      startrow=startrow, startcol=startcol)

    @staticmethod
    def read_excel_with_openpyxl(file_path, sheet_name):
        """
        :return: 文件内容的字符串
        """
        try:
            # 加载工作簿
            wb = openpyxl.load_workbook(file_path)
            # 检查工作表是否存在
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            # 获取工作表
            ws = wb[sheet_name]
            # 读取内容并转换为字符串
            content = []
            for row in ws.iter_rows(values_only=True):
                content.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            return "\n".join(content)
        except Exception as e:
            # 如果出错，返回错误信息
            return f"读取文件时出错: {e}"