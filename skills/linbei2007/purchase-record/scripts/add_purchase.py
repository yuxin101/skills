#!/usr/bin/env python3
"""
采购记录写入脚本
将采购信息追加到 Excel 表格
"""

import openpyxl
import re
import sys
from pathlib import Path

# Excel 文件路径
EXCEL_PATH = r"C:\Users\Administrator.rjazz-2022BWPUD\Desktop\purchase_record.xlsx"


def parse_purchase_command(message: str) -> dict | None:
    """
    解析采购命令
    格式：采购 MMDD 物品名称 X 元/X 块/XX 数字
    返回：{date, item_name, price} 或 None（如果格式不对）
    """
    # 匹配"采购 MMDD 物品名称 X 元/X 块钱/XX"格式
    pattern = r'^采购\s+(\d{4})\s+(.+?)\s+([\d.]+)\s*(?:元 | 块钱？|刀)?$/i'
    match = re.match(pattern, message.strip(), re.IGNORECASE)

    if not match:
        return None

    date_code, item_name, price_str = match.groups()

    # 解析日期（MMDD -> MM-DD）
    month = int(date_code[0:2])
    day = int(date_code[2:4])
    
    if not (1 <= month <= 12):
        return None
    
    formatted_date = f"{month:02d}-{day:02d}"

    # 解析价格
    try:
        price = float(price_str)
    except ValueError:
        return None

    return {
        'date': formatted_date,
        'item_name': item_name.strip(),
        'price': price
    }


def add_to_excel(data: dict) -> bool:
    """
    将采购记录写入 Excel
    找到第一非空白行并写入数据
    """
    try:
        # 加载工作簿（如果不存在则创建）
        if Path(EXCEL_PATH).exists():
            wb = openpyxl.load_workbook(EXCEL_PATH)
        else:
            wb = openpyxl.Workbook()

        ws = wb.active

        # 设置表头（如果是新文件）
        if ws.max_row == 1 and not all(ws.cell(row=1, column=c).value for c in range(1, 4)):
            ws['A1'] = '日期'
            ws['B1'] = '物品名称'
            ws['C1'] = '价格'

        # 查找第一非空白行（跳过表头）
        first_empty_row = None
        for row in range(2, ws.max_row + 2):
            if all(ws.cell(row=row, column=col).value is None for col in [1, 2, 3]):
                first_empty_row = row
                break

        # 如果没有找到空白行，在最后一行后面添加
        if not first_empty_row:
            first_empty_row = ws.max_row + 1

        # 写入数据
        ws.cell(row=first_empty_row, column=1, value=data['date'])
        ws.cell(row=first_empty_row, column=2, value=data['item_name'])
        ws.cell(row=first_empty_row, column=3, value=data['price'])

        # 保存并关闭
        wb.save(EXCEL_PATH)
        wb.close()

        return True
    except Exception as e:
        print(f"❌ 写入 Excel 失败：{e}", file=sys.stderr)
        return False


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("用法：python add_purchase.py \"采购 MMDD 物品名称 价格\"")
        sys.exit(1)

    message = ' '.join(sys.argv[1:])
    
    # 解析命令
    data = parse_purchase_command(message)
    if not data:
        print("❌ 格式不正确。请使用：采购 MMDD 物品名称 价格")
        print("例如：采购 0312 螺丝 3 元")
        sys.exit(1)

    # 写入 Excel
    success = add_to_excel(data)
    
    if success:
        reply = f"✅ **采购记录已添加**\n\n📅 日期：{data['date']}\n🛒 物品：{data['item_name']}\n💰 价格：{data['price']}元"
        print(reply)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
