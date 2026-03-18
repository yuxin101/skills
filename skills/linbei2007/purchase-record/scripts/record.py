#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
采购记录录入工具
将输入的数据写入 purchase_record.xlsx 的最后一行之后的下一行（追加模式）
"""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime


def parse_date(date_str):
    """解析月日格式为 YYYY-MM-DD，默认为当前年份"""
    if len(date_str) != 4 or not date_str.isdigit():
        raise ValueError("日期必须是四位数字，如 '0312'")

    month = int(date_str[:2])
    day = int(date_str[2:])

    if not (1 <= month <= 12):
        raise ValueError("月份必须在 1-12 之间")
    if not (1 <= day <= 31):
        raise ValueError("日期必须在 1-31 之间")

    current_year = datetime.now().year
    return "{}-{:02d}-{:02d}".format(current_year, month, day)


def parse_price(price_str):
    """解析价格，只保留数字部分"""
    # 提取所有数字和小数点
    numbers = re.findall(r'\d+\.?\d*', price_str)
    if not numbers:
        raise ValueError("价格必须是数字")

    return float(numbers[0])


def find_last_data_row(wb):
    """找到最后一个有数据的行号，返回 (last_row, sheet_name)"""
    # 获取所有工作表（支持不同大小写）
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 从第 2 行开始检查（假设第 1 行是表头）
        last_row = None
        for row in range(2, ws.max_row + 1):
            has_data = False
            for col in [1, 2, 3]:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != '':
                    has_data = True
                    break

            if has_data:
                last_row = row

        if last_row or last_row == 1:  # 如果至少找到一行数据或只有表头
            return (last_row, sheet_name)

    # 如果没有找到任何数据行，返回第 2 行（假设第 1 行是表头）
    first_sheet = wb.sheetnames[0]
    ws = wb[first_sheet]
    if ws.max_row >= 1:
        return (ws.max_row, first_sheet)
    return (1, first_sheet)


def record_purchase(date_str, name, price_str):
    """录入采购记录"""

    # 解析参数
    date = parse_date(date_str)
    price = parse_price(price_str)

    # Excel 文件路径
    excel_path = Path.home() / "Desktop" / "purchase_record.xlsx"

    if not excel_path.exists():
        raise FileNotFoundError("找不到 Excel 文件：" + str(excel_path))

    # 尝试加载工作簿，处理可能的锁定问题
    import time
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb = load_workbook(str(excel_path), keep_vba=False)
            break
        except PermissionError as e:
            if attempt < max_retries - 1:
                time.sleep(2)
            else:
                raise RuntimeError("无法打开 Excel 文件，可能是被其他程序占用：" + str(e))

    # 查找最后一个有数据的行（追加模式）
    last_row, sheet_name = find_last_data_row(wb)
    write_row = last_row + 1
    ws = wb[sheet_name]

    # 写入数据
    ws.cell(row=write_row, column=1).value = date      # A 列：日期
    ws.cell(row=write_row, column=2).value = name      # B 列：名称
    ws.cell(row=write_row, column=3).value = price     # C 列：价格

    # 保存并关闭
    wb.save(excel_path)
    print("采购记录已录入！")
    print("   日期：" + date)
    print("   名称：" + name)
    print("   价格：" + str(price))

    wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("用法错误！")
        print("格式：python record.py [日期] [物品名称] [价格]")
        print("示例：python record.py 0312 螺丝 3")
        sys.exit(1)

    date = sys.argv[1]
    name = " ".join(sys.argv[2:-1])  # 除了第一个和最后一个参数，中间的都是名称
    price = sys.argv[-1]

    try:
        record_purchase(date, name, price)
    except Exception as e:
        print("错误：" + str(e))
        sys.exit(1)
