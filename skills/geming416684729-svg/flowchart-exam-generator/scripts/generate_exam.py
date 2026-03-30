#!/usr/bin/env python3
"""
流程图考试题生成脚本
根据流程图分析结果生成标准格式的应知应会考试题Excel
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 答案格式：A=1, B=2, C=3, D=4
ANSWER_MAP = {'A': '1', 'B': '2', 'C': '3', 'D': '4'}

def create_excel(exam_questions, output_path='考试题.xlsx'):
    """
    创建考试题Excel文件

    Args:
        exam_questions: 考试题列表，每道题为dict，包含题型、题目、选项、正确答案、解析、出处
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "考试题库"

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 60
    ws.column_dimensions['I'].width = 15

    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = ['题型', '题目', '选项1', '选项2', '选项3', '选项4', '正确答案', '解析', '出处']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 写入数据
    for row, question in enumerate(exam_questions, 2):
        ws.cell(row=row, column=1, value=question.get('题型', '')).border = thin_border
        ws.cell(row=row, column=2, value=question.get('题目', '')).border = thin_border
        ws.cell(row=row, column=3, value=question.get('选项1', '')).border = thin_border
        ws.cell(row=row, column=4, value=question.get('选项2', '')).border = thin_border
        ws.cell(row=row, column=5, value=question.get('选项3', '')).border = thin_border
        ws.cell(row=row, column=6, value=question.get('选项4', '')).border = thin_border
        ws.cell(row=row, column=7, value=question.get('正确答案', '')).border = thin_border
        ws.cell(row=row, column=8, value=question.get('解析', '')).border = thin_border
        ws.cell(row=row, column=9, value=question.get('出处', '')).border = thin_border

        # 设置对齐方式
        for col in range(1, 10):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True
            )

    # 保存文件
    wb.save(output_path)
    print(f"考试题已生成并保存至: {output_path}")

    # 统计
    single_count = sum(1 for q in exam_questions if q.get('题型') == '单选题')
    multi_count = sum(1 for q in exam_questions if q.get('题型') == '多选题')
    print(f"共计 {len(exam_questions)} 道题目")
    print(f"- 单选题: {single_count} 道")
    print(f"- 多选题: {multi_count} 道")
    print("\n答案格式说明：")
    print("- A=1, B=2, C=3, D=4")
    print("- 多选题答案示例: 1,3,4 表示选A、C、D")


def convert_answer(answer_str):
    """
    将字母格式的答案转换为数字格式
    例如: 'ABCD' -> '1,2,3,4', 'AC' -> '1,3'
    """
    result = []
    for char in answer_str.upper():
        if char in ANSWER_MAP:
            result.append(ANSWER_MAP[char])
    return ','.join(result)


if __name__ == '__main__':
    # 示例数据
    sample_questions = [
        {
            '题型': '单选题',
            '题目': '测试流程，流程从（）部门开始。',
            '选项1': 'A部门',
            '选项2': 'B部门',
            '选项3': 'C部门',
            '选项4': 'D部门',
            '正确答案': '2',
            '解析': '根据流程图，流程从B部门开始。',
            '出处': '测试流程'
        }
    ]

    if len(sys.argv) > 1:
        # 从命令行参数读取题目数据(JSON格式)
        try:
            questions = json.loads(sys.argv[1])
            output = sys.argv[2] if len(sys.argv) > 2 else '考试题.xlsx'
            create_excel(questions, output)
        except json.JSONDecodeError as e:
            print(f"JSON解析错误: {e}")
            sys.exit(1)
    else:
        create_excel(sample_questions)
