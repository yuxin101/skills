"""表格拆分 Skill。

将单个 Excel/CSV 文件拆分为多个独立的表格文件。

Usage:
    python split_table_skill.py <input_file> [output_dir]

Arguments:
    input_file: 输入文件路径 (.xlsx, .xls, .csv)
    output_dir: 输出目录（可选，默认为输入文件所在目录）

Example:
    python split_table_skill.py /path/to/input.xlsx /path/to/output
    python split_table_skill.py /path/to/input.xlsx  # 输出到 /path/to/
"""

import os
import sys
import asyncio
import argparse
from typing import Any, Dict, List

import pandas as pd

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(script_dir))

from script.src import (
    split_workbook_all_sheets_return_worksheets,
    preprocess_excel,
    preprocess_merge,
    remove_blank_rows_and_columns,
    ws_to_df,
    split_df_by_blank_rows,
    preprocess_merge_xlrd,
)


class SplitTableSkill:
    """表格拆分技能。

    将单个文件拆分为多个独立的表格文件。
    输入：文件路径
    输出：拆分后的文件路径列表
    """

    def __init__(self):
        pass

    async def run(self, input_file: str, output_dir: str = None) -> Dict[str, Any]:
        """执行表格拆分。

        Args:
            input_file: 输入文件路径
            output_dir: 输出目录（可选）

        Returns:
            包含以下键的字典：
                - split_files: 拆分后的文件路径列表
                - table_count: 表格数量
                - success: 是否成功
        """
        if not os.path.exists(input_file):
            return {
                "split_files": [],
                "table_count": 0,
                "success": False,
                "error": f"文件不存在: {input_file}"
            }

        if output_dir is None:
            output_dir = os.path.dirname(os.path.abspath(input_file))

        os.makedirs(output_dir, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        split_files = []
        try:
            if input_file.endswith('.csv'):
                split_files = await self._split_csv(input_file, output_dir, base_name)
            else:
                split_files = await self._split_excel(input_file, output_dir, base_name)
            return {
                "split_files": split_files,
                "table_count": len(split_files),
                "success": True,
                "output_dir": output_dir
            }
        except Exception as e:
            return {
                "split_files": [],
                "table_count": 0,
                "success": False,
                "error": str(e)
            }

    async def _split_csv(self, file_path: str, output_dir: str, base_name: str) -> List[str]:
        """拆分 CSV 文件。

        Args:
            file_path: CSV 文件路径
            output_dir: 输出目录
            base_name: 文件基础名

        Returns:
            拆分后的文件路径列表
        """
        df = self._read_csv(file_path)
        df = remove_blank_rows_and_columns(df)
        sub_dfs = self._split_by_empty_rows(df)
        split_files = []
        for idx, sub_df in enumerate(sub_dfs):
            if idx == 0:
                out_name = f"{base_name}.csv"
            else:
                out_name = f"{base_name}_{idx + 1}.csv"
            out_path = os.path.join(output_dir, out_name)
            sub_df.to_csv(out_path, index=False, header=False, encoding='utf-8')
            split_files.append(out_path)
        return split_files

    async def _split_excel(self, file_path: str, output_dir: str, base_name: str) -> List[str]:
        """拆分 Excel 文件。

        Args:
            file_path: Excel 文件路径
            output_dir: 输出目录
            base_name: 文件基础名

        Returns:
            拆分后的文件路径列表
        """
        try:
            return await self._split_excel_spire(file_path, output_dir, base_name)
        except Exception as e:
            print(f"Spire 拆分失败，使用回退方案: {e}")
            return await self._split_excel_fallback(file_path, output_dir, base_name)

    async def _split_excel_spire(self, file_path: str, output_dir: str, base_name: str) -> List[str]:
        """使用 Spire 拆分 Excel。

        Args:
            file_path: Excel 文件路径
            output_dir: 输出目录
            base_name: 文件基础名

        Returns:
            拆分后的文件路径列表
        """
        worksheets = split_workbook_all_sheets_return_worksheets(file_path)

        from spire.xls import Workbook
        wb_template = Workbook()
        template_path = os.path.join(script_dir, '2012 最新版土地开发整理项目人工预算单价计算表.xls')
        standard_style = None
        if os.path.exists(template_path):
            wb_template.LoadFromFile(template_path)
            standard_style = wb_template.Worksheets[0]['A1'].Style

        split_files = []
        for idx, sheet in enumerate(worksheets):
            if sheet.IsEmpty:
                continue

            df = preprocess_excel(sheet, remove_chars=[','], standard_style=standard_style)
            df = remove_blank_rows_and_columns(df)

            if idx == 0:
                out_name = f"{base_name}.csv"
            else:
                out_name = f"{base_name}_{idx + 1}.csv"

            out_path = os.path.join(output_dir, out_name)
            df.to_csv(out_path, index=False, header=False, encoding='utf-8')
            split_files.append(out_path)

        return split_files

    async def _split_excel_fallback(self, file_path: str, output_dir: str, base_name: str) -> List[str]:
        """使用 openpyxl/xlrd 拆分 Excel（回退方案）。

        Args:
            file_path: Excel 文件路径
            output_dir: 输出目录
            base_name: 文件基础名

        Returns:
            拆分后的文件路径列表
        """
        split_files = []

        if file_path.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                preprocess_merge(ws)
                # 将 ws 转换为 DataFrame
                df = ws_to_df(ws)
                
                # 先移除整个 sheet 开头和结尾的空白行列
                df = remove_blank_rows_and_columns(df)
                
                if df.empty:
                    continue
                # 再按空行分割
                tables, filenames = split_df_by_blank_rows(df, sheet_name)

                for table, filename in zip(tables, filenames):
                    if table.empty:
                        continue

                    out_name = f"{filename}.csv"
                    out_path = os.path.join(output_dir, out_name)
                    table.to_csv(out_path, index=False, header=False, encoding='utf-8')
                    split_files.append(out_path)
        else:
            from xlrd import open_workbook
            excel = pd.ExcelFile(
                open_workbook(file_path, formatting_info=True),
                engine="xlrd"
            )

            for sheet_name in excel.sheet_names:
                sheet = excel.book[sheet_name]
                if sheet.nrows > 0 and sheet.ncols > 0:
                    
                    # 预处理合并单元格
                    merge_map = {(r, c): v for (r, c, v) in preprocess_merge_xlrd(sheet)}
                    
                    # 将 xlrd sheet 转换为 DataFrame
                    data = []
                    for i in range(sheet.nrows):
                        row = [merge_map.get((i, j), sheet.cell_value(i, j)) for j in range(sheet.ncols)]
                        data.append(row)
                    df = pd.DataFrame(data)
                    # 先移除整个 sheet 开头和结尾的空白行列
                    df = remove_blank_rows_and_columns(df)
                    if df.empty:
                        continue

                    # 再按空行分割
                    tables, filenames = split_df_by_blank_rows(df, sheet_name)

                    for table, filename in zip(tables, filenames):
                        if table.empty:
                            continue

                        out_name = f"{filename}.csv"
                        out_path = os.path.join(output_dir, out_name)
                        table.to_csv(out_path, index=False, header=False, encoding='utf-8')
                        split_files.append(out_path)

        return split_files

    def _read_csv(self, file_path: str) -> pd.DataFrame:
        """读取 CSV，支持编码回退。

        Args:
            file_path: CSV 文件路径

        Returns:
            DataFrame
        """
        try:
            return pd.read_csv(file_path, header=None)
        except:
            return pd.read_csv(file_path, encoding='gbk', header=None)

    def _split_by_empty_rows(self, df: pd.DataFrame) -> List[pd.DataFrame]:
        """按空行分割 DataFrame。

        Args:
            df: 输入 DataFrame

        Returns:
            分割后的 DataFrame 列表
        """
        df_str = df.astype(str)
        empty_mask = (df_str == 'nan') | (df_str == '') | df_str.isnull()
        empty_rows = empty_mask.all(axis=1)
        empty_indices = empty_rows[empty_rows].index.tolist()
        if not empty_indices:
            return [df]
        split_points = [-1] + empty_indices + [len(df)]
        sub_dfs = []
        for i in range(len(split_points) - 1):
            start = split_points[i] + 1
            end = split_points[i + 1]
            if start < end:
                sub_dfs.append(df.iloc[start:end].copy())
        return sub_dfs


async def run_skill(input_file: str, output_dir: str = None) -> Dict[str, Any]:
    """运行表格拆分技能。

    Args:
        input_file: 输入文件路径
        output_dir: 输出目录

    Returns:
        拆分结果
    """
    skill = SplitTableSkill()
    return await skill.run(input_file, output_dir)


def main():
    parser = argparse.ArgumentParser(description="表格拆分 Skill")
    parser.add_argument("input_file", help="输入文件路径 (.xlsx, .xls, .csv)")
    parser.add_argument("output_dir", nargs="?", default=None, help="输出目录（可选）")

    args = parser.parse_args()
    print("=" * 60, "Split Table Skill", "=" * 60, sep="\n")
    print(f"Input file: {args.input_file}")
    print(f"Output directory: {args.output_dir or '默认与输入同目录'}\n")
    
    result = asyncio.run(run_skill(args.input_file, args.output_dir))
    if result["success"]:
        print(f"✓ 拆分成功！共 {result['table_count']} 个表格\n")
        print("输出文件:")
        for f in result["split_files"]:
            print(f"  - {os.path.basename(f)}")
    else:
        print(f"✗ 拆分失败: {result.get('error', '未知错误')}\n")
    print("=" * 60)
    
    return 0 if result["success"] else 1


if __name__ == "__main__":
    sys.exit(main())
