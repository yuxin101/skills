#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""表结构处理脚本"""
import os, sys, glob, shutil

def ensure_openpyxl():
    try:
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        return openpyxl, load_workbook, Font, Alignment, Border, Side, PatternFill, get_column_letter
    except ImportError:
        lib_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib")
        sys.path.insert(0, lib_dir)
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        return openpyxl, load_workbook, Font, Alignment, Border, Side, PatternFill, get_column_letter

openpyxl, load_workbook, Font, Alignment, Border, Side, PatternFill, get_column_letter = ensure_openpyxl()


# ── 表格样式（从 data/cellStyle.txt 读取并应用）────────────────────────
def apply_cell_styles(ws, max_row, max_col):
    skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    style_file = os.path.join(skill_dir, "data", "cellStyle.txt")
    if not os.path.exists(style_file):
        print("  - 未找到 cellStyle.txt，跳过样式应用")
        return

    # 基准字体：宋体 11
    base_font = Font(name="宋体", size=11)

    # 第一行：白字 + 绿底
    white_font = Font(name="宋体", size=11, color="FFFFFF", bold=True)
    green_fill = PatternFill(fill_type="solid", fgColor="00B050")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = white_font
        cell.fill = green_fill

    # 从第二行起：对齐 + 字体
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    center_cols = {"A","B","E","H","I","J","K","L"}
    left_cols   = {"C","D","F","G","M","N"}
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            letter = get_column_letter(col)
            if letter in center_cols:
                cell.alignment = center_align
            elif letter in left_cols:
                cell.alignment = left_align
            if not isinstance(cell.font, Font) or cell.font.name != "宋体":
                cell.font = base_font

    # 所有列宽自适应（英文×1.2，中文按实际宽度×1.8）
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_w = 0
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                w = sum(2.0 if '\u4e00' <= c <= '\u9fff' else 1.2 for c in str(val))
                max_w = max(max_w, w)
        ws.column_dimensions[letter].width = max(max_w + 2, 8)

    print("  - 表格样式应用 ✅")


def convert_to_pinyin_initial(chinese_text):
    """将中文转换为拼音首字母，'-'转换为'_'；未知字符用 pypinyin 兜底"""
    PYMAP = {
        "个":"G","人":"R","变":"B","动":"D","数":"S","据":"J","账":"Z","户":"H",
        "冻":"D","结":"J","明":"M","细":"X","信":"X","息":"X","住":"Z","房":"F",
        "贷":"D","款":"K","业":"Y","务":"W","过":"G","程":"C","逾":"Y","期":"Q",
        "注":"Z","销":"X","撤":"C","核":"H","对":"D","余":"Y","额":"E","转":"Z",
        "移":"Y","冲":"C","正":"Z","调":"T","整":"Z","利":"L","息":"X","还":"H",
        "抵":"D","提":"T","取":"Q","汇":"H","出":"C","账":"Z","户":"H","公":"G",
        "积":"J","金":"J","买":"M","受":"S","让":"R","租":"Z","赁":"L","银":"Y",
        "行":"H","卡":"K","年":"N","月":"Y","日":"R","序":"X","号":"H","标":"B",
        "识":"S","码":"M","描":"M","述":"S","类":"L","型":"X","长":"C","度":"D",
        "约束":"YS","索":"S","引":"Y","主":"Z","键":"J","非":"F","空":"K",
        "唯一":"WYZ","创":"C","建":"J","说":"S","明":"M","质":"Z","量":"L",
        "规":"G","则":"Z","填":"T","写":"X","格":"G","式":"S","属":"S","性":"X",
        "值":"Z","域":"Y","范":"F","围":"W","计":"J","算":"S","查":"C","校":"J",
        "准":"Z","错":"C","误":"W","报":"B","告":"G","审":"S","批":"P","通":"T",
        "过":"G","状":"Z","态":"T","启":"Q","用":"Y","禁":"J","止":"Z","删":"S",
        "除":"C","修":"X","改":"G","看":"K","览":"L","编":"B","辑":"J","名":"M",
        "称":"C","英":"Y","文":"W","中":"Z","汉":"H","拼":"P","音":"Y","首":"S",
        "字":"Z","母":"M","面":"M","积":"J","总":"Z","高":"G","宽":"K","比":"B",
        "例":"L","率":"V","含":"H","义":"Y","固":"G","定":"D","名":"M","格":"G",
        "称":"C","描":"M","述":"S","码":"M","编":"B","号":"H","性":"X","质":"Z",
    }
    # pypinyin 懒人模式兜底
    try:
        importlib = __import__('importlib')
        importlib.invalidate_caches()
        from pypinyin import lazy_pinyin as _lp
        HAS_PYPINYIN = True
    except ImportError:
        HAS_PYPINYIN = False

    result = []
    unknown_chars = []
    for ch in chinese_text:
        if ch == "-":
            result.append("_")
        elif "\u4e00" <= ch <= "\u9fff":
            pinyin_ch = PYMAP.get(ch)
            if pinyin_ch:
                result.append(pinyin_ch)
            else:
                unknown_chars.append(ch)
                result.append(ch)   # 暂存原字，后续替换
        else:
            result.append(ch)

    # 对未知汉字用 pypinyin 补全
    if unknown_chars and HAS_PYPINYIN:
        pinyins = _lp(unknown_chars)
        # 将原位置的汉字替换为拼音首字母
        pi_idx = 0
        for i, ch in enumerate(result):
            if ch in unknown_chars:
                result[i] = pinyins[pi_idx][0].upper()
                pi_idx += 1

    return "".join(result)


# ── 表头从 data/addFields.md 读取 ──────────────────────────────────────────
def load_headers():
    skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    header_file = os.path.join(skill_dir, "data", "addFields.md")
    with open(header_file, "r", encoding="utf-8") as f:
        content = f.read().strip()
    if "，" in content or "," in content:
        sep = "，" if "，" in content else ","
        headers = [h.strip() for h in content.split(sep) if h.strip()]
    else:
        headers = [h.strip() for h in content.split("\n") if h.strip()]
    if not headers:
        raise ValueError("addFields.md 内容为空或格式错误")
    return headers


# ── 插入行从 data/addRow.md 读取 ─────────────────────────────────────────────
def load_add_row():
    skill_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    row_file = os.path.join(skill_dir, "data", "addRow.md")
    result = {}
    with open(row_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or "|" not in line:
                continue
            # 跳过 markdown 表格表头行（如"列号|列中文名"）和分隔行（如"|----|"）
            if any(kw in line for kw in ["列号", "列中文名", "----", "|---"]):
                continue
            parts = line.split("|")
            if len(parts) >= 3:
                col = parts[1].strip().upper()
                val = parts[2].strip()
                if col and val:
                    result[col] = val
    return result


HEADERS = load_headers()
START_COL = 6  # F 列


def find_input_file(specified_name=None):
    user_dir = "/workspace/user_input_files"
    if specified_name:
        path = os.path.join(user_dir, specified_name)
        if os.path.exists(path):
            return path
        if not specified_name.endswith(".xlsx"):
            path2 = path + ".xlsx"
            if os.path.exists(path2):
                return path2
        raise FileNotFoundError(f"文件不存在: {specified_name}")
    files = glob.glob(os.path.join(user_dir, "*.xlsx"))
    if not files:
        raise FileNotFoundError(f"{user_dir} 目录中没有 Excel 文件")
    return max(files, key=os.path.getmtime)


def process(input_path, output_path=None):
    print(f"[输入] {input_path}")

    if output_path is None:
        base = os.path.splitext(os.path.basename(input_path))[0]
        output_dir = "/workspace/skills_output"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, base + "_processed.xlsx")

    # 读取源文件工作表名称
    wb_src = load_workbook(input_path)
    sheet_name = wb_src.sheetnames[0]
    wb_src.close()
    pinyin_name = convert_to_pinyin_initial(sheet_name)
    print(f"  - 工作表名称: {sheet_name}")
    print(f"  - 拼音首字母: {pinyin_name}")

    shutil.copy(input_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Step 0: 取消所有合并单元格
    for merged_range in list(ws.merged_cells):
        ws.unmerge_cells(str(merged_range))

    # Step 1: 删除第一行（原标题行）
    ws.delete_rows(1)
    print("  - 删除第一行 ✅")

    # Step 2: 在新的第一行 F~N 写入表头
    for i, header in enumerate(HEADERS):
        col = START_COL + i
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    print(f"  - F列起写入{len(HEADERS)}列表头 ✅")

    # Step 3: 在第2行之上插入一行，按 addRow.md 填充
    ws.insert_rows(2)
    add_row_values = load_add_row()
    col_map = {c: idx + 1 for idx, c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ")}
    for col_letter, val in add_row_values.items():
        if col_letter in col_map:
            ws.cell(row=2, column=col_map[col_letter]).value = val
            print(f"  - 插入行 {col_letter}2 = {val}")
        else:
            print(f"  - 跳过列 {col_letter}（超出范围）")

    # Step 4: A列序号重编（从第2行起，1,2,3...）
    max_row_now = ws.max_row
    for seq, row_idx in enumerate(range(2, max_row_now + 1), start=1):
        ws.cell(row=row_idx, column=1).value = seq
    print("  - A列序号重编 ✅")

    # Step 5: E/F/G/H 列填充（E列从第3行起，F/G/H从第2行起）
    for row_idx in range(2, max_row_now + 1):
        ws.cell(row=row_idx, column=6).value = pinyin_name
        ws.cell(row=row_idx, column=7).value = sheet_name
        ws.cell(row=row_idx, column=8).value = "VARCHAR2"
    for row_idx in range(3, max_row_now + 1):
        ws.cell(row=row_idx, column=5).value = "M"

    # Step 6: 行高 30
    for row in range(1, max_row_now + 1):
        ws.row_dimensions[row].height = 30

    # Step 7: 所有单元格加 thin 边框
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for row in range(1, max_row_now + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = thin

    # Step 8: F~N 列宽（英文×1，中文×1.8，自适应含表头+数据）
    for i in range(len(HEADERS)):
        col = START_COL + i
        letter = get_column_letter(col)
        max_w = 0
        for row in range(1, max_row_now + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                w = sum(2.0 if '\u4e00' <= c <= '\u9fff' else 1.2 for c in str(val))
                max_w = max(max_w, w)
        ws.column_dimensions[letter].width = max(max_w + 2, 10)

    # Step 9: 应用表格样式（根据 cellStyle.txt）
    apply_cell_styles(ws, max_row_now, ws.max_column)

    wb.save(output_path)
    print(f"[输出] {output_path}")
    return output_path


if __name__ == "__main__":
    args = sys.argv[1:]
    input_name = args[0] if args else None
    output_name = args[1] if len(args) > 1 else None
    try:
        input_path = find_input_file(input_name)
        output_path = None
        if output_name:
            output_path = os.path.join("/workspace/user_input_files", output_name)
        result = process(input_path, output_path)
        print(f"\n✅ 处理完成\n文件已保存至: {result}")
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
