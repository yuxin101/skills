#!/usr/bin/env python3
import urllib.request, urllib.parse, json, xml.etree.ElementTree as ET
import datetime, sys, os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('正在安装 openpyxl...')
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

TODAY = datetime.date.today().isoformat()

ARXIV_CATEGORIES = ['q-fin.TR','q-fin.PM','q-fin.RM','q-fin.MF','q-fin.PR','q-fin.CP','q-fin.ST','q-fin.EC','q-fin.GN']

def fetch_arxiv(max_results=30):
    print('📡 抓取 arXiv q-fin...')
    papers = []
    cats = '+OR+'.join(f'cat:{c}' for c in ARXIV_CATEGORIES)
    query = urllib.parse.quote(f'({cats})')
    url = (f'http://export.arxiv.org/api/query?search_query={query}&sortBy=submittedDate&sortOrder=descending&max_results={max_results}')
    try:
        with urllib.request.urlopen(url, timeout=20) as r:
            xml_data = r.read()
        ns = {'atom':'http://www.w3.org/2005/Atom'}
        root = ET.fromstring(xml_data)
        for entry in root.findall('atom:entry', ns):
            title = entry.find('atom:title', ns)
            summary = entry.find('atom:summary', ns)
            published = entry.find('atom:published', ns)
            link = entry.find('atom:id', ns)
            authors = entry.findall('atom:author', ns)
            cats_el = entry.findall('atom:category', ns)
            category = cats_el[0].get('term','') if cats_el else ''
            author_names = [a.find('atom:name', ns).text for a in authors[:3] if a.find('atom:name', ns) is not None]
            papers.append({'来源':'arXiv','标题':(title.text or '').strip().replace('\n',' ') if title is not None else '','作者':', '.join(author_names),'摘要':(summary.text or '').strip().replace('\n',' ')[:500] if summary is not None else '','发布日期':(published.text or '')[:10] if published is not None else '','分类':category,'链接':link.text.strip() if link is not None else ''})
    except Exception as e:
        print(f'  ⚠️ arXiv 失败: {e}')
    print(f'  ✓ arXiv: {len(papers)} 篇')
    return papers

def fetch_semantic_scholar(max_results=20):
    print('📡 抓取 Semantic Scholar...')
    papers = []
    keywords = ['quantitative finance','asset pricing','machine learning finance','risk management','algorithmic trading']
    seen = set()
    for kw in keywords:
        if len(papers) >= max_results: break
        query = urllib.parse.quote(kw)
        url = (f'https://api.semanticscholar.org/graph/v1/paper/search?query={query}&fields=title,authors,abstract,year,externalIds,publicationDate,url&limit=10&sort=publicationDate')
        try:
            req = urllib.request.Request(url, headers={'User-Agent':'finance-paper-bot/1.0'})
            with urllib.request.urlopen(req, timeout=20) as r:
                data = json.loads(r.read())
            for p in data.get('data', []):
                pid = p.get('paperId','')
                if pid in seen: continue
                seen.add(pid)
                pub_date = p.get('publicationDate') or str(p.get('year',''))
                authors = [a['name'] for a in p.get('authors',[])[:3]]
                ext = p.get('externalIds',{})
                link = p.get('url') or (f'https://arxiv.org/abs/{ext["ArXiv"]}' if ext.get('ArXiv') else f'https://www.semanticscholar.org/paper/{pid}')
                papers.append({'来源':'Semantic Scholar','标题':p.get('title',''),'作者':', '.join(authors),'摘要':(p.get('abstract') or '')[:500],'发布日期':pub_date[:10] if len(pub_date)>=10 else pub_date,'分类':kw,'链接':link})
        except Exception as e:
            print(f'  ⚠️ SS ({kw}): {e}')
    print(f'  ✓ Semantic Scholar: {len(papers)} 篇')
    return papers

def fetch_openalex(max_results=20):
    print('📡 抓取 OpenAlex...')
    papers = []
    concept_ids = ['C144235515','C162324750','C27206212','C144133560']
    seen = set()
    for cid in concept_ids:
        if len(papers) >= max_results: break
        from_date = (datetime.date.today()-datetime.timedelta(days=30)).isoformat()
        url = (f'https://api.openalex.org/works?filter=concepts.id:{cid},from_publication_date:{from_date}&sort=publication_date:desc&per-page=10&mailto=research@example.com')
        try:
            req = urllib.request.Request(url, headers={'User-Agent':'finance-paper-bot/1.0'})
            with urllib.request.urlopen(req, timeout=20) as r:
                data = json.loads(r.read())
            for w in data.get('results',[]):
                wid = w.get('id','')
                if wid in seen: continue
                seen.add(wid)
                title = w.get('title') or ''
                authors = [a['author']['display_name'] for a in w.get('authorships',[])[:3] if a.get('author')]
                abstract_inv = w.get('abstract_inverted_index') or {}
                abstract = ''
                if abstract_inv:
                    words = {pos:word for word,positions in abstract_inv.items() for pos in positions}
                    abstract = ' '.join(words[i] for i in sorted(words.keys()))[:500]
                pub_date = w.get('publication_date') or ''
                doi = w.get('doi') or ''
                link = doi if doi.startswith('http') else (f'https://doi.org/{doi}' if doi else wid)
                concepts = [c['display_name'] for c in w.get('concepts',[])[:2]]
                papers.append({'来源':'OpenAlex','标题':title,'作者':', '.join(authors),'摘要':abstract,'发布日期':pub_date[:10],'分类':', '.join(concepts),'链接':link})
        except Exception as e:
            print(f'  ⚠️ OpenAlex ({cid}): {e}')
    print(f'  ✓ OpenAlex: {len(papers)} 篇')
    return papers

def fetch_google_scholar(max_results=15):
    print('📡 尝试 Google Scholar (scholarly)...')
    papers = []
    try:
        try:
            from scholarly import scholarly as sc
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable,'-m','pip','install','scholarly','-q'])
            from scholarly import scholarly as sc
        keywords = ['quantitative finance 2024','asset pricing machine learning 2024']
        seen = set()
        for kw in keywords:
            if len(papers) >= max_results: break
            try:
                results = sc.search_pubs(kw)
                for _ in range(8):
                    try:
                        pub = next(results)
                        bib = pub.get('bib',{})
                        t = bib.get('title','')
                        if t in seen: continue
                        seen.add(t)
                        authors = bib.get('author','')
                        if isinstance(authors,list): authors = ', '.join(authors[:3])
                        papers.append({'来源':'Google Scholar','标题':t,'作者':authors,'摘要':bib.get('abstract','')[:500],'发布日期':str(bib.get('pub_year','')),'分类':kw,'链接':pub.get('pub_url','')})
                    except StopIteration: break
            except Exception as e:
                print(f'  ⚠️ scholarly ({kw}): {e}')
    except Exception as e:
        print(f'  ⚠️ Google Scholar 不可用: {e}')
    print(f'  ✓ Google Scholar: {len(papers)} 篇')
    return papers

def dedup(papers):
    seen = set()
    result = []
    for p in papers:
        key = p['标题'].lower().strip()
        if key and key not in seen:
            seen.add(key)
            result.append(p)
    return result

def write_excel(papers, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '每日论文'
    header_fill = PatternFill('solid', fgColor='1F3864')
    source_colors = {'arXiv':PatternFill('solid',fgColor='DCEEFB'),'Semantic Scholar':PatternFill('solid',fgColor='E8F5E9'),'OpenAlex':PatternFill('solid',fgColor='FFF8E1'),'Google Scholar':PatternFill('solid',fgColor='FCE4EC')}
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    headers = ['#','来源','标题','作者','摘要','发布日期','分类','链接']
    col_widths = [4,16,50,30,80,12,20,50]
    for col,(h,w) in enumerate(zip(headers,col_widths),1):
        cell = ws.cell(row=1,column=col,value=h)
        cell.font = Font(bold=True,color='FFFFFF',size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    for idx,p in enumerate(papers,1):
        row = idx+1
        src = p.get('来源','')
        fill = source_colors.get(src, PatternFill('solid',fgColor='F5F5F5'))
        values = [idx,src,p.get('标题',''),p.get('作者',''),p.get('摘要',''),p.get('发布日期',''),p.get('分类',''),p.get('链接','')]
        for col,val in enumerate(values,1):
            cell = ws.cell(row=row,column=col,value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='top',wrap_text=(col in [3,5]))
        ws.row_dimensions[row].height = 80
    ws2 = wb.create_sheet('来源统计')
    ws2.append(['来源','论文数量'])
    from collections import Counter
    cnt = Counter(p['来源'] for p in papers)
    for src,n in cnt.most_common():
        ws2.append([src,n])
    ws2.append(['合计',len(papers)])
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12
    wb.save(output_path)
    print(f'\n✅ 已保存: {output_path}')
    print(f'   共 {len(papers)} 篇论文（去重后）')

def main():
    desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
    os.makedirs(desktop, exist_ok=True)
    output_path = os.path.join(desktop, f'{TODAY}_finance_papers.xlsx')
    print(f'🚀 金融论文日报 {TODAY}')
    print('='*50)
    all_papers = []
    all_papers += fetch_arxiv(30)
    all_papers += fetch_semantic_scholar(20)
    all_papers += fetch_openalex(20)
    all_papers += fetch_google_scholar(15)
    papers = dedup(all_papers)
    papers.sort(key=lambda p: p.get('发布日期',''), reverse=True)
    print(f'\n📊 汇总: 共 {len(papers)} 篇（去重后）')
    write_excel(papers, output_path)
    return output_path

if __name__ == '__main__':
    main()
