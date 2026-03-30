#!/usr/bin/env python3
"""
发票 OCR 识别脚本
使用阿里云 OCR API 识别发票并提取信息
支持混贴发票识别，自动识别 17+ 种发票类型
输出 Excel xlsx 格式
"""

import os
import sys
import json
import base64
import hashlib
import hmac
import urllib.request
import urllib.parse
import time
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 支持的文件格式
SUPPORTED_FORMATS = {'.pdf', '.ofd', '.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp'}

# 表头（按用户要求的顺序）
HEADERS = [
    "发票号码",
    "开票日期",
    "购买方信息",
    "销售方信息",
    "项目名称",
    "规格型号",
    "单位",
    "数量",
    "单价",
    "金额",
    "税率",
    "税额",
    "价税合计",
]


def load_config():
    """加载配置文件"""
    config_path = Path(__file__).parent.parent / "config.json"
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_config(config: dict):
    """保存配置文件"""
    config_path = Path(__file__).parent.parent / "config.json"
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)


def percent_encode(s: str) -> str:
    """URL 编码（阿里云规范）"""
    result = urllib.parse.quote(str(s), safe='')
    return result.replace('+', '%20').replace('*', '%2A').replace('%7E', '~')


def create_signature(params: dict, secret: str, method: str = 'POST') -> str:
    """
    创建阿里云 API 签名
    
    Args:
        params: 请求参数字典
        secret: AccessKey Secret
        method: HTTP 方法（必须与实际请求方法一致）
    
    Returns:
        签名字符串
    
    阿里云签名规范:
    - StringToSign = HTTPMethod + "&" + percent_encode("/") + "&" + percent_encode(CanonicalizedQueryString)
    - 签名方法必须与实际 HTTP 请求方法一致
    """
    # 1. 参数按字母序排序
    sorted_params = sorted(params.items())
    
    # 2. 构造规范化查询字符串
    canonicalized_query = '&'.join([f"{percent_encode(k)}={percent_encode(v)}" for k, v in sorted_params])
    
    # 3. 构造待签名字符串（关键：方法必须与实际请求一致）
    string_to_sign = f'{method}&%2F&' + percent_encode(canonicalized_query)
    
    # 4. HMAC-SHA1 签名
    signature = hmac.new(
        (secret + '&').encode('utf-8'),
        string_to_sign.encode('utf-8'),
        hashlib.sha1
    ).digest()
    
    # 5. Base64 编码
    return base64.b64encode(signature).decode('utf-8')


def call_aliyun_ocr(image_data: bytes, access_key_id: str, access_key_secret: str, 
                    api_action: str = "RecognizeMixedInvoices") -> dict:
    """调用阿里云 OCR API"""
    endpoint = "https://ocr-api.cn-hangzhou.aliyuncs.com"
    
    timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    nonce = str(int(time.time() * 1000000))
    
    params = {
        "AccessKeyId": access_key_id,
        "Action": api_action,
        "Format": "JSON",
        "SignatureMethod": "HMAC-SHA1",
        "SignatureNonce": nonce,
        "SignatureVersion": "1.0",
        "Timestamp": timestamp,
        "Version": "2021-07-07",
    }
    
    signature = create_signature(params, access_key_secret)
    params["Signature"] = signature
    
    query_string = '&'.join([f"{urllib.parse.quote(k)}={urllib.parse.quote(str(v))}" for k, v in params.items()])
    url = f"{endpoint}?{query_string}"
    
    try:
        req = urllib.request.Request(url, data=image_data, method='POST')
        req.add_header('Content-Type', 'application/octet-stream')
        with urllib.request.urlopen(req, timeout=60) as response:
            result = json.loads(response.read().decode('utf-8'))
            return result
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8')
        return {"Code": "HttpError", "Message": f"HTTP {e.code}: {error_body}"}
    except Exception as e:
        return {"Code": "Error", "Message": str(e)}


def parse_invoice_data(data: dict) -> list:
    """解析发票数据，返回发票信息列表"""
    results = []
    
    # 混贴发票识别返回 subMsgs
    if "subMsgs" in data:
        for sub_msg in data.get("subMsgs", []):
            invoice_data = sub_msg.get("result", {}).get("data", {})
            invoice_info = extract_invoice_fields(invoice_data)
            results.append(invoice_info)
    else:
        # 单张发票识别
        invoice_data = data.get("data", data)
        invoice_info = extract_invoice_fields(invoice_data)
        results.append(invoice_info)
    
    return results


def extract_invoice_fields(invoice: dict) -> dict:
    """从发票数据中提取字段，按用户要求的字段名"""
    # 默认值为空字符串
    result = {k: "" for k in HEADERS}
    
    # 发票号码 = 发票代码 + 发票号码
    invoice_code = invoice.get("invoiceCode", "")
    invoice_number = invoice.get("invoiceNumber", "")
    if invoice_code and invoice_number:
        result["发票号码"] = f"{invoice_code} {invoice_number}"
    elif invoice_number:
        result["发票号码"] = invoice_number
    elif invoice_code:
        result["发票号码"] = invoice_code
    
    result["开票日期"] = invoice.get("invoiceDate", "")
    result["购买方信息"] = invoice.get("purchaserName", "")
    result["销售方信息"] = invoice.get("sellerName", "")
    result["价税合计"] = invoice.get("totalAmount", "")
    
    # 金额信息
    result["金额"] = invoice.get("invoiceAmountPreTax", "")
    result["税额"] = invoice.get("invoiceTax", "")
    
    # 提取明细（取第一条）
    details = invoice.get("invoiceDetails", [])
    if details:
        # 合并所有明细的项目名称
        item_names = [d.get("itemName", "") for d in details if d.get("itemName")]
        result["项目名称"] = "; ".join(item_names)
        
        # 第一个明细的详细信息
        first_detail = details[0]
        result["规格型号"] = first_detail.get("specification", "")
        result["单位"] = first_detail.get("unit", "")
        result["数量"] = str(first_detail.get("quantity", ""))
        result["单价"] = first_detail.get("unitPrice", "")
        result["金额"] = first_detail.get("amount", result["金额"])
        result["税率"] = first_detail.get("taxRate", "")
        result["税额"] = first_detail.get("tax", result["税额"])
    
    return result


def read_file_bytes(file_path: str) -> bytes:
    """读取文件二进制数据"""
    with open(file_path, 'rb') as f:
        return f.read()


def scan_folder(folder_path: str) -> list:
    """扫描文件夹中的所有发票文件"""
    folder = Path(folder_path)
    if not folder.exists():
        print(f"错误：文件夹不存在: {folder_path}")
        return []
    
    files = []
    for file in folder.rglob('*'):
        if file.is_file() and file.suffix.lower() in SUPPORTED_FORMATS:
            files.append(str(file))
    
    return files


def recognize_invoice(file_path: str, access_key_id: str, access_key_secret: str) -> list:
    """识别单个发票文件"""
    print(f"正在识别: {os.path.basename(file_path)}")
    
    try:
        image_data = read_file_bytes(file_path)
        
        # 使用混贴发票识别 API
        ocr_result = call_aliyun_ocr(image_data, access_key_id, access_key_secret, "RecognizeMixedInvoices")
        
        if "Code" in ocr_result and ocr_result["Code"] != "Success":
            # 尝试单发票识别
            ocr_result = call_aliyun_ocr(image_data, access_key_id, access_key_secret, "RecognizeInvoice")
            
            if "Code" in ocr_result and ocr_result["Code"] != "Success":
                error_msg = ocr_result.get("Message", "识别失败")
                return [{"发票号码": "识别失败", "开票日期": error_msg}]
        
        # 解析 Data 字段
        data = ocr_result.get("Data", {})
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                return [{"发票号码": "解析失败", "开票日期": "响应数据解析失败"}]
        
        return parse_invoice_data(data)
        
    except Exception as e:
        return [{"发票号码": "异常", "开票日期": str(e)}]


def export_to_xlsx(results: list, output_path: str):
    """导出到 Excel xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "发票信息"
    
    # 样式定义
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, invoice in enumerate(results, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            value = invoice.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # 自动调整列宽
    for col_idx, header in enumerate(HEADERS, 1):
        max_length = len(header) * 2  # 中文字符宽度
        for row in ws.iter_rows(min_row=2, max_row=len(results) + 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # 中文字符算2个宽度
                    char_count = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_length = max(max_length, char_count)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 保存文件
    wb.save(output_path)
    print(f"\n✓ Excel 文件已保存: {output_path}")
    print(f"  共 {len(results)} 条记录")


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("发票 OCR 识别工具")
        print("")
        print("用法:")
        print("  python recognize_invoices.py <发票文件夹路径> [--output <输出文件>]")
        print("  python recognize_invoices.py --config")
        print("  python recognize_invoices.py --list-config")
        print("")
        print("支持的发票类型:")
        print("  增值税发票、火车票、出租车票、机票行程单、定额发票、")
        print("  机动车销售发票、网约车行程单、过路过桥费发票等 17+ 种")
        print("")
        print("支持的文件格式:")
        print("  PDF, OFD, JPG, PNG, BMP, GIF, TIFF, WebP")
        sys.exit(1)
    
    # 处理配置命令
    if sys.argv[1] == "--config":
        print("请输入阿里云配置信息:")
        access_key_id = input("AccessKey ID: ").strip()
        access_key_secret = input("AccessKey Secret: ").strip()
        config = {
            "aliyun_access_key_id": access_key_id,
            "aliyun_access_key_secret": access_key_secret
        }
        save_config(config)
        print("✓ 配置已保存!")
        return
    
    if sys.argv[1] == "--list-config":
        config = load_config()
        if config:
            print("当前配置:")
            print(f"  AccessKey ID: {config.get('aliyun_access_key_id', '(未设置)')}")
            print(f"  AccessKey Secret: {'*' * 8 if config.get('aliyun_access_key_secret') else '(未设置)'}")
        else:
            print("尚未配置，请运行: python recognize_invoices.py --config")
        return
    
    # 加载配置
    config = load_config()
    access_key_id = config.get("aliyun_access_key_id")
    access_key_secret = config.get("aliyun_access_key_secret")
    
    if not access_key_id or not access_key_secret:
        print("❌ 错误：未配置阿里云凭证")
        print("")
        print("请先运行以下命令配置凭证:")
        print("  python recognize_invoices.py --config")
        sys.exit(1)
    
    # 解析参数
    folder_path = sys.argv[1]
    output_file = None
    
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_file = sys.argv[idx + 1]
    
    # 扫描文件夹
    files = scan_folder(folder_path)
    if not files:
        print("❌ 未找到发票文件")
        print(f"   支持的格式: {', '.join(sorted(SUPPORTED_FORMATS))}")
        return
    
    print(f"📁 找到 {len(files)} 个发票文件\n")
    
    # 识别所有发票
    results = []
    for i, file_path in enumerate(files, 1):
        print(f"[{i}/{len(files)}] ", end="")
        invoices = recognize_invoice(file_path, access_key_id, access_key_secret)
        
        for inv in invoices:
            invoice_number = inv.get("发票号码", "")
            if "失败" in invoice_number or "异常" in invoice_number:
                print(f"  ❌ 失败: {inv.get('开票日期', '未知错误')}")
            else:
                print(f"  ✓ 发票号码: {invoice_number}")
        
        results.extend(invoices)
    
    # 输出结果
    if not output_file:
        output_file = "发票汇总.xlsx"
    elif not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    export_to_xlsx(results, output_file)


if __name__ == "__main__":
    main()