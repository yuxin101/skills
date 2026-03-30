#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
艾威板材费用计算器 - 无模板独立生成版
从激光切割 PDF 提取 Sheet dimension 数据，生成艾威格式 Excel
"""

import os
import re
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


class SheetMaterialCalculator:
    """板材费用计算器（无模板独立生成）"""
    
    def __init__(self, output_dir: str = None):
        """
        初始化计算器
        
        Args:
            output_dir: 输出目录，默认为输入文件所在目录
        """
        self.output_dir = output_dir
        
        # 厚度映射（PDF 厚度 → 模板厚度，负公差）
        self.thickness_map = {
            1.5: 1.5,
            2: 2,
            3: 2.75,
            4: 3.75,
            5: 4.75,
            6: 5.75,
            8: 7.75,
            10: 9.75,
            12: 11.75,
        }
    
    def parse_pdf_filename(self, filename: str) -> tuple:
        """
        从 PDF 文件名提取厚度和材质
        
        Args:
            filename: PDF 文件名
            
        Returns:
            (厚度，材质) 元组
        """
        name = filename.replace('.pdf', '')
        
        # 处理前缀（如 dxb-, sus-, nm- 等）
        if name.lower().startswith('sus-'):
            material = 'SUS'
            thickness_str = name[4:]  # 去掉 'sus-'
        elif name.lower().startswith('dxb-'):
            material = '镀锌板'
            thickness_str = name[4:]  # 去掉 'dxb-'
        elif name.lower().startswith('nm-') or name.lower().startswith('nm'):
            material = '耐磨板'
            thickness_str = re.sub(r'^[Nn][Mm]-?', '', name)
        elif 'Mn' in name:
            material = 'Q345'
            thickness_str = name.replace('Mn', '')
        elif 'SUS' in name:
            material = 'SUS'
            thickness_str = name.replace('SUS', '')
        else:
            material = 'Q235'
            thickness_str = name
        
        # 去掉 mm 后缀再提取数字
        thickness_str = thickness_str.replace('mm', '')
        thickness_num = re.match(r'([\d.]+)', thickness_str)
        thickness = float(thickness_num.group(1)) if thickness_num else float(thickness_str)
        
        return thickness, material
    
    def get_template_material(self, thickness: float, material: str) -> str:
        """
        获取模板显示的材质
        
        Args:
            thickness: 厚度
            material: 材质
            
        Returns:
            模板显示的材质
        """
        if material == 'Q235' and thickness <= 2:
            return 'Q235 冷板'
        return material
    
    def extract_pdf_data(self, pdf_path: str) -> dict:
        """
        从 PDF 提取 Sheet dimension 数据
        
        提取逻辑：
        - Material data 只在第一页提取（用于验证，不填充表格）
        - 每个 Plan data 块下有 Sheet dimension 就提取
        - 不去重（即使数据相同也保留）
        - 提取完后验证板材使用量（对比 Material data）
        
        Args:
            pdf_path: PDF 文件路径
            
        Returns:
            包含板材数据的字典
        """
        import PyPDF2
        
        reader = PyPDF2.PdfReader(pdf_path)
        
        data = {
            'thickness': None,
            'material': None,
            'sheets': [],
            'material_used': {}  # Material data 中的使用量（仅用于验证）
        }
        
        # 从文件名获取厚度和材质
        thickness, material = self.parse_pdf_filename(pdf_path.name)
        data['thickness'] = thickness
        data['material'] = material
        
        # 1. 只从第一页提取 Material data（用于验证）
        if len(reader.pages) > 0:
            page = reader.pages[0]
            text = page.extract_text()
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            in_material_data = False
            for i, line in enumerate(lines):
                if 'Material data' in line:
                    in_material_data = True
                    continue
                
                # 遇到 Plan data 就停止（Material data 只在第一页，且到 Plan data 为止）
                if in_material_data and 'Plan data' in line:
                    break
                
                if in_material_data:
                    # 查找 Dimension 和 Used
                    dim_match = re.search(r'(\d{3,4})\s*x\s*(\d{3,4})\s*mm', line)
                    if dim_match:
                        dim_key = f"{dim_match.group(1)}x{dim_match.group(2)}"
                        # 向后查找 Used
                        for j in range(i+1, min(i+5, len(lines))):
                            used_match = re.search(r'(\d+)\s*/\s*(\d+)', lines[j])
                            if used_match:
                                used = int(used_match.group(1))
                                data['material_used'][dim_key] = used
                                break
        
        # 2. 从所有页提取 Sheet dimension 数据
        for page_num, page in enumerate(reader.pages):
            text = page.extract_text()
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            header_idx = -1
            for i, line in enumerate(lines):
                if line == 'Sheet dimension':
                    header_idx = i
                    break
            
            if header_idx == -1:
                continue
            
            # 从表头下一行开始提取
            row_idx = header_idx + 1
            
            while row_idx < len(lines):
                line = lines[row_idx]
                
                # 遇到 Flat part data 就停止
                if 'Flat part data' in line:
                    break
                
                # 查找尺寸数据
                match = re.search(r'(\d{3,4})\s*x\s*(\d{3,4})\s*mm', line)
                
                if match:
                    # 检查前面是否有另一个尺寸（Plan dimension）
                    prev_lines = ' '.join(lines[max(header_idx, row_idx-3):row_idx])
                    has_plan_dim = bool(re.search(r'\d{3,4}\s*x\s*\d{3,4}\s*mm', prev_lines))
                    
                    # 只提取 Sheet dimension（前面有 Plan dimension 的第二个尺寸）
                    if has_plan_dim:
                        length = int(match.group(1))
                        width = int(match.group(2))
                        
                        # 向后查找 Cycles 和 Waste
                        cycles = None
                        waste = None
                        
                        for j in range(row_idx+1, min(row_idx+10, len(lines))):
                            next_line = lines[j]
                            # 跳过尺寸行
                            if re.match(r'\d{3,4}\s*x\s*\d{3,4}\s*mm', next_line):
                                continue
                            # 查找 Cycles（纯数字）
                            if cycles is None and re.match(r'^\d+$', next_line):
                                cycles = int(next_line)
                            # 查找 Waste（百分比）
                            elif re.match(r'^[\d.]+%$', next_line):
                                waste = float(next_line.replace('%', ''))
                                break
                        
                        # 保存完整数据（不去重）
                        if cycles is not None and waste is not None:
                            data['sheets'].append({
                                'length': length,
                                'width': width,
                                'cycles': cycles,
                                'waste_rate': waste
                            })
                
                row_idx += 1
        
        # 3. 验证提取的板材数量是否和 Material data 一致
        self._verify_material_usage(data, pdf_path.name)
        
        return data
    
    def _verify_material_usage(self, data: dict, filename: str = None):
        """
        核对提取的板材和 Material data 中已使用的板材
        
        1. Material data 里 Used > 0 的规格，是否都提取到了
        2. 提取到的规格，数量是否和 Used 一致
        没使用的（Used=0）不管
        
        Args:
            data: 包含板材数据的字典
            filename: PDF 文件名（用于打印提示）
        """
        if not data['material_used']:
            return
        
        # 统计提取的每种规格的板材数量
        extracted_usage = {}
        for sheet in data['sheets']:
            dim_key = f"{sheet['length']}x{sheet['width']}"
            extracted_usage[dim_key] = extracted_usage.get(dim_key, 0) + sheet['cycles']
        
        issues = []
        
        # 1. 检查 Material data 里 Used > 0 的是否都提取到了，数量是否一致
        for dim_key, used in data['material_used'].items():
            if used <= 0:
                continue  # 没使用的不管
            extracted = extracted_usage.get(dim_key, 0)
            if extracted == 0:
                issues.append(f"❌ {dim_key}: Used={used} 但未提取到")
            elif extracted != used:
                issues.append(f"⚠️  {dim_key}: 提取{extracted} vs Used{used}")
        
        # 打印核对结果
        total_extracted = sum(extracted_usage.values())
        
        if issues:
            print(f"📋 {filename or 'PDF'}: 提取 {total_extracted} 张，核对发现问题：")
            for issue in issues:
                print(f"   {issue}")
        else:
            print(f"✅ {filename or 'PDF'}: 核对通过 - {total_extracted} 张板材")
    
    def process_folder(self, folder_path: str) -> dict:
        """
        处理文件夹中的所有 PDF 文件
        
        Args:
            folder_path: 文件夹路径（字符串或 Path 对象）
            
        Returns:
            包含所有板材数据的字典
        """
        from pathlib import Path
        
        all_data = {'sheets': {}}
        folder = Path(folder_path)
        
        pdf_files = [f for f in folder.iterdir() if f.suffix == '.pdf']
        
        for pdf_file in sorted(pdf_files, key=lambda x: x.name):
            data = self.extract_pdf_data(pdf_file)
            
            key = f"{data['thickness']}_{data['material']}"
            if key not in all_data['sheets']:
                all_data['sheets'][key] = {'sheets': []}
            
            all_data['sheets'][key]['sheets'].extend(data['sheets'])
        
        return all_data
    
    def load_price_table(self, price_table_path: str, month: str = None) -> dict:
        """
        加载价格表，提取当月板材采购单价
        
        价格表格式（震源板材价格）：
        - 工作表名：'26年X月份板材价格' (带可能的空格)
        - F 列（6）：厚度（mm）
        - C 列（3）：材质名称（"开平板 Q235" / "冷板 ST12" / "锰板 Q345"）
        - 采购价格列：位置不固定，通过表头匹配 "X月采购价格" 或 "X月\n采购价格"
        
        Args:
            price_table_path: 价格表文件路径（Excel）
            month: 月份数字（如 3 表示 3 月），默认使用当前月份
            
        Returns:
            单价字典 {(厚度, 材质): 单价（元/吨）}
            材质分类：'Q235'（热轧开平板）、'Q235冷板'（冷板ST12）、'Q345'、'SUS'
        """
        from datetime import datetime
        from openpyxl import load_workbook
        
        if month is None:
            month = datetime.now().month
        else:
            month = int(month)
        
        price_dict = {}
        
        # 加载价格表（data_only=True 读取公式计算结果）
        wb = load_workbook(price_table_path, data_only=True)
        
        # 查找月份工作表（模糊匹配，忽略空格）
        ws = None
        for name in wb.sheetnames:
            name_clean = name.replace(' ', '')
            if f'{month}月' in name_clean and '板材价格' in name_clean:
                ws = wb[name]
                print(f"✅ 找到工作表：'{name}'")
                break
        
        if ws is None:
            print(f"⚠️  未找到{month}月份价格表")
            return price_dict
        
        # 动态查找采购价格列（表头包含 "X月" + "采购价格"）
        price_col = None
        target_month_strs = [f'{month}月', f'{month} 月']
        
        for col in range(1, ws.max_column + 1):
            # 合并第1行和第2行的表头文本
            h1 = str(ws.cell(row=1, column=col).value or '')
            h2 = str(ws.cell(row=2, column=col).value or '')
            header = (h1 + h2).replace('\n', '').replace(' ', '')
            
            # 匹配 "X月采购价格" 或 "X月\n采购价格"
            has_month = any(ms.replace(' ', '') in header for ms in target_month_strs)
            has_price = '采购价格' in header
            
            if has_month and has_price:
                price_col = col
                display = (h1 + h2).replace('\n', ' ')[:30]
                print(f"✅ 找到{month}月采购价格列：第{col}列 ('{display}')")
                break
        
        if price_col is None:
            print(f"⚠️  未找到{month}月采购价格列")
            # 尝试找最后一个带"采购价格"的列
            for col in range(ws.max_column, 0, -1):
                h1 = str(ws.cell(row=1, column=col).value or '')
                h2 = str(ws.cell(row=2, column=col).value or '')
                header = (h1 + h2).replace('\n', '').replace(' ', '')
                if '采购价格' in header:
                    price_col = col
                    print(f"⚠️  回退使用最后一个采购价格列：第{col}列")
                    break
        
        if price_col is None:
            print(f"❌ 完全找不到采购价格列")
            return price_dict
        
        # 提取价格数据（从第3行开始，第2行是表头）
        for row in range(3, ws.max_row + 1):
            thickness = ws.cell(row=row, column=6).value   # F列 - 厚度
            material_raw = str(ws.cell(row=row, column=3).value or '').strip()  # C列 - 材质名称
            price_val = ws.cell(row=row, column=price_col).value  # 采购价格
            
            if thickness is None or price_val is None:
                continue
            
            try:
                thickness = float(thickness)
                price_per_ton = float(price_val)
            except (ValueError, TypeError):
                continue
            
            # 分类材质
            if '冷板' in material_raw or 'ST12' in material_raw:
                material = 'Q235冷板'
            elif 'Q345' in material_raw or '锰板' in material_raw:
                material = 'Q345'
            elif 'SUS' in material_raw or '不锈钢' in material_raw:
                material = 'SUS'
            else:
                material = 'Q235'  # 开平板 Q235 等
            
            key = (thickness, material)
            price_dict[key] = price_per_ton
            
        print(f"✅ 已加载 {len(price_dict)} 条价格（{month}月采购价）")
        for (t, m), p in sorted(price_dict.items(), key=lambda x: (x[0][1], x[0][0])):
            print(f"   {m} {t}mm → {p} 元/吨")
        
        return price_dict
    
    def _lookup_price(self, price_dict: dict, raw_thickness: float, mapped_thickness: float, material: str) -> float:
        """
        智能匹配价格：先试映射厚度，再试原始厚度
        
        Args:
            price_dict: 价格字典 {(厚度, 材质): 元/吨}
            raw_thickness: PDF 文件名中的原始厚度（如 3, 5, 8）
            mapped_thickness: 负公差映射后的厚度（如 2.75, 4.75, 7.75）
            material: PDF 材质（'Q235', 'Q345', 'SUS'）
            
        Returns:
            单价（元/吨），未找到返回 None
        """
        # 确定价格表中的材质分类
        if material == 'Q235' and raw_thickness <= 2:
            price_material = 'Q235冷板'
        else:
            price_material = material
        
        # 1. 先试映射后的厚度
        price = price_dict.get((mapped_thickness, price_material))
        if price is not None:
            return price
        
        # 2. 再试原始厚度
        price = price_dict.get((raw_thickness, price_material))
        if price is not None:
            return price
        
        # 3. 冷板没匹配到，试 Q235 开平板
        if price_material == 'Q235冷板':
            price = price_dict.get((mapped_thickness, 'Q235'))
            if price is not None:
                return price
            price = price_dict.get((raw_thickness, 'Q235'))
            if price is not None:
                return price
        
        return None
    
    def generate_excel(self, all_data: dict, output_path: str, waste_price: float = None, price_table_path: str = None, month: str = None):
        """
        生成 Excel 文件（无模板独立生成）
        
        Args:
            all_data: 包含所有板材数据的字典
            output_path: 输出文件路径
            waste_price: 废料单价（元/Kg），可选
            price_table_path: 价格表文件路径，可选
            month: 月份（格式：YYYY-MM），默认使用当前月份
        """
        # 加载价格表
        price_dict = {}
        if price_table_path:
            price_dict = self.load_price_table(price_table_path, month)
        wb = Workbook()
        ws = wb.active
        ws.title = '震源'
        
        # 设置样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头
        headers = [
            '序号', '日期', '订单号', '材质（Q235/Q345/SUS）',
            '长（mm）', '宽（mm）', '厚（mm）', '数量',
            '重量    （单重 Kg）', '合计重量（Kg）', '板材利用率',
            '废料重量（Kg）', '单价（Kg）', '板材价格', '废料率%'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # 排序规则：
        # Q235冷板/Q235/Q345 → 先按厚度从小到大，同厚度按材质排序（常用材质）
        # 其他（SUS/镀锌板/耐磨板等）→ 排在后面，各自按材质分类再按厚度
        main_materials = {'Q235', 'Q345'}  # Q235 含冷板
        
        main_keys = [k for k in all_data['sheets'].keys() if k.split('_')[1] in main_materials]
        other_keys = [k for k in all_data['sheets'].keys() if k.split('_')[1] not in main_materials]
        
        # 常用材质：厚度优先，同厚度按材质字母排序
        main_keys.sort(key=lambda x: (float(x.split('_')[0]), x.split('_')[1]))
        # 其他材质：按材质分类，再按厚度
        other_order = {'SUS': 0, '镀锌板': 1, '耐磨板': 2}
        other_keys.sort(key=lambda x: (other_order.get(x.split('_')[1], 99), float(x.split('_')[0])))
        
        sorted_keys = main_keys + other_keys
        
        # 填充数据
        row_idx = 2
        for key in sorted_keys:
            sheet_info = all_data['sheets'][key]
            parts = key.split('_')
            thickness = float(parts[0])
            material = parts[1] if len(parts) > 1 else 'Q235'
            template_thickness = self.thickness_map.get(thickness, thickness)
            template_material = self.get_template_material(thickness, material)
            
            # 从价格表获取单价（元/吨）
            unit_price_ton = self._lookup_price(price_dict, thickness, template_thickness, material) if price_dict else None
            if unit_price_ton is not None:
                unit_price_kg = round(unit_price_ton / 1000, 3)  # 转换为元/Kg
            else:
                unit_price_kg = None
            
            for sheet in sheet_info['sheets']:
                ws.cell(row=row_idx, column=1, value=row_idx - 1)  # A 列 - 序号
                ws.cell(row=row_idx, column=4, value=template_material)  # D 列 - 材质
                ws.cell(row=row_idx, column=5, value=sheet['length'])  # E 列 - 长
                ws.cell(row=row_idx, column=6, value=sheet['width'])  # F 列 - 宽
                ws.cell(row=row_idx, column=7, value=template_thickness)  # G 列 - 厚
                ws.cell(row=row_idx, column=8, value=sheet['cycles'])  # H 列 - 数量
                
                # M 列 - 单价（元/Kg，从价格表获取）
                if unit_price_kg is not None:
                    ws.cell(row=row_idx, column=13, value=unit_price_kg)
                # 否则 M 列留空
                
                # I 列 - 重量（不锈钢密度 7.95，其他 7.85）
                density = 7.95 if material == 'SUS' else 7.85
                ws.cell(row=row_idx, column=9, value=f'=E{row_idx}*F{row_idx}*G{row_idx}/1000000*{density}')  # I 列
                ws.cell(row=row_idx, column=10, value=f'=I{row_idx}*H{row_idx}')  # J 列
                ws.cell(row=row_idx, column=11, value=f'=1-O{row_idx}')  # K 列
                ws.cell(row=row_idx, column=12, value=f'=J{row_idx}*(1-K{row_idx})*0.85')  # L 列
                ws.cell(row=row_idx, column=14, value=f'=J{row_idx}*M{row_idx}')  # N 列
                ws.cell(row=row_idx, column=15, value=sheet['waste_rate']/100)  # O 列 - 废料率（小数格式）
                
                # 设置单元格样式
                for col in range(1, 16):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                row_idx += 1
        
        last_data_row = row_idx - 1
        total_row = last_data_row + 1
        
        # 合计行
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=11)
        ws.cell(row=total_row, column=1, value='合计')
        ws.cell(row=total_row, column=12, value=f'=SUM(L2:L{last_data_row})')
        ws.cell(row=total_row, column=14, value=f'=SUM(N2:N{last_data_row})')
        
        for col in range(1, 16):
            cell = ws.cell(row=total_row, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.cell(row=total_row, column=12).number_format = '0.00'
        ws.cell(row=total_row, column=14).number_format = '0.00'
        
        # 废料单价行（L 列）
        ws.merge_cells(start_row=total_row+1, start_column=1, end_row=total_row+1, end_column=11)
        ws.cell(row=total_row+1, column=1, value='废料单价')
        if waste_price is not None:
            ws.cell(row=total_row+1, column=12, value=waste_price)
        
        for col in range(1, 16):
            cell = ws.cell(row=total_row+1, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.cell(row=total_row+1, column=12).number_format = '0.00'
        
        # 废料价格行（L 列）
        ws.merge_cells(start_row=total_row+2, start_column=1, end_row=total_row+2, end_column=11)
        ws.cell(row=total_row+2, column=1, value='废料价格')
        ws.cell(row=total_row+2, column=12, value=f'=L{total_row}*L{total_row+1}')
        
        for col in range(1, 16):
            cell = ws.cell(row=total_row+2, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.cell(row=total_row+2, column=12).number_format = '0.00'
        
        # 设置百分比格式
        for row in range(2, last_data_row + 1):
            ws.cell(row=row, column=11).number_format = '0.00%'  # K 列 - 利用率
            ws.cell(row=row, column=15).number_format = '0.00%'  # O 列 - 废料率
            ws.cell(row=row, column=9).number_format = '0.00'    # I 列 - 重量
            ws.cell(row=row, column=10).number_format = '0.00'   # J 列 - 合计重量
            ws.cell(row=row, column=12).number_format = '0.00'   # L 列 - 废料重量
            ws.cell(row=row, column=14).number_format = '0.00'   # N 列 - 板材价格
        
        # 设置列宽（参考模板）
        column_widths = {
            'A': 9, 'B': 11.5, 'C': 18, 'D': 12,
            'E': 10.5, 'F': 13, 'G': 13, 'H': 11.5,
            'I': 10.5, 'J': 10.5, 'K': 10.5, 'L': 10.5,
            'M': 10.5, 'N': 10.5, 'O': 12.5
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置第 1 行行高
        ws.row_dimensions[1].height = 54
        
        # 保存
        wb.save(output_path)


# 主程序
if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("用法：python3 calculator.py <PDF 文件夹> [输出文件] [废料单价] [价格表] [月份]")
        print("示例：python3 calculator.py /tmp/060_processing /tmp/060_艾威板材统计.xlsx 0.02 /tmp/价格表.xlsx 2026-03")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    waste_price = float(sys.argv[3]) / 100 if len(sys.argv) > 3 else None  # 转换为小数
    price_table_path = sys.argv[4] if len(sys.argv) > 4 else None
    month = sys.argv[5] if len(sys.argv) > 5 else None
    
    if not output_path:
        # 默认命名：{压缩包名/文件夹名}_Board material cost statistics.xlsx
        folder_name = os.path.basename(folder_path.rstrip('/'))
        output_path = os.path.join(os.path.dirname(folder_path.rstrip('/')), f'{folder_name}_Board material cost statistics.xlsx')
    
    calculator = SheetMaterialCalculator()
    all_data = calculator.process_folder(folder_path)
    calculator.generate_excel(all_data, output_path, waste_price, price_table_path, month)
    
    print(f"✅ 已生成：{output_path}")
    total_sheets = sum(len(v['sheets']) for v in all_data['sheets'].values())
    print(f"   共 {total_sheets} 张板材")
