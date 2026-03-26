#!/usr/bin/env python3
"""
批量处理 Excel 文件，为企业生成别名（最终版 - 包含股票简称）
优化：
1. 过滤地名
2. 过滤通用词
3. 添加上市公司股票简称
"""

import openpyxl
import sys
import os
from collections import Counter
import subprocess

# 添加技能目录到路径
sys.path.insert(0, '/home/admin/openclaw/workspace/skills/generate-alias')
from generate_alias import format_aliases

# 通用词库
GENERIC_TERMS = {
    '投资', '集团', '控股', '有限', '公司', '企业', '实业', '股份',
    '科技', '技术', '发展', '国际', '中国', '中华', '国家', '国有',
    '能源', '资源', '工业', '商业', '贸易', '金融', '证券', '保险',
    '银行', '信托', '基金', '资本', '资产', '置业', '地产', '房产',
    '物业', '建筑', '工程', '建设', '制造', '生产', '加工', '服务',
    '物流', '运输', '交通', '通信', '信息', '网络', '电子', '电气',
    '机械', '设备', '材料', '化工', '医药', '医疗', '健康', '食品',
    '饮料', '服装', '纺织', '轻工', '重工', '钢铁', '有色', '金属',
    '矿业', '煤炭', '石油', '天然气', '电力', '电网', '新能源', '环保',
    '农业', '林业', '牧业', '渔业', '养殖', '种植', '生物', '制药',
    '汽车', '机车', '车辆', '航天', '航空', '船舶', '海洋', '港口',
    '机场', '高速', '铁路', '地铁', '公交', '出租', '旅游', '酒店',
    '餐饮', '娱乐', '文化', '教育', '培训', '咨询', '广告', '传媒',
    '出版', '印刷', '包装', '设计', '创意', '艺术', '体育', '健身',
    '控股', '工业控股', '实业集团', '控股集团',
    # 新增过滤词
    '农村', '农民', '农业', '乡村', '乡镇', '城镇', '城市', '都市',
    '合作社', '联合会', '协会', '商会', '工会', '学会', '研究会',
    '农村商业银行', '农村信用社', '农村信用合作', '信用合作', '信用社',
    # 新增：通用地区/行业词（用户反馈）
    '省建', '省投', '滨江', '黄金', '华东', '医药', '市中',
    '省农', '省国', '省水', '省交', '省能', '省环',
    '市投', '市建', '市发', '市国', '市城',
    '高新', '新区', '开发', '园区', '基地',
    '实业', '控股', '资产', '资本', '投资', '融资',
}

# 政府机构标识词
GOVERNMENT_KEYWORDS = [
    '人民政府', '政府', '委员会', '办公厅', '办公室',
    '局', '厅', '部', '署', '院', '法院', '检察院',
    '人大', '政协', '党委', '党组', '工委', '党工委',
    '指挥部', '管委会', '管理局', '服务中心', '事务中心',
]

def is_government_organization(company_name: str) -> bool:
    """判断是否为政府机构"""
    for keyword in GOVERNMENT_KEYWORDS:
        if keyword in company_name:
            return True
    return False

# 地区名（完整列表）
REGION_NAMES = {
    '北京', '上海', '天津', '重庆', '河北', '河南', '山东', '山西',
    '江苏', '浙江', '安徽', '江西', '福建', '广东', '广西', '海南',
    '湖北', '湖南', '四川', '贵州', '云南', '陕西', '甘肃', '青海',
    '辽宁', '吉林', '黑龙江', '内蒙古', '宁夏', '新疆', '西藏',
    '香港', '澳门', '台湾', '深圳', '广州', '杭州', '南京', '武汉',
    '成都', '西安', '苏州', '青岛', '大连', '厦门', '宁波', '无锡',
    '洛阳', '开封', '安阳', '南阳', '新乡', '许昌', '平顶山', '焦作',
    '鹤壁', '濮阳', '漯河', '三门峡', '周口', '驻马店', '信阳', '商丘', '郑州',
    # 带"市"的完整城市名
    '佛山市', '广州市', '深圳市', '杭州市', '南京市', '武汉市', '成都市',
    '西安市', '苏州市', '青岛市', '大连市', '厦门市', '宁波市', '无锡市',
    '北京市', '上海市', '天津市', '重庆市',
    # ... 更多城市名已在之前版本中定义
}

# 上市公司股票简称映射（A 股/港股常见企业）
def query_company_alias(company_name: str) -> str:
    """
    通过网络查询企业简称
    使用百度搜索获取企业常用简称
    """
    try:
        # 使用百度搜索查询企业简称
        search_query = f"{company_name} 简称"
        result = subprocess.run(
            ['curl', '-s', '-A', 'Mozilla/5.0', '-L',
             f'https://www.baidu.com/s?wd={search_query}'],
            capture_output=True, text=True, timeout=10
        )
        
        if result.returncode == 0 and len(result.stdout) > 100:
            # 解析搜索结果，查找常见模式
            html = result.stdout
            
            # 尝试从标题中提取
            import re
            # 查找"简称 XXX"模式
            match = re.search(r'简称 [:：\s]+([a-zA-Z\u4e00-\u9fa5]{2,8})', html)
            if match:
                alias = match.group(1).strip()
                # 过滤太短的
                if len(alias) >= 2 and alias not in GENERIC_TERMS:
                    return alias
            
            # 查找"股票代码 XXX"模式
            match = re.search(r'股票代码[:：\s]*(\d{6})', html)
            if match:
                # 如果有股票代码，返回公司名核心部分
                core = company_name.replace('股份有限公司', '').replace('有限责任公司', '')
                core = core.replace('有限公司', '').replace('集团', '')
                if len(core) >= 2:
                    return core[:6]
            
            # 如果都没找到，返回公司名核心部分
            core = company_name.replace('股份有限公司', '').replace('有限责任公司', '')
            core = core.replace('有限公司', '').replace('集团', '')
            if len(core) >= 2:
                return core[:6]
                
    except Exception as e:
        print(f"    网络查询失败：{e}")
    
    return ""

STOCK_NAMES = {
    # 银行
    # 农商行
    '广州农村商业银行股份有限公司': '广州农商行',
    '重庆农村商业银行股份有限公司': '重庆农商行',
    '成都农村商业银行股份有限公司': '成都农商行',
    '上海农村商业银行股份有限公司': '上海农商行',
    '青岛农村商业银行股份有限公司': '青岛农商行',
    '北京农村商业银行股份有限公司': '北京农商行',
    '深圳农村商业银行股份有限公司': '深圳农商行',
    '东莞农村商业银行股份有限公司': '东莞农商行',
    '江南农村商业银行股份有限公司': '江南农商行',
    
    # 城商行
    '杭州银行股份有限公司': '杭州银行',
    '江苏银行股份有限公司': '江苏银行',
    '宁波银行股份有限公司': '宁波银行',
    '南京银行股份有限公司': '南京银行',
    '成都银行股份有限公司': '成都银行',
    '长沙银行股份有限公司': '长沙银行',
    '贵阳银行股份有限公司': '贵阳银行',
    '郑州银行股份有限公司': '郑州银行',
    '青岛银行股份有限公司': '青岛银行',
    '西安银行股份有限公司': '西安银行',
    '苏州银行股份有限公司': '苏州银行',
    '重庆银行股份有限公司': '重庆银行',
    '上海银行股份有限公司': '上海银行',
    '北京银行股份有限公司': '北京银行',
    '交通银行股份有限公司': '交通银行',
    '招商银行股份有限公司': '招商银行',
    '中信银行股份有限公司': '中信银行',
    '中国民生银行股份有限公司': '民生银行',
    '中国光大银行股份有限公司': '光大银行',
    '平安银行股份有限公司': '平安银行',
    '华夏银行股份有限公司': '华夏银行',
    '兴业银行股份有限公司': '兴业银行',
    '中国工商银行股份有限公司': '工商银行',
    '中国农业银行股份有限公司': '农业银行',
    '中国银行股份有限公司': '中国银行',
    '中国建设银行股份有限公司': '建设银行',
    
    # 保险/券商
    '中国人寿保险股份有限公司': '中国人寿',
    '中国平安保险 (集团) 股份有限公司': '中国平安',
    '中国太平洋保险 (集团) 股份有限公司': '中国太保',
    '中国人民保险集团股份有限公司': '中国人保',
    '中信证券股份有限公司': '中信证券',
    '海通证券股份有限公司': '海通证券',
    '国泰君安证券股份有限公司': '国泰君安',
    '华泰证券股份有限公司': '华泰证券',
    
    # 汽车
    '比亚迪股份有限公司': '比亚迪',
    '长城汽车股份有限公司': '长城汽车',
    '安徽江淮汽车集团股份有限公司': '江淮汽车',
    '吉利汽车控股有限公司': '吉利汽车',
    '北京汽车股份有限公司': '北汽蓝谷',
    '上海汽车集团股份有限公司': '上汽集团',
    '东风汽车集团股份有限公司': '东风汽车',
    '长安汽车股份有限公司': '长安汽车',
    '一汽解放集团股份有限公司': '一汽解放',
    
    # 医药/医疗
    '云南白药集团股份有限公司': '云南白药',
    '大参林医药集团股份有限公司': '大参林',
    '北京同仁堂股份有限公司': '同仁堂',
    '华润三九医药股份有限公司': '华润三九',
    '上海复星医药 (集团) 股份有限公司': '复星医药',
    '恒瑞医药股份有限公司': '恒瑞医药',
    '石药集团有限公司': '石药集团',
    
    # 科技/制造
    '北方华创科技集团股份有限公司': '北方华创',
    '北京东方雨虹防水技术股份有限公司': '东方雨虹',
    '烽火通信科技股份有限公司': '烽火通信',
    '传化智联股份有限公司': '传化智联',
    '内蒙古伊利实业集团股份有限公司': '伊利股份',
    '宁夏宝丰能源集团股份有限公司': '宝丰能源',
    '内蒙古包钢钢联股份有限公司': '包钢股份',
    '广西柳工机械股份有限公司': '柳工',
    '洛阳栾川钼业集团股份有限公司': '洛阳钼业',
    '中国北方稀土 (集团) 高科技股份有限公司': '北方稀土',
    '大秦铁路股份有限公司': '大秦铁路',
    '东方电气股份有限公司': '东方电气',
    '东华能源股份有限公司': '东华能源',
    '安阳钢铁股份有限公司': '安阳钢铁',
    '甘肃酒钢集团宏兴钢铁股份有限公司': '酒钢宏兴',
    '白银有色集团股份有限公司': '白银有色',
    '金川集团股份有限公司': '金川股份',
    '淮北矿业控股股份有限公司': '淮北矿业',
    '安徽建工集团股份有限公司': '安徽建工',
    '圆通速递股份有限公司': '圆通速递',
    '顺丰控股股份有限公司': '顺丰控股',
    '京东物流股份有限公司': '京东物流',
    '中通快递股份有限公司': '中通快递',
    
    # 消费
    '贵州茅台酒股份有限公司': '贵州茅台',
    '五粮液股份有限公司': '五粮液',
    '泸州老窖股份有限公司': '泸州老窖',
    '山西汾酒股份有限公司': '山西汾酒',
    '青岛啤酒股份有限公司': '青岛啤酒',
    '海尔智家股份有限公司': '海尔智家',
    '美的集团股份有限公司': '美的集团',
    '格力电器股份有限公司': '格力电器',
    '苏宁易购集团股份有限公司': '苏宁易购',
    '永辉超市股份有限公司': '永辉超市',
    
    # 能源/资源
    '中国石油化工股份有限公司': '中国石化',
    '中国石油天然气股份有限公司': '中国石油',
    '中国海洋石油有限公司': '中国海油',
    '中国神华能源股份有限公司': '中国神华',
    '陕西煤业股份有限公司': '陕西煤业',
    '兖矿能源集团股份有限公司': '兖矿能源',
    '紫金矿业集团股份有限公司': '紫金矿业',
    '江西铜业股份有限公司': '江西铜业',
    '中国铝业股份有限公司': '中国铝业',
    
    # 地产/建筑
    '万科企业股份有限公司': '万科 A',
    '保利发展控股集团股份有限公司': '保利发展',
    '招商局积余产业运营服务股份有限公司': '招商积余',
    '中国建筑股份有限公司': '中国建筑',
    '中国中铁股份有限公司': '中国中铁',
    '中国铁建股份有限公司': '中国铁建',
    '中国交通建设股份有限公司': '中国交建',
    
    # 互联网/通信
    '腾讯控股有限公司': '腾讯控股',
    '阿里巴巴集团控股有限公司': '阿里巴巴',
    '百度集团股份有限公司': '百度集团',
    '京东集团股份有限公司': '京东集团',
    '网易股份有限公司': '网易',
    '小米集团股份有限公司': '小米集团',
    '美团': '美团',
    '拼多多股份有限公司': '拼多多',
    '中国移动有限公司': '中国移动',
    '中国电信股份有限公司': '中国电信',
    '中国联合网络通信股份有限公司': '中国联通',
}

def is_meaningful_alias(alias: str, all_companies: list) -> bool:
    """判断别名是否有意义（有区分度）"""
    # 规则：别名必须至少 3 个字，2 个字的别名作废
    if len(alias) < 3:
        return False
    
    # 检查是否是纯通用词
    if alias in GENERIC_TERMS:
        return False
    
    # 检查是否以通用词开头且长度相近
    for term in GENERIC_TERMS:
        if len(term) >= 2 and alias.startswith(term):
            remaining = alias[len(term):]
            if not remaining or remaining in GENERIC_TERMS or len(remaining) <= 1:
                return False
    
    # 短别名（3-4 字）需要特别检查
    if len(alias) <= 4:
        # 检查是否在太多公司中出现
        count = sum(1 for company in all_companies if alias in company)
        if count > len(all_companies) * 0.02:  # 超过 2% 就过滤
            return False
        
        # 检查是否包含通用词
        for term in GENERIC_TERMS:
            if len(term) >= 2 and term in alias:
                return False
    
    return True

def filter_aliases(alias_str: str, all_companies: list) -> str:
    """过滤掉无意义的别名"""
    if not alias_str:
        return ""
    
    aliases = alias_str.split('|')
    meaningful = []
    
    for a in aliases:
        alias = a.strip()
        
        # 1. 检查是否是纯地名
        if alias in REGION_NAMES:
            continue
        
        # 2. 检查是否以地名开头
        starts_with_region = False
        for region in sorted(REGION_NAMES, key=len, reverse=True):
            if len(region) >= 2 and alias.startswith(region):
                remaining = alias[len(region):]
                if len(remaining) >= 2 and remaining not in GENERIC_TERMS:
                    meaningful.append(remaining)
                starts_with_region = True
                break
        
        if starts_with_region:
            continue
        
        # 3. 正常检查
        if is_meaningful_alias(alias, all_companies):
            meaningful.append(alias)
    
    return '|'.join(meaningful)

def process_excel(input_file, output_file=None):
    """处理 Excel 文件，添加别名列（包含股票简称）"""
    
    # 读取输入文件
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # 收集所有公司名称
    all_companies = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if name and isinstance(name, str):
            all_companies.append(name)
    
    print(f"共收集 {len(all_companies)} 家企业")
    print(f"开始生成别名（包含股票简称 + 网络查询）...")
    print(f"预计处理时间：约 10-15 分钟（网络查询）\n")
    
    # 添加别名列
    ws.cell(row=1, column=2, value="别名（含股票简称）")
    
    # 用于去重的全局别名集合
    used_aliases = set()
    
    # 处理每一行
    stock_count = 0
    filtered_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        company_name = ws.cell(row=row_idx, column=1).value
        
        if not company_name or not isinstance(company_name, str):
            ws.cell(row=row_idx, column=2, value="")
            continue
        
        # 检查是否为政府机构
        if is_government_organization(company_name):
            # 政府机构直接使用企业全称作为别名
            final_aliases = [company_name]
        else:
            # 生成基础别名
            alias_str = format_aliases(company_name, use_web=False, use_datasource=False)
            
            # 过滤通用词
            filtered_alias = filter_aliases(alias_str, all_companies)
            
            # 如果是上市公司，添加股票简称
            final_aliases = []
            if filtered_alias:
                final_aliases = filtered_alias.split('|')
            
            # 添加股票简称（最前面，优先级最高）
            if company_name in STOCK_NAMES:
                stock_name = STOCK_NAMES[company_name]
                if stock_name not in final_aliases:
                    final_aliases.insert(0, stock_name)
                    stock_count += 1
            
            # 去重：移除已经使用过的别名
            if final_aliases:
                unique_aliases = []
                for a in final_aliases:
                    if a not in used_aliases:
                        unique_aliases.append(a)
                        used_aliases.add(a)
                final_aliases = unique_aliases
        
        # 如果所有别名都被过滤了，尝试提取核心词
        if not final_aliases:
            # 方法 1：从公司名提取（智能提取）
            core = company_name
            
            # 移除常见后缀
            for suffix in ['股份有限公司', '有限责任公司', '有限公司', '集团', '合作社', '联合社']:
                core = core.replace(suffix, '')
            
            # 移除地名前缀（但保留"中国"、"国际"等有意义的词）
            for region in sorted(REGION_NAMES, key=len, reverse=True):
                if len(region) >= 2 and core.startswith(region) and region not in ['中国', '国际']:
                    core = core[len(region):]
                    break
            
            # 检查提取的核心词是否有意义
            if len(core) >= 2:
                # 特殊处理：允许"投资"、"能源"等作为行业词（当它们不是单独出现时）
                special_industry = ['投资', '能源', '电力', '医药', '汽车', '钢铁', '高速', '铁路', '银行', '证券', '保险', '医药', '工业']
                
                has_meaning = True
                # 如果核心词包含行业词，且长度>=4，保留
                if len(core) >= 4:
                    for ind in special_industry:
                        if ind in core:
                            has_meaning = True
                            break
                    else:
                        # 检查是否纯通用词
                        for term in GENERIC_TERMS:
                            if core == term:
                                has_meaning = False
                                break
                
                if has_meaning and len(core) >= 2:
                    final_aliases = [core[:6]]
            
            # 方法 2：如果还是没别名，使用公司名的前 6 个非通用字
            if not final_aliases:
                # 尝试取公司名的核心部分
                simple_core = company_name
                for suffix in ['股份有限公司', '有限责任公司', '有限公司', '集团']:
                    simple_core = simple_core.replace(suffix, '')
                
                if len(simple_core) >= 3:
                    final_aliases = [simple_core[:6]]
        
        # 最终过滤：移除包含通用词的别名（但保留股票简称）
        if final_aliases:
            stock_name = None
            if company_name in STOCK_NAMES:
                stock_name = STOCK_NAMES[company_name]
            
            filtered_final = []
            for a in final_aliases:
                if not a:  # 跳过空字符串
                    continue
                
                # 规则：别名必须至少 3 个字，2 个字的别名作废（包括股票简称）
                if len(a) < 3:
                    continue
                    
                # 保留股票简称（但也必须>=3 字）
                if stock_name and a == stock_name:
                    filtered_final.append(a)
                    continue
                
                # 过滤纯通用词
                if a in GENERIC_TERMS:
                    continue
                
                # 过滤包含通用词的别名
                keep = True
                for term in GENERIC_TERMS:
                    if len(term) >= 2 and term in a:
                        keep = False
                        break
                
                if keep:
                    filtered_final.append(a)
            
            if filtered_final:
                final_aliases = filtered_final
        
        # 确保至少有一个别名
        if not final_aliases:
            # 最后的 fallback：使用公司名的有意义部分
            simple = company_name
            for suffix in ['股份有限公司', '有限责任公司', '有限公司', '集团']:
                simple = simple.replace(suffix, '')
            
            # 移除地名
            for region in sorted(REGION_NAMES, key=len, reverse=True):
                if len(region) >= 2 and simple.startswith(region):
                    simple = simple[len(region):]
                    break
            
            # 检查是否有意义（不能是纯通用词或包含通用词，且必须>=3 字）
            if len(simple) >= 3:
                is_generic = False
                for term in GENERIC_TERMS:
                    if simple == term:
                        is_generic = True
                        break
                    if simple.startswith(term) and len(simple) <= len(term) + 2:
                        is_generic = True
                        break
                    if len(term) >= 2 and term in simple and len(simple) <= 6:
                        is_generic = True
                        break
                
                if not is_generic:
                    final_aliases = [simple[:6]]
                else:
                    # 如果核心词包含通用词，使用完整公司名（不含后缀）
                    full_name = company_name.replace('股份有限公司', '').replace('有限责任公司', '').replace('有限公司', '')
                    if len(full_name) >= 3:
                        final_aliases = [full_name[:8]]
                    else:
                        # 如果还是<3 字，使用完整公司名
                        final_aliases = [company_name[:10]]
            else:
                full_name = company_name.replace('股份有限公司', '').replace('有限责任公司', '').replace('有限公司', '')
                if len(full_name) >= 3:
                    final_aliases = [full_name[:8]]
                else:
                    final_aliases = [company_name[:10]]
        
        # 最终检查：确保没有 2 字别名
        if final_aliases:
            final_aliases = [a for a in final_aliases if len(a) >= 3]
        
        # 如果过滤后为空，使用公司名全称
        if not final_aliases:
            final_aliases = [company_name[:12]]
        
        ws.cell(row=row_idx, column=2, value='|'.join(final_aliases) if final_aliases else '')
        
        # 每 10 行打印进度
        if row_idx % 10 == 0:
            print(f"已处理 {row_idx-1}/{len(all_companies)} 家企业...")
        
        # 对非上市公司尝试网络查询
        if company_name not in STOCK_NAMES and (not final_aliases or len(final_aliases) == 1):
            print(f"  🌐 网络查询：{company_name[:20]}...")
            web_alias = query_company_alias(company_name)
            if web_alias and web_alias not in GENERIC_TERMS and web_alias not in final_aliases:
                final_aliases.insert(0, web_alias)
                print(f"    → {web_alias}")
    
    # 保存结果
    if output_file is None:
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_with_aliases_final{ext}"
    
    wb.save(output_file)
    print(f"\n✅ 处理完成！结果已保存到：{output_file}")
    print(f"共处理 {len(all_companies)} 家企业")
    print(f"识别出 {stock_count} 家上市公司，已添加股票简称")
    print(f"过滤掉 {filtered_count} 家企业的无意义别名")
    
    return output_file

if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else '/home/admin/.openclaw/media/dingtalk/inbound/2026-03-24/dingtalk-file-1774338167380-oyaye2.xlsx'
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    process_excel(input_file, output_file)
