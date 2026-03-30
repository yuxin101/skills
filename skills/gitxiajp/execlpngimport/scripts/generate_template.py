# -*- coding: utf-8 -*-
"""
Excel 导入模板生成器主程序
用法：
    python generate_template.py <模板路径> <数据源路径> <表中文名> [输出路径]
示例：
    python generate_template.py Excel导入模板.xlsx 功能字段.xlsx 资产负债表
    python generate_template.py Excel导入模板.png 现金流量.png 现金流量表
"""
import sys
import os

def detect_field_type(name):
    """根据字段名智能识别类型"""
    amount_keywords = ['资金', '借款', '存款', '资产', '负债', '权益', '资本', 
                       '股本', '账款', '票据', '投资', '准备金', '预付款', 
                       '应收', '应付', '存货', '商誉', '费用', '收益', '利润',
                       '股利', '薪酬', '税费', '保费', '佣金', '手续费', 
                       '垫款', '合计', '总计', '公积金', '储备', '利息', 
                       '租金', '债券', '税额', '收入', '成本']
    if any(kw in name for kw in amount_keywords):
        return '金额'
    
    if any(kw in name for kw in ['日期']):
        return '日期时间'
    
    if any(kw in name for kw in ['类型']):
        return '单选'
    
    if any(kw in name for kw in ['状态']):
        return '状态标签'
    
    return '单行文本'


def process_excel(source_path):
    """从Excel文件提取字段"""
    import openpyxl
    fields = []
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        for val in row:
            if val and '不用新增' not in str(val):
                fields.append(str(val).strip())
    
    return fields


def process_image(image_path):
    """从图片OCR识别字段"""
    from cnocr import CnOcr
    
    print(f'正在识别图片: {image_path}')
    ocr = CnOcr()
    result = ocr.ocr(image_path)
    
    # 排除占位符
    keywords_to_exclude = ['请输入', '年/月/日']
    fields = []
    
    for item in result:
        text = item['text'].strip()
        if text and not any(kw in text for kw in keywords_to_exclude):
            fields.append(text)
    
    print(f'识别到 {len(fields)} 个字段')
    return fields


def generate_template(template_path, source_path, table_name, output_path=None):
    """生成导入模板"""
    import openpyxl
    from scripts.field_mapping import get_english_name
    
    # 确定输出路径
    if not output_path:
        base_dir = os.path.dirname(template_path)
        output_path = os.path.join(base_dir, f'{table_name}_导入模板.xlsx')
    
    print(f'\n开始生成模板...')
    print(f'  模板: {template_path}')
    print(f'  数据源: {source_path}')
    print(f'  表名: {table_name}')
    print(f'  输出: {output_path}')
    
    # 读取模板
    tpl_wb = openpyxl.load_workbook(template_path)
    tpl_ws = tpl_wb.active
    
    # 设置表名 ⚠️ 表名放在 B1
    tpl_ws['B1'] = table_name
    
    # 清除旧数据
    for row in range(5, tpl_ws.max_row + 1):
        for col in range(1, 6):
            tpl_ws.cell(row=row, column=col).value = None
    
    # 判断数据源类型并提取字段
    ext = os.path.splitext(source_path)[1].lower()
    if ext in ['.png', '.jpg', '.jpeg', '.bmp', '.gif']:
        fields = process_image(source_path)
    else:
        fields = process_excel(source_path)
    
    # 写入字段
    for i, name in enumerate(fields):
        row_num = i + 5
        eng_name = get_english_name(name, i)
        field_type = detect_field_type(name)
        
        tpl_ws.cell(row=row_num, column=1, value=name)
        tpl_ws.cell(row=row_num, column=2, value=eng_name)
        tpl_ws.cell(row=row_num, column=3, value=field_type)
        tpl_ws.cell(row=row_num, column=4, value='否')
    
    # 保存
    tpl_wb.save(output_path)
    print(f'\n✅ 完成！共写入 {len(fields)} 个字段')
    print(f'✅ 文件已保存: {output_path}')
    
    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    
    template_path = sys.argv[1]
    source_path = sys.argv[2]
    table_name = sys.argv[3]
    output_path = sys.argv[4] if len(sys.argv) > 4 else None
    
    generate_template(template_path, source_path, table_name, output_path)
