#!/usr/bin/env python3
"""
_render.py — Report rendering layer for email-summarizer.

NOT a standalone script. Imported by build_report.py.

Provides:
  build_report_html(contacts, label, date_range, now_str, total_emails) → str
  build_excel(contacts, label, date_range, now_str)                     → str (temp file path)
  SMTP_MAP                                                               → dict
"""

import os
import tempfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from _core import html_esc


# ── Source badge helpers ──────────────────────────────────────────────────────

_SOURCE_COLORS = {"From": "#1565c0", "To": "#2e7d32", "CC": "#6a1b9a"}


def _source_badges(source_str: str) -> str:
    if not source_str or source_str == "—":
        return '<span style="color:#bbb;">—</span>'
    parts = [s.strip() for s in source_str.split("/")]
    spans = []
    for p in parts:
        col = _SOURCE_COLORS.get(p, "#78909c")
        spans.append(
            f'<span style="background:{col};color:#fff;font-size:10px;font-weight:600;'
            f'padding:2px 7px;border-radius:8px;white-space:nowrap;'
            f'display:inline-block;margin:1px 2px 1px 0;">{p}</span>'
        )
    return "".join(spans)


def _email_badge(total: int, received: int, sent: int) -> str:
    if   total >= 10: bg = "#c62828"
    elif total >= 5:  bg = "#e65100"
    elif total >= 2:  bg = "#1565c0"
    else:             bg = "#78909c"
    return (
        f'<span style="background:{bg};color:#fff;font-size:11px;font-weight:700;'
        f'padding:3px 9px;border-radius:12px;white-space:nowrap;display:inline-block;">'
        f'{total}</span>'
        f'<div style="font-size:10px;color:#aaa;margin-top:2px;white-space:nowrap;">'
        f'{received}&nbsp;/&nbsp;{sent}</div>'
    )


def _th(label: str, align: str = "left") -> str:
    return (
        f'<th nowrap="nowrap" style="padding:10px 13px;color:#fff;font-size:10px;'
        f'text-align:{align};font-weight:700;letter-spacing:0.7px;white-space:nowrap;">'
        f"{label}</th>"
    )


def _table_row(i: int, c: dict) -> str:
    bg  = "#ffffff" if i % 2 == 0 else "#f9fafb"
    td  = ('nowrap="nowrap" style="padding:9px 13px;border-bottom:1px solid #eaecef;'
           'font-size:12px;white-space:nowrap;"')
    tds = 'style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:12px;"'
    return f"""
    <tr style="background:{bg};">
      <td {td} style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:12px;text-align:center;white-space:nowrap;">{i}</td>
      <td {td}><b>{html_esc(c["preferred_name"])}</b></td>
      <td {td} style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:11px;color:#555;white-space:nowrap;">{html_esc(c["email"])}</td>
      <td {td}>{html_esc(c["company"])}</td>
      <td {tds} style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:12px;color:#555;">{html_esc(c["position"])}</td>
      <td {tds} style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:11px;color:#444;max-width:200px;">{html_esc(c["subject_summary"])}</td>
      <td nowrap="nowrap" style="padding:9px 13px;border-bottom:1px solid #eaecef;white-space:nowrap;">{_source_badges(c["source"])}</td>
      <td nowrap="nowrap" style="padding:9px 13px;border-bottom:1px solid #eaecef;text-align:center;white-space:nowrap;">{_email_badge(c["total"], c["received"], c["sent"])}</td>
    </tr>"""


def _contact_card(c: dict) -> str:
    t  = c["total"]
    hc = "#c62828" if t >= 10 else "#e65100" if t >= 5 else "#1565c0" if t >= 2 else "#78909c"
    return f"""
<div style="background:#fff;border:1px solid #e0e4ea;border-left:4px solid {hc};
            border-radius:8px;padding:14px 20px;margin-bottom:10px;">
  <table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td style="font-size:14px;font-weight:700;color:#1a1a2e;">{html_esc(c["preferred_name"])}</td>
    <td align="right" nowrap="nowrap" style="white-space:nowrap;">
      <span style="background:{hc};color:#fff;font-size:11px;font-weight:700;padding:2px 9px;border-radius:10px;">{t} emails</span>
      &nbsp;{_source_badges(c["source"])}
    </td>
  </tr></table>
  <div style="font-size:11px;color:#888;margin-top:3px;">
    {html_esc(c["email"])} &nbsp;&middot;&nbsp;
    <b>{html_esc(c["company"])}</b> &nbsp;&middot;&nbsp;
    {html_esc(c["position"])}
  </div>
  <div style="font-size:12px;color:#555;margin-top:8px;padding:8px 10px;background:#f7f9fc;border-radius:6px;">
    {html_esc(c["subject_summary"])}
  </div>
</div>"""


# ── HTML report ───────────────────────────────────────────────────────────────

def build_report_html(contacts: list, label: str, date_range: str,
                      now_str: str, total_emails: int) -> str:
    """Return a fully self-contained HTML contact profile report string."""
    n   = len(contacts)
    tw  = sum(1 for c in contacts if c["sent"] > 0)
    act = sum(1 for c in contacts if c["total"] >= 5)
    rows  = "".join(_table_row(i + 1, c) for i, c in enumerate(contacts))
    cards = "".join(_contact_card(c) for c in contacts[:5])

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Contact Profile Report</title></head>
<body style="margin:0;padding:0;background:#f0f2f5;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f2f5;padding:28px 0;">
<tr><td align="center">
<table width="900" cellpadding="0" cellspacing="0"
       style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 16px rgba(0,0,0,0.09);">
  <tr><td style="background:linear-gradient(135deg,#0d3b8e 0%,#1565c0 60%,#1976d2 100%);padding:30px 40px;">
    <div style="font-size:22px;font-weight:700;color:#fff;">&#128101; Contact Profile Report</div>
    <div style="margin-top:6px;font-size:13px;color:rgba(255,255,255,0.75);">
      {html_esc(label)} &nbsp;&middot;&nbsp; {html_esc(date_range)} &nbsp;&middot;&nbsp; Generated {now_str}
    </div>
  </td></tr>
  <tr><td style="background:#f7f9fc;padding:18px 40px;border-bottom:1px solid #e4e8ef;">
    <table cellpadding="0" cellspacing="0"><tr>
      <td style="padding:0 24px 0 0;border-right:1px solid #dce1ea;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#1565c0;">{n}</div>
        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.8px;">Contacts</div>
      </td>
      <td style="padding:0 24px;border-right:1px solid #dce1ea;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#2e7d32;">{tw}</div>
        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.8px;">Two-way</div>
      </td>
      <td style="padding:0 24px;border-right:1px solid #dce1ea;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#c62828;">{act}</div>
        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.8px;">Active (&ge;5)</div>
      </td>
      <td style="padding:0 0 0 24px;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#e65100;">{total_emails}</div>
        <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.8px;">Total Emails</div>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:28px 40px 20px;">
    <div style="font-size:15px;font-weight:700;color:#1a1a2e;margin-bottom:14px;">
      &#128202; Contact Overview
      <span style="font-size:12px;font-weight:400;color:#aaa;margin-left:8px;">sorted by total interactions</span>
    </div>
    <div style="overflow-x:auto;border:1px solid #e0e4ea;border-radius:10px;">
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;min-width:700px;">
      <tr style="background:#1565c0;">
        {_th("#", "center")}{_th("PREFERRED&nbsp;NAME")}{_th("EMAIL")}{_th("COMPANY")}
        {_th("POSITION")}{_th("SUBJECT&nbsp;SUMMARY")}{_th("SOURCE","center")}{_th("EMAILS","center")}
      </tr>
      {rows}
    </table></div>
  </td></tr>
  <tr><td style="padding:0 40px 32px;">
    <div style="font-size:15px;font-weight:700;color:#1a1a2e;margin-bottom:14px;">&#128269; Top 5 Highlights</div>
    {cards}
  </td></tr>
  <tr><td style="background:#f7f9fc;padding:16px 40px;border-top:1px solid #e4e8ef;text-align:center;">
    <span style="font-size:11px;color:#bbb;">Generated by OpenClaw email-summarizer &nbsp;&middot;&nbsp; {now_str}</span>
  </td></tr>
</table></td></tr></table></body></html>"""


# ── Excel report ──────────────────────────────────────────────────────────────

def build_excel(contacts: list, label: str, date_range: str, now_str: str) -> str:
    """Write contact profile to a temp .xlsx file and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contact Profile"

    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor="1565C0")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin       = Side(style="thin", color="DDDDDD")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    even_fill  = PatternFill("solid", fgColor="F7F9FC")
    odd_fill   = PatternFill("solid", fgColor="FFFFFF")

    ws.merge_cells("A1:H1")
    ws["A1"] = "Contact Profile Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{label}  |  {date_range}  |  Generated {now_str}"
    ws["A2"].font = Font(name="Calibri", italic=True, color="888888", size=9)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    n  = len(contacts)
    tw = sum(1 for c in contacts if c["sent"] > 0)
    ac = sum(1 for c in contacts if c["total"] >= 5)
    ws.merge_cells("A3:H3")
    ws["A3"] = f"Contacts: {n}    Two-way: {tw}    Active (>=5): {ac}"
    ws["A3"].font = Font(name="Calibri", bold=True, color="1565C0", size=10)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 6

    COLS       = ["#", "Preferred Name", "Email", "Company", "Position",
                  "Subject Summary", "Source", "Emails (Recv/Sent)"]
    COL_WIDTHS = [4, 20, 28, 18, 22, 45, 14, 16]
    HDR_ROW    = 5
    for ci, (col_label, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=col_label)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HDR_ROW].height = 18

    wrap_al   = Alignment(vertical="top", wrap_text=True)
    nowrap_al = Alignment(vertical="top", wrap_text=False)
    for i, c in enumerate(contacts):
        row  = HDR_ROW + 1 + i
        fill = even_fill if i % 2 == 0 else odd_fill
        vals = [
            i + 1, c["preferred_name"], c["email"], c["company"],
            c["position"], c["subject_summary"], c["source"],
            f"{c['total']}  ({c['received']} recv / {c['sent']} sent)",
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = fill; cell.border = border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = wrap_al if ci in (5, 6) else nowrap_al
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)
    ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(COLS))}{HDR_ROW + n}"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="contact_profile_")
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


# ── SMTP presets ──────────────────────────────────────────────────────────────

SMTP_MAP: dict = {
    "163":     ("smtp.163.com",       465),
    "qq":      ("smtp.qq.com",        465),
    "exmail":  ("smtp.exmail.qq.com", 465),
    "gmail":   ("smtp.gmail.com",     465),
    "outlook": ("smtp.office365.com", 587),
}
