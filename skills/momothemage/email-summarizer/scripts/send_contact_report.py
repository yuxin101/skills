#!/usr/bin/env python3
"""
send_contact_report.py - Generate and email a contact profile report.

Fields per contact:
  #               — rank by total emails
  Preferred Name  — inferred from display name, email local-part, or body signature
  Email           — sender/recipient address
  Company         — inferred from domain or body content
  Position        — inferred from body/subject; "Unknown" if not determinable
  Subject Summary — one-line synthesis of all subjects with this contact
  Source          — From / To / CC badges (where this contact appeared)
  Emails          — total count badge + Recv / Sent breakdown

Usage:
  EMAIL_USER=you@163.com EMAIL_PASS=xxx python3 send_contact_report.py \
      --days 30 --preset 163 [--to recipient@email.com]
"""

import json, os, re, smtplib, subprocess, sys, argparse, tempfile
from collections import defaultdict
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))


# ── Low-level helpers ─────────────────────────────────────────────────────────

def strip_html(text: str) -> str:
    text = re.sub(r'<[^>]+>', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def parse_addr(raw: str):
    raw = raw.strip()
    m = re.match(r'^"?([^"<]+?)"?\s*<([^>]+)>', raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    if '@' in raw:
        return raw, raw
    return raw, ''


def get_domain(addr: str) -> str:
    m = re.search(r'@([\w.\-]+)', addr)
    return m.group(1).lower() if m else ''


def html_esc(t: object) -> str:
    return str(t).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


# ── Inference: Company ────────────────────────────────────────────────────────

DOMAIN_COMPANY = {
    'github.com':                   'GitHub',
    'linkedin.com':                 'LinkedIn',
    'xing.com':                     'XING',
    'mail.xing.com':                'XING',
    'lietou-edm.com':               'Liepin',
    'liepin.com':                   'Liepin',
    'maimai.mobi':                  'Maimai',
    'patreon.com':                  'Patreon',
    'discord.com':                  'Discord',
    'mail.hsbc.com.hk':             'HSBC',
    'notification.hsbc.com.hk':     'HSBC',
    'hsbc.com.hk':                  'HSBC',
    'bank.za.group':                'ZA Bank',
    'hello.stackoverflow.email':    'Stack Overflow',
    'stackoverflow.email':          'Stack Overflow',
    'mail.coursera.org':            'Coursera / Stanford',
    'e.cathaypacific.com':          'Cathay Pacific',
    'conductor.build':              'Conductor',
    'email.webook.com':             'Webook',
    'nia.gov.cn':                   'China Immigration Bureau',
    'service.netease.com':          'NetEase (163 Mail)',
    '163.com':                      '163 Mail (NetEase)',
    'quora.com':                    'Quora',
    'tavily.com':                   'Tavily',
}

def infer_company(domain: str, name: str, bodies: list) -> str:
    d = domain.lower()
    for k, v in DOMAIN_COMPANY.items():
        if d == k or d.endswith('.' + k):
            return v
    combined = ' '.join(bodies[:3])
    for pat in [r'\bat\s+([A-Z][A-Za-z0-9 &,.\-]{2,40})',
                r'Company[:\s]+([A-Za-z0-9 &,.\-]{2,40})']:
        m = re.search(pat, combined)
        if m:
            return m.group(1).strip().rstrip('.,')
    for part in d.split('.'):
        if part not in ('www', 'mail', 'smtp', 'imap', 'com', 'org', 'net', 'cn', 'hk', 'io'):
            return part.capitalize()
    return domain


# ── Inference: Preferred Name ─────────────────────────────────────────────────

_SYSTEM_KW = {
    'noreply', 'no-reply', 'mailrobot', 'notification', 'service',
    'donotreply', 'do-not-reply', 'support', 'info', 'admin', 'alert',
    'newsletter', 'mailer', 'robot', 'bot', 'auto', 'jobs', 'digest',
    'news', 'mkt', 'help', 'safe', 'security',
}

def infer_preferred_name(display_name: str, addr: str, bodies: list) -> str:
    n = display_name.strip()
    local = addr.split('@')[0].lower()
    if n and not any(k in n.lower() for k in _SYSTEM_KW) and len(n) <= 50:
        if re.search(r'[A-Za-z\u4e00-\u9fff]', n):
            return n
    combined = '\n'.join(bodies[:3])
    for pat in [
        r'(?:Best|Regards|Thanks|Cheers|Sincerely|Warm regards)[,\s]+([A-Z][a-z]+(?: [A-Z][a-z]+)?)',
        r'--\s*\n([A-Z][a-z]+(?: [A-Z][a-z]+)?)',
    ]:
        m = re.search(pat, combined)
        if m:
            return m.group(1).strip()
    if any(k in local for k in _SYSTEM_KW):
        return display_name or local
    return re.sub(r'[\._\-\+\d]+', ' ', local).strip().title() or display_name or local


# ── Inference: Position ───────────────────────────────────────────────────────

def infer_position(display_name: str, domain: str, subjects: list, bodies: list) -> str:
    d = domain.lower()
    if 'lietou' in d or 'liepin' in d:  return 'Headhunter / Recruiter'
    if 'linkedin.com' in d:             return 'Platform Bot (LinkedIn)'
    if 'xing.com' in d:
        return 'Job Recommendation Bot' if 'jobs' in display_name.lower() else 'Platform Bot (XING)'
    if 'github.com' in d:               return 'Platform Bot (GitHub)'
    if 'discord.com' in d:              return 'Platform Bot (Discord)'
    if 'hsbc' in d:                     return 'Banking Service'
    if 'za.group' in d:                 return 'Banking Service'
    if 'stackoverflow' in d or 'coursera' in d or 'stanford' in d:
        return 'Platform Bot (Education)'
    if 'cathaypacific' in d or 'cathay' in d: return 'Airline Service'
    if 'nia.gov.cn' in d:               return 'Government Service'
    if 'netease' in d or '163.com' in d: return 'Platform Bot (Mail Provider)'
    if 'patreon.com' in d:              return 'Content Creator'

    body_text = ' '.join(bodies[:5])
    for pat in [
        r'(?:Title|Position|Role)[:\s]+([A-Za-z &/\-]{3,50})',
        r'\n([A-Z][a-z]+(?: [A-Z][a-z]+)?\s*\|)',
    ]:
        m = re.search(pat, body_text)
        if m:
            c = m.group(1).strip(' |')
            if 3 < len(c) < 60:
                return c

    subj_text = ' '.join(subjects).lower()
    if any(k in subj_text for k in ['recruit', 'opportunity', 'position', 'job', 'career', 'hiring']):
        return 'Recruiter'
    if any(k in subj_text for k in ['invoice', 'payment', 'purchase', 'order']):
        return 'Finance / Billing'
    if any(k in subj_text for k in ['security', 'alert', 'verify', 'confirm', 'login']):
        return 'Security Service'
    return 'Unknown'


# ── Inference: Subject Summary ────────────────────────────────────────────────

def summarise_subjects(subjects: list) -> str:
    if not subjects:
        return '—'
    unique = list(dict.fromkeys(s.strip() for s in subjects if s.strip()))
    if len(unique) == 1:
        return unique[0][:90]
    clean = []
    for s in unique:
        s = re.sub(r'^(?:Re|Fwd|FW|AW|回复|转发)[:\s]+', '', s, flags=re.IGNORECASE).strip()
        s = re.sub(r'^\[.*?\]\s*', '', s).strip()
        if s:
            clean.append(s)
    if not clean:
        return unique[0][:90]
    words = [set(re.findall(r'[a-zA-Z]{4,}', s.lower())) for s in clean]
    if len(words) >= 2:
        common = words[0].intersection(*words[1:])
        common -= {'your', 'with', 'this', 'that', 'from', 'have', 'will', 'been', 'about'}
        if common:
            theme = max(common, key=len).title()
            return f'{theme}-related ({len(unique)} emails)'
    preview = ' · '.join(clean[:3])
    if len(unique) > 3:
        preview += f' (+{len(unique)-3} more)'
    return preview[:120]


# ── Build contact list ────────────────────────────────────────────────────────

MERGE_MAP = {
    'service@mail8.lietou-edm.com':  ('Liepin', 'service@liepin.com'),
    'service@mail19.lietou-edm.com': ('Liepin', 'service@liepin.com'),
}

def build_contacts(emails: list, owner: str) -> list:
    raw: dict = defaultdict(lambda: {
        'display_name': '', 'email': '', 'domain': '',
        'received': 0, 'sent': 0,
        'subjects': [], 'bodies': [],
        'sources': set(),
    })

    for e in emails:
        direction = e.get('direction', 'received')
        subj      = e.get('subject', '').strip()
        body_raw  = e.get('body', '')
        body      = (strip_html(body_raw) if '<' in body_raw else body_raw)[:500]

        if direction == 'received':
            name, addr = parse_addr(e.get('from', ''))
            if not addr:
                continue
            raw_key = addr.lower()
            if raw_key in MERGE_MAP:
                name, addr = MERGE_MAP[raw_key]
            key = addr.lower()
            c = raw[key]
            if not c['email']:
                c['display_name'] = name
                c['email']        = addr
                c['domain']       = get_domain(addr)
            c['received'] += 1
            c['subjects'].append(subj)
            c['bodies'].append(body)
            c['sources'].add('From')
            # Track anyone CC'd on received mail
            for part in e.get('cc', '').split(','):
                _, cc_addr = parse_addr(part.strip())
                if cc_addr and owner not in cc_addr.lower():
                    ck = cc_addr.lower()
                    raw[ck]['sources'].add('CC')
                    if not raw[ck]['email']:
                        raw[ck]['email']  = cc_addr
                        raw[ck]['domain'] = get_domain(cc_addr)
        else:
            # Sent mail: track To and CC recipients
            for part in e.get('to', '').split(','):
                _, addr = parse_addr(part.strip())
                if not addr or owner in addr.lower():
                    continue
                key = addr.lower()
                raw[key]['sent'] += 1
                raw[key]['sources'].add('To')
                if not raw[key]['email']:
                    raw[key]['email']  = addr
                    raw[key]['domain'] = get_domain(addr)
            for part in e.get('cc', '').split(','):
                _, addr = parse_addr(part.strip())
                if not addr or owner in addr.lower():
                    continue
                key = addr.lower()
                raw[key]['sources'].add('CC')
                if not raw[key]['email']:
                    raw[key]['email']  = addr
                    raw[key]['domain'] = get_domain(addr)

    raw.pop(owner, None)

    order = {'From': 0, 'To': 1, 'CC': 2}
    result = []
    for c in raw.values():
        total     = c['received'] + c['sent']
        pref_name = infer_preferred_name(c['display_name'], c['email'], c['bodies'])
        company   = infer_company(c['domain'], c['display_name'], c['bodies'])
        position  = infer_position(c['display_name'], c['domain'], c['subjects'], c['bodies'])
        summary   = summarise_subjects(c['subjects'])
        source    = ' / '.join(sorted(c['sources'], key=lambda x: order.get(x, 9))) or '—'
        result.append({
            'preferred_name':  pref_name,
            'email':           c['email'],
            'company':         company,
            'position':        position,
            'subject_summary': summary,
            'source':          source,
            'received':        c['received'],
            'sent':            c['sent'],
            'total':           total,
        })

    return sorted(result, key=lambda x: -x['total'])


# ── HTML helpers ──────────────────────────────────────────────────────────────

def _th(label: str, align: str = 'left') -> str:
    return (f'<th nowrap="nowrap" style="padding:10px 13px;color:#fff;font-size:10px;'
            f'text-align:{align};font-weight:700;letter-spacing:0.7px;'
            f'white-space:nowrap;">{label}</th>')


SOURCE_COLORS = {'From': '#1565c0', 'To': '#2e7d32', 'CC': '#6a1b9a'}

def _source_badges(source_str: str) -> str:
    if not source_str or source_str == '—':
        return '<span style="color:#bbb;">—</span>'
    out = []
    for p in [s.strip() for s in source_str.split('/')]:
        col = SOURCE_COLORS.get(p, '#78909c')
        out.append(
            f'<span style="background:{col};color:#fff;font-size:10px;font-weight:600;'
            f'padding:2px 7px;border-radius:8px;white-space:nowrap;'
            f'display:inline-block;margin:1px 2px 1px 0;">{p}</span>'
        )
    return ''.join(out)


def _email_badge(total: int, received: int, sent: int) -> str:
    if total >= 10:  bg = '#c62828'
    elif total >= 5: bg = '#e65100'
    elif total >= 2: bg = '#1565c0'
    else:            bg = '#78909c'
    return (
        f'<span style="background:{bg};color:#fff;font-size:11px;font-weight:700;'
        f'padding:3px 9px;border-radius:12px;white-space:nowrap;'
        f'display:inline-block;">{total}</span>'
        f'<div style="font-size:10px;color:#aaa;margin-top:2px;white-space:nowrap;">'
        f'{received}&nbsp;/&nbsp;{sent}</div>'
    )


def build_table_row(i: int, c: dict) -> str:
    bg  = '#ffffff' if i % 2 == 0 else '#f9fafb'
    td  = ('nowrap="nowrap" style="padding:9px 13px;border-bottom:1px solid #eaecef;'
           'font-size:12px;white-space:nowrap;"')
    tds =  'style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:12px;"'

    return f'''
    <tr style="background:{bg};">
      <td {td} style="padding:9px 13px;border-bottom:1px solid #eaecef;font-size:12px;
                      text-align:center;white-space:nowrap;">{i}</td>
      <td {td}><b>{html_esc(c["preferred_name"])}</b></td>
      <td {td} style="padding:9px 13px;border-bottom:1px solid #eaecef;
                      font-size:11px;color:#555;white-space:nowrap;">{html_esc(c["email"])}</td>
      <td {td}>{html_esc(c["company"])}</td>
      <td {tds} style="padding:9px 13px;border-bottom:1px solid #eaecef;
                       font-size:12px;color:#555;">{html_esc(c["position"])}</td>
      <td {tds} style="padding:9px 13px;border-bottom:1px solid #eaecef;
                       font-size:11px;color:#444;max-width:200px;">{html_esc(c["subject_summary"])}</td>
      <td nowrap="nowrap" style="padding:9px 13px;border-bottom:1px solid #eaecef;
                                  white-space:nowrap;">{_source_badges(c["source"])}</td>
      <td nowrap="nowrap" style="padding:9px 13px;border-bottom:1px solid #eaecef;
                                  text-align:center;white-space:nowrap;">
        {_email_badge(c["total"], c["received"], c["sent"])}
      </td>
    </tr>'''


def build_contact_card(c: dict) -> str:
    t  = c['total']
    hc = '#c62828' if t >= 10 else '#e65100' if t >= 5 else '#1565c0' if t >= 2 else '#78909c'
    return f'''
<div style="background:#fff;border:1px solid #e0e4ea;border-left:4px solid {hc};
            border-radius:8px;padding:14px 20px;margin-bottom:10px;">
  <table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td style="font-size:14px;font-weight:700;color:#1a1a2e;">{html_esc(c["preferred_name"])}</td>
    <td align="right" nowrap="nowrap" style="white-space:nowrap;">
      <span style="background:{hc};color:#fff;font-size:11px;font-weight:700;
                   padding:2px 9px;border-radius:10px;">{t} emails</span>
      &nbsp;{_source_badges(c["source"])}
    </td>
  </tr></table>
  <div style="font-size:11px;color:#888;margin-top:3px;">
    {html_esc(c["email"])} &nbsp;&middot;&nbsp;
    <b>{html_esc(c["company"])}</b> &nbsp;&middot;&nbsp;
    {html_esc(c["position"])}
  </div>
  <div style="font-size:12px;color:#555;margin-top:8px;padding:8px 10px;
              background:#f7f9fc;border-radius:6px;">
    {html_esc(c["subject_summary"])}
  </div>
</div>'''


# ── Full HTML report ──────────────────────────────────────────────────────────

def build_report_html(contacts: list, email_user: str, date_range: str,
                      now_str: str, total_emails: int) -> str:
    n_contacts = len(contacts)
    two_way    = sum(1 for c in contacts if c['sent'] > 0)
    active     = sum(1 for c in contacts if c['total'] >= 5)

    table_rows = ''.join(build_table_row(i + 1, c) for i, c in enumerate(contacts))
    cards      = ''.join(build_contact_card(c) for c in contacts[:5])

    return f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>Contact Profile Report</title></head>
<body style="margin:0;padding:0;background:#f0f2f5;
  font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0"
       style="background:#f0f2f5;padding:28px 0;">
<tr><td align="center">
<table width="900" cellpadding="0" cellspacing="0"
       style="background:#fff;border-radius:12px;overflow:hidden;
              box-shadow:0 2px 16px rgba(0,0,0,0.09);">

  <!-- Header -->
  <tr><td style="background:linear-gradient(135deg,#0d3b8e 0%,#1565c0 60%,#1976d2 100%);
                  padding:30px 40px;">
    <div style="font-size:22px;font-weight:700;color:#fff;">&#128101; Contact Profile Report</div>
    <div style="margin-top:6px;font-size:13px;color:rgba(255,255,255,0.75);">
      {html_esc(email_user)} &nbsp;&middot;&nbsp; {html_esc(date_range)}
      &nbsp;&middot;&nbsp; Generated {now_str}
    </div>
  </td></tr>

  <!-- Stats bar -->
  <tr><td style="background:#f7f9fc;padding:18px 40px;border-bottom:1px solid #e4e8ef;">
    <table cellpadding="0" cellspacing="0"><tr>
      <td style="padding:0 24px 0 0;border-right:1px solid #dce1ea;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#1565c0;">{n_contacts}</div>
        <div style="font-size:10px;color:#888;margin-top:2px;text-transform:uppercase;letter-spacing:.8px;">Contacts</div>
      </td>
      <td style="padding:0 24px;border-right:1px solid #dce1ea;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#2e7d32;">{two_way}</div>
        <div style="font-size:10px;color:#888;margin-top:2px;text-transform:uppercase;letter-spacing:.8px;">Two-way</div>
      </td>
      <td style="padding:0 24px;border-right:1px solid #dce1ea;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#c62828;">{active}</div>
        <div style="font-size:10px;color:#888;margin-top:2px;text-transform:uppercase;letter-spacing:.8px;">Active (&ge;5)</div>
      </td>
      <td style="padding:0 0 0 24px;text-align:center;">
        <div style="font-size:28px;font-weight:800;color:#e65100;">{total_emails}</div>
        <div style="font-size:10px;color:#888;margin-top:2px;text-transform:uppercase;letter-spacing:.8px;">Total Emails</div>
      </td>
    </tr></table>
  </td></tr>

  <!-- Overview table -->
  <tr><td style="padding:28px 40px 20px;">
    <div style="font-size:15px;font-weight:700;color:#1a1a2e;margin-bottom:14px;">
      &#128202; Contact Overview
      <span style="font-size:12px;font-weight:400;color:#aaa;margin-left:8px;">sorted by total interactions</span>
    </div>
    <div style="overflow-x:auto;border:1px solid #e0e4ea;border-radius:10px;">
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;min-width:700px;">
      <tr style="background:#1565c0;">
        {_th('#',                    'center')}
        {_th('PREFERRED&nbsp;NAME')}
        {_th('EMAIL')}
        {_th('COMPANY')}
        {_th('POSITION')}
        {_th('SUBJECT&nbsp;SUMMARY')}
        {_th('SOURCE',               'center')}
        {_th('EMAILS',               'center')}
      </tr>
      {table_rows}
    </table>
    </div>
    <div style="margin-top:9px;font-size:10px;color:#bbb;">
      Source badges: <span style="background:#1565c0;color:#fff;padding:1px 6px;border-radius:5px;font-size:9px;">From</span>
      <span style="background:#2e7d32;color:#fff;padding:1px 6px;border-radius:5px;font-size:9px;">To</span>
      <span style="background:#6a1b9a;color:#fff;padding:1px 6px;border-radius:5px;font-size:9px;">CC</span>
      &nbsp;&nbsp;|&nbsp;&nbsp;
      Emails badge: &#128308;&nbsp;Heavy(&ge;10) &#128992;&nbsp;Active(5-9) &#128994;&nbsp;Moderate(2-4) &#9898;&nbsp;Light(1)
      &nbsp;&nbsp;|&nbsp;&nbsp; Recv&nbsp;/&nbsp;Sent shown below badge
    </div>
  </td></tr>

  <!-- Top-5 cards -->
  <tr><td style="padding:0 40px 32px;">
    <div style="font-size:15px;font-weight:700;color:#1a1a2e;margin-bottom:14px;">
      &#128269; Top 5 Contact Highlights
    </div>
    {cards}
  </td></tr>

  <!-- Footer -->
  <tr><td style="background:#f7f9fc;padding:16px 40px;border-top:1px solid #e4e8ef;text-align:center;">
    <span style="font-size:11px;color:#bbb;">
      Generated by OpenClaw email-summarizer &nbsp;&middot;&nbsp; {now_str}
    </span>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>'''


# ── Excel export ─────────────────────────────────────────────────────────────

def build_excel(contacts: list, email_user: str, date_range: str, now_str: str) -> str:
    """Write contact profile to a temp .xlsx file and return its path."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contact Profile'

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill   = PatternFill('solid', fgColor='1565C0')
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=False)

    meta_font  = Font(name='Calibri', italic=True, color='888888', size=9)
    title_font = Font(name='Calibri', bold=True, size=13)

    cell_align  = Alignment(vertical='top', wrap_text=True)
    nowrap_align = Alignment(vertical='top', wrap_text=False)

    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    even_fill = PatternFill('solid', fgColor='F7F9FC')
    odd_fill  = PatternFill('solid', fgColor='FFFFFF')

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Contact Profile Report'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'{email_user}  |  {date_range}  |  Generated {now_str}'
    ws['A2'].font = meta_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # ── Stats row ─────────────────────────────────────────────────────────────
    total_contacts = len(contacts)
    two_way  = sum(1 for c in contacts if c['sent'] > 0)
    active   = sum(1 for c in contacts if c['total'] >= 5)

    ws.merge_cells('A3:H3')
    ws['A3'] = (f'Contacts: {total_contacts}    Two-way: {two_way}'
                f'    Active (>=5): {active}')
    ws['A3'].font = Font(name='Calibri', bold=True, color='1565C0', size=10)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 16

    # blank separator
    ws.row_dimensions[4].height = 6

    # ── Header row ────────────────────────────────────────────────────────────
    COLS = ['#', 'Preferred Name', 'Email', 'Company',
            'Position', 'Subject Summary', 'Source', 'Emails (Recv/Sent)']
    COL_WIDTHS = [4, 20, 28, 18, 22, 45, 14, 16]

    HDR_ROW = 5
    for ci, (label, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=label)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = hdr_align
        cell.border  = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[HDR_ROW].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, c in enumerate(contacts):
        row  = HDR_ROW + 1 + i
        fill = even_fill if i % 2 == 0 else odd_fill

        values = [
            i + 1,
            c['preferred_name'],
            c['email'],
            c['company'],
            c['position'],
            c['subject_summary'],
            c['source'],
            f"{c['total']}  ({c['received']} recv / {c['sent']} sent)",
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(name='Calibri', size=10)
            # wrap only for summary and position columns
            cell.alignment = cell_align if ci in (5, 6) else nowrap_align

        ws.row_dimensions[row].height = 28

    # freeze panes below header
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # auto-filter on header row
    ws.auto_filter.ref = (
        f'A{HDR_ROW}:{get_column_letter(len(COLS))}{HDR_ROW + len(contacts)}'
    )

    # ── Save ──────────────────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(
        suffix='.xlsx', delete=False,
        prefix='contact_profile_'
    )
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Build and email a contact profile report.')
    parser.add_argument('--days',   type=int, default=30,
                        help='Look-back window in days (default 30). Ignored if --since is set.')
    parser.add_argument('--since',  type=str, default=None,
                        help='Start date, inclusive. Format: YYYY-MM-DD (e.g. 2026-03-01)')
    parser.add_argument('--until',  type=str, default=None,
                        help='End date, exclusive. Format: YYYY-MM-DD (e.g. 2026-03-25). Defaults to today.')
    parser.add_argument('--max',    type=int, default=100, help='Max emails to fetch per folder')
    parser.add_argument('--preset', default='163',         help='Mailbox preset')
    parser.add_argument('--to',        default='',  help='Override recipient address')
    parser.add_argument('--from-json', default='',  help='Load emails from a local JSON file (skip IMAP fetch)')
    args = parser.parse_args()

    email_user = os.environ.get('EMAIL_USER', '')
    email_pass = os.environ.get('EMAIL_PASS', '')
    recipient  = args.to or email_user

    if not email_user or not email_pass:
        print('ERROR: EMAIL_USER and EMAIL_PASS must be set.', file=sys.stderr)
        sys.exit(1)

    # ── Resolve display date range ────────────────────────────────────────────
    def parse_ymd(s):
        return datetime.strptime(s, '%Y-%m-%d')

    if args.since:
        since_dt  = parse_ymd(args.since)
    else:
        since_dt  = datetime.now() - timedelta(days=args.days)

    until_dt   = parse_ymd(args.until) if args.until else datetime.now()
    date_range = (f'{since_dt.strftime("%b %d, %Y")} – {until_dt.strftime("%b %d, %Y")}')
    # ─────────────────────────────────────────────────────────────────────────

    # ── Load emails: local JSON or IMAP fetch ─────────────────────────────────
    from_json = getattr(args, 'from_json', '')
    if from_json:
        print(f'[load] Reading emails from {from_json}')
        with open(from_json, 'r', encoding='utf-8') as f:
            raw = json.load(f)
        # Support both bare list and profile dict (with "emails" key)
        if isinstance(raw, dict) and 'emails' in raw:
            emails = raw['emails']
            date_range = (f'{raw.get("since", since_dt.strftime("%b %d, %Y"))} – '
                          f'{raw.get("until", until_dt.strftime("%b %d, %Y"))}')
        else:
            emails = raw
        print(f'[load] Got {len(emails)} emails from local file.')
    else:
        since_arg = ['--since', args.since] if args.since else ['--days', str(args.days)]
        until_arg = ['--until', args.until] if args.until else []
        print(f'[fetch] Date range: {date_range} (max {args.max} per folder, preset={args.preset})')
        result = subprocess.run(
            [sys.executable, os.path.join(HERE, 'fetch_emails.py'),
             *since_arg, *until_arg,
             '--preset', args.preset,
             '--max',    str(args.max),
             '--with-sent'],
            capture_output=True, text=True,
            env={**os.environ, 'EMAIL_USER': email_user, 'EMAIL_PASS': email_pass},
        )
        if result.returncode != 0:
            print('ERROR:\n' + result.stderr, file=sys.stderr); sys.exit(1)
        emails = json.loads(result.stdout)
        print(f'[fetch] Got {len(emails)} emails.')

    contacts   = build_contacts(emails, email_user.lower())
    now_str    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    html       = build_report_html(contacts, email_user, date_range, now_str, len(emails))

    # Build Excel attachment
    print('[excel] Generating Excel report…')
    xlsx_path = build_excel(contacts, email_user, date_range, now_str)
    xlsx_name = (f'contact_profile_'
                 f'{since_dt.strftime("%Y%m%d")}-{until_dt.strftime("%Y%m%d")}.xlsx')
    print(f'[excel] Saved to {xlsx_path}')

    # Compose email
    msg = MIMEMultipart('mixed')
    msg['Subject'] = f'Contact Profile Report — {date_range} ({email_user})'
    msg['From']    = email_user
    msg['To']      = recipient

    body_part = MIMEMultipart('alternative')
    body_part.attach(MIMEText(html, 'html', 'utf-8'))
    msg.attach(body_part)

    with open(xlsx_path, 'rb') as f:
        xlsx_data = f.read()
    att = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    att.set_payload(xlsx_data)
    encoders.encode_base64(att)
    att.add_header('Content-Disposition', 'attachment', filename=xlsx_name)
    msg.attach(att)
    os.unlink(xlsx_path)

    smtp_map = {
        '163':     ('smtp.163.com',       465),
        'qq':      ('smtp.qq.com',        465),
        'exmail':  ('smtp.exmail.qq.com', 465),
        'gmail':   ('smtp.gmail.com',     465),
        'outlook': ('smtp.office365.com', 587),
    }
    host, port = smtp_map.get(args.preset, ('smtp.163.com', 465))
    print(f'[smtp] Connecting to {host}:{port}…')
    if port == 587:
        with smtplib.SMTP(host, port) as s:
            s.starttls(); s.login(email_user, email_pass)
            s.sendmail(email_user, recipient, msg.as_bytes())
    else:
        with smtplib.SMTP_SSL(host, port) as s:
            s.login(email_user, email_pass)
            s.sendmail(email_user, recipient, msg.as_bytes())

    print(f'[done] Sent to {recipient} — {len(contacts)} contacts + Excel ({date_range}).')


if __name__ == '__main__':
    main()
