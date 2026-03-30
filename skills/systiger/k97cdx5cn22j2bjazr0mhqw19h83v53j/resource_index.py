# -*- coding: utf-8 -*-
"""
AI Animation Studio - 资源索引生成器
扫描 D:\AI视频资源 目录，生成结构化索引
"""

import os
import json
from pathlib import Path

RESOURCE_DIR = r"D:\AI视频资源"

def scan_resources():
    """扫描资源目录，生成索引"""
    index = {
        "分镜模板": [],
        "特效素材": {},
        "音效素材": {}
    }
    
    # 扫描根目录文档
    for f in os.listdir(RESOURCE_DIR):
        path = os.path.join(RESOURCE_DIR, f)
        if os.path.isfile(path):
            ext = os.path.splitext(f)[1].lower()
            if ext in ['.xlsx', '.csv', '.docx']:
                index["分镜模板"].append({
                    "name": f,
                    "path": path,
                    "type": ext
                })
    
    # 扫描特效素材包
    effects_dir = os.path.join(RESOURCE_DIR, "短视频特效素材包")
    if os.path.exists(effects_dir):
        for category in os.listdir(effects_dir):
            cat_path = os.path.join(effects_dir, category)
            if os.path.isdir(cat_path):
                files = []
                for root, _, filenames in os.walk(cat_path):
                    for filename in filenames:
                        if filename.endswith(('.mov', '.mp4', '.avi')):
                            files.append({
                                "name": filename,
                                "path": os.path.join(root, filename)
                            })
                if files:
                    index["特效素材"][category] = files
    
    # 扫描音效素材包
    audio_dir = os.path.join(RESOURCE_DIR, "音效素材包")
    if os.path.exists(audio_dir):
        for category in os.listdir(audio_dir):
            cat_path = os.path.join(audio_dir, category)
            if os.path.isdir(cat_path):
                files = []
                for root, _, filenames in os.walk(cat_path):
                    for filename in filenames:
                        if filename.endswith(('.mp3', '.wav', '.aac')):
                            files.append({
                                "name": filename,
                                "path": os.path.join(root, filename)
                            })
                if files:
                    index["音效素材"][category] = files
    
    return index

def load_storyboard_template():
    """加载分镜模板数据"""
    from openpyxl import load_workbook
    
    template_path = os.path.join(RESOURCE_DIR, "AI视频脚本分镜模板_共300条.xlsx")
    if not os.path.exists(template_path):
        return []
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    templates = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 有序号
            templates.append({
                headers[i]: row[i] for i in range(len(headers)) if row[i]
            })
    
    return templates

def load_style_prompts():
    """加载电影风格提示词"""
    import csv
    
    csv_path = os.path.join(RESOURCE_DIR, "300+电影风格提示词.csv")
    if not os.path.exists(csv_path):
        return []
    
    styles = []
    with open(csv_path, 'r', encoding='gbk') as f:
        reader = csv.DictReader(f)
        for row in reader:
            styles.append({
                "category": row.get("分类", ""),
                "prompt": row.get("风格提示词", "")
            })
    
    return styles

def load_shot_prompts():
    """加载镜头提示词"""
    from openpyxl import load_workbook
    
    xlsx_path = os.path.join(RESOURCE_DIR, "【先看这个】分镜画面提示词.xlsx")
    if not os.path.exists(xlsx_path):
        return []
    
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    prompts = []
    current_category = None
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            current_category = row[0]
        if row[1]:
            prompts.append({
                "category": current_category,
                "prompt": row[1]
            })
    
    return prompts

def load_ai_prompts():
    """加载AI生图指令"""
    from openpyxl import load_workbook
    
    xlsx_path = os.path.join(RESOURCE_DIR, "即梦100多组（700+个）神级指令合集.xlsx")
    if not os.path.exists(xlsx_path):
        return []
    
    wb = load_workbook(xlsx_path)
    ws = wb.active
    
    # 第二行是列名
    headers = [cell.value for cell in ws[2]]
    prompts = []
    
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            prompts.append({
                headers[i]: row[i] for i in range(min(len(headers), len(row))) if row[i]
            })
    
    return prompts

def get_random_effect(effect_type="背景"):
    """随机获取特效素材"""
    import random
    
    index = scan_resources()
    
    if effect_type == "背景":
        categories = ["2. 款竖版短视频动态背景素材", "4. 竖版短视频动态效果素材"]
    elif effect_type == "转场":
        categories = ["1. 划屏元素素材"]
    elif effect_type == "特效":
        categories = ["5. 抖音快手透明点赞心形素材", "7. 竖版唯美叠加炫光素材"]
    else:
        categories = list(index["特效素材"].keys())
    
    for cat in categories:
        if cat in index["特效素材"]:
            files = index["特效素材"][cat]
            if files:
                return random.choice(files)["path"]
    
    return None

def get_random_audio(audio_type="bgm"):
    """随机获取音效素材"""
    import random
    
    index = scan_resources()
    
    if audio_type == "bgm":
        categories = ["卡点音乐", "网红爆款bgm", "搞笑bgm"]
    elif audio_type == "sfx":
        categories = ["字幕弹出音效", "短音效", "常用音效汇总"]
    elif audio_type == "laugh":
        categories = ["各种笑"]
    else:
        categories = list(index["音效素材"].keys())
    
    for cat in categories:
        if cat in index["音效素材"]:
            files = index["音效素材"][cat]
            if files:
                return random.choice(files)["path"]
    
    return None

if __name__ == "__main__":
    # 测试
    print("=== 分镜模板 ===")
    templates = load_storyboard_template()
    print(f"共 {len(templates)} 条分镜模板")
    if templates:
        print("示例:", templates[0])
    
    print("\n=== 电影风格 ===")
    styles = load_style_prompts()
    print(f"共 {len(styles)} 种风格")
    if styles:
        print("示例:", styles[0])
    
    print("\n=== 镜头提示词 ===")
    shots = load_shot_prompts()
    print(f"共 {len(shots)} 个镜头提示词")
    if shots:
        print("示例:", shots[0])
    
    print("\n=== AI指令 ===")
    ai_prompts = load_ai_prompts()
    print(f"共 {len(ai_prompts)} 条AI指令")
    if ai_prompts:
        print("示例:", ai_prompts[0])
    
    print("\n=== 资源统计 ===")
    index = scan_resources()
    print(f"分镜模板: {len(index['分镜模板'])} 个")
    print(f"特效素材: {sum(len(v) for v in index['特效素材'].values())} 个")
    print(f"音效素材: {sum(len(v) for v in index['音效素材'].values())} 个")
