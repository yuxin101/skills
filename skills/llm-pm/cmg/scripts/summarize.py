#!/usr/bin/env python3
"""
CMG 扫描结果汇总脚本
用法: python3 summarize.py <scan_result.xlsx>
"""
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("缺少依赖，请先执行: pip3 install openpyxl")
    sys.exit(1)


def summarize(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path)

    total = 0
    product_counts = {}          # 产品 -> 实例数
    region_counts = defaultdict(int)   # 地域 -> 实例数
    product_region = defaultdict(lambda: defaultdict(int))  # 产品 -> 地域 -> 实例数

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = list(rows[0])
        data_rows = rows[1:]
        if not data_rows:
            continue

        # 查找包含"地域"的列，没有则跳过（全球级产品/子资源，如 VSwitch、RAM、DDOS 等）
        region_idx = None
        region_col_name = None
        for i, h in enumerate(headers):
            if h and "地域" in str(h):
                region_idx = i
                region_col_name = str(h)
                break
        if region_idx is None:
            # 该 Sheet 没有地域列，跳过（或可单独统计全球级服务）
            continue

        # 提取产品名：Sheet 名格式通常为 ACS_ECS_Instance / ACS_VPC_VPC 等
        parts = sheet_name.split("_")
        product = parts[1] if len(parts) >= 2 else sheet_name

        count = len(data_rows)
        product_counts[product] = product_counts.get(product, 0) + count
        total += count

        for row in data_rows:
            region = row[region_idx] or "unknown"
            region_counts[region] += 1
            product_region[product][region] += 1

    # 输出汇总
    print("=" * 50)
    print(f"扫描结果汇总：{xlsx_path}")
    print("=" * 50)
    print(f"\n总实例数：{total}")

    print(f"\n【按产品】")
    for product, count in sorted(product_counts.items(), key=lambda x: -x[1]):
        print(f"  {product:<12} {count} 个")

    print(f"\n【按地域】")
    for region, count in sorted(region_counts.items(), key=lambda x: -x[1]):
        print(f"  {region:<20} {count} 个")

    if len(product_counts) > 1:
        print(f"\n【按产品 × 地域】")
        for product, regions in sorted(product_region.items()):
            for region, count in sorted(regions.items(), key=lambda x: -x[1]):
                print(f"  {product:<12} {region:<20} {count} 个")

    print()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"用法: python3 {sys.argv[0]} <scan_result.xlsx>")
        sys.exit(1)
    summarize(sys.argv[1])
