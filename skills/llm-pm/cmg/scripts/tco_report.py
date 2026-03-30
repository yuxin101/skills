#!/usr/bin/env python3
"""
CMG TCO 分析报告生成器
用法: python3 tco_report.py <pricing_data.json> [--title "项目名称"]

从询价 JSON 数据生成：
  1. Excel 报告（询价明细 + 按厂商/产品汇总 + 成本对比）
  2. HTML 可视化报告（图表 + 表格）
"""
from __future__ import annotations

import sys
import os
import json
import argparse
from datetime import datetime
from typing import Any, TypedDict

try:
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
except ImportError:
    print("缺少依赖，请先执行: pip3 install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 类型定义
# ---------------------------------------------------------------------------

class VendorSummary(TypedDict):
    monthly: float
    yearly: float
    count: int


class ProductSummary(TypedDict):
    source_monthly: float
    target_monthly: float
    source_yearly: float
    target_yearly: float


class Summary(TypedDict):
    total_items: int
    source_items: list[dict[str, Any]]
    target_items: list[dict[str, Any]]
    by_vendor: dict[str, VendorSummary]
    by_product: dict[str, ProductSummary]
    source_total_monthly: float
    source_total_yearly: float
    target_total_monthly: float
    target_total_yearly: float
    saving_monthly: float
    saving_yearly: float
    saving_pct: float


PricingData = dict[str, Any]


# ---------------------------------------------------------------------------
# 数据加载与校验
# ---------------------------------------------------------------------------

def load_pricing_data(json_path: str) -> PricingData:
    """加载询价 JSON 数据"""
    with open(json_path, "r", encoding="utf-8") as f:
        data: PricingData = json.load(f)

    # 兼容：如果传入的是数组，包装成标准结构
    if isinstance(data, list):
        data = {
            "project_name": "云迁移TCO分析",
            "analysis_date": datetime.now().strftime("%Y-%m-%d"),
            "source_vendor": "",
            "target_vendor": "tencent",
            "currency": "CNY",
            "pricing_items": data,
        }

    items: list[dict[str, Any]] = data.get("pricing_items", [])
    if not items:
        print("错误：pricing_items 为空")
        sys.exit(1)

    # 校验必填字段
    required = ["vendor", "product", "subtotal_monthly"]
    for i, item in enumerate(items):
        for field in required:
            if field not in item or item[field] is None:
                print(f"警告：第 {i+1} 条记录缺少字段 '{field}'，已跳过")

    return data


# ---------------------------------------------------------------------------
# 统计计算
# ---------------------------------------------------------------------------

def compute_summary(data: PricingData) -> Summary:
    """计算各维度汇总"""
    items: list[dict[str, Any]] = data["pricing_items"]

    by_vendor: dict[str, VendorSummary] = {}
    by_product: dict[str, ProductSummary] = {}
    source_items: list[dict[str, Any]] = []
    target_items: list[dict[str, Any]] = []
    source_total_monthly: float = 0
    source_total_yearly: float = 0
    target_total_monthly: float = 0
    target_total_yearly: float = 0

    for item in items:
        side: str = item.get("side", "source")
        vendor: str = item.get("vendor", "unknown")
        product: str = item.get("product", "unknown")
        monthly: float = float(item.get("subtotal_monthly", 0) or 0)
        yearly: float = float(item.get("subtotal_yearly", 0) or monthly * 12)

        # 按厂商
        if vendor not in by_vendor:
            by_vendor[vendor] = {"monthly": 0, "yearly": 0, "count": 0}
        by_vendor[vendor]["monthly"] += monthly
        by_vendor[vendor]["yearly"] += yearly
        by_vendor[vendor]["count"] += 1

        # 按产品+side
        if product not in by_product:
            by_product[product] = {"source_monthly": 0, "target_monthly": 0, "source_yearly": 0, "target_yearly": 0}

        if side == "source":
            source_items.append(item)
            source_total_monthly += monthly
            source_total_yearly += yearly
            by_product[product]["source_monthly"] += monthly
            by_product[product]["source_yearly"] += yearly
        else:
            target_items.append(item)
            target_total_monthly += monthly
            target_total_yearly += yearly
            by_product[product]["target_monthly"] += monthly
            by_product[product]["target_yearly"] += yearly

    # 节省计算
    saving_monthly: float = 0
    saving_yearly: float = 0
    saving_pct: float = 0
    if source_total_monthly > 0:
        saving_monthly = source_total_monthly - target_total_monthly
        saving_yearly = saving_monthly * 12
        saving_pct = (saving_monthly / source_total_monthly) * 100

    return {
        "total_items": len(items),
        "source_items": source_items,
        "target_items": target_items,
        "by_vendor": by_vendor,
        "by_product": by_product,
        "source_total_monthly": source_total_monthly,
        "source_total_yearly": source_total_yearly,
        "target_total_monthly": target_total_monthly,
        "target_total_yearly": target_total_yearly,
        "saving_monthly": saving_monthly,
        "saving_yearly": saving_yearly,
        "saving_pct": saving_pct,
    }


# ---------------------------------------------------------------------------
# Excel 报告
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
MONEY_FORMAT = '#,##0.00'

SOURCE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TARGET_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SAVING_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _style_header(ws: Any, row: int, col_count: int) -> None:
    """给表头行加样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws: Any) -> None:
    """自动列宽"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 40)


def generate_excel(data: PricingData, summary: Summary, output_path: str) -> None:
    """生成 Excel 报告"""
    wb = openpyxl.Workbook()

    # --- Sheet 1: 询价明细 ---
    ws1: Any = wb.active
    assert ws1 is not None
    ws1.title = "询价明细"
    headers = ["序号", "类别", "厂商", "产品", "资源ID", "资源名称", "地域", "规格摘要",
               "计费方式", "月单价", "数量", "月小计", "年小计", "币种", "数据来源", "询价日期", "备注"]
    ws1.append(headers)
    _style_header(ws1, 1, len(headers))

    items: list[dict[str, Any]] = data["pricing_items"]
    for i, item in enumerate(items, 1):
        side_label = "源端" if item.get("side") == "source" else "目标端"
        monthly = float(item.get("subtotal_monthly", 0) or 0)
        yearly = float(item.get("subtotal_yearly", 0) or monthly * 12)
        row = [
            i,
            side_label,
            item.get("vendor", ""),
            item.get("product", ""),
            item.get("resource_id", ""),
            item.get("resource_name", ""),
            item.get("region", ""),
            item.get("spec_summary", ""),
            item.get("billing_mode", ""),
            float(item.get("unit_price_monthly", 0) or 0),
            int(item.get("quantity", 1) or 1),
            monthly,
            yearly,
            item.get("currency", "CNY"),
            item.get("price_source", ""),
            item.get("query_time", ""),
            item.get("notes", ""),
        ]
        ws1.append(row)
        row_num = i + 1
        # 给源端/目标端上色
        fill = SOURCE_FILL if item.get("side") == "source" else TARGET_FILL
        for col in range(1, len(headers) + 1):
            ws1.cell(row=row_num, column=col).border = THIN_BORDER
            ws1.cell(row=row_num, column=col).fill = fill
        # 金额格式
        for col in [10, 12, 13]:
            ws1.cell(row=row_num, column=col).number_format = MONEY_FORMAT

    _auto_width(ws1)

    # --- Sheet 2: 按厂商汇总 ---
    ws2 = wb.create_sheet("按厂商汇总")
    headers2 = ["厂商", "资源数", "月度总费用", "年度总费用"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    for vendor, info in sorted(summary["by_vendor"].items()):
        ws2.append([vendor, info["count"], info["monthly"], info["yearly"]])
    row_num = ws2.max_row
    for r in range(2, row_num + 1):
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=r, column=c).border = THIN_BORDER
        ws2.cell(row=r, column=3).number_format = MONEY_FORMAT
        ws2.cell(row=r, column=4).number_format = MONEY_FORMAT
    _auto_width(ws2)

    # --- Sheet 3: 按产品汇总 ---
    ws3 = wb.create_sheet("按产品汇总")
    headers3 = ["产品", "源端月费用", "目标端月费用", "源端年费用", "目标端年费用", "月节省", "节省比例"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))

    for product, info in sorted(summary["by_product"].items()):
        src_m = info["source_monthly"]
        tgt_m = info["target_monthly"]
        src_y = info["source_yearly"]
        tgt_y = info["target_yearly"]
        saving_m = src_m - tgt_m
        pct = (saving_m / src_m * 100) if src_m > 0 else 0
        ws3.append([product, src_m, tgt_m, src_y, tgt_y, saving_m, f"{pct:.1f}%"])
    row_num = ws3.max_row
    for r in range(2, row_num + 1):
        for c in range(1, len(headers3) + 1):
            ws3.cell(row=r, column=c).border = THIN_BORDER
        for c in [2, 3, 4, 5, 6]:
            ws3.cell(row=r, column=c).number_format = MONEY_FORMAT
        # 节省为正的标绿，为负的标红
        saving_cell = ws3.cell(row=r, column=6)
        try:
            if float(saving_cell.value or 0) > 0:
                saving_cell.fill = SAVING_FILL
            elif float(saving_cell.value or 0) < 0:
                saving_cell.fill = LOSS_FILL
        except (ValueError, TypeError):
            pass
    _auto_width(ws3)

    # --- Sheet 4: 成本对比 ---
    ws4 = wb.create_sheet("成本对比")
    ws4.append(["项目", "值"])
    _style_header(ws4, 1, 2)

    comparison_rows = [
        ["项目名称", data.get("project_name", "")],
        ["分析日期", data.get("analysis_date", "")],
        ["源端厂商", data.get("source_vendor", "")],
        ["目标厂商", data.get("target_vendor", "tencent")],
        ["", ""],
        ["源端资源数", len(summary["source_items"])],
        ["目标端资源数", len(summary["target_items"])],
        ["", ""],
        ["源端月度总成本", summary["source_total_monthly"]],
        ["目标端月度总成本", summary["target_total_monthly"]],
        ["月度节省金额", summary["saving_monthly"]],
        ["", ""],
        ["源端年度总成本", summary["source_total_yearly"]],
        ["目标端年度总成本", summary["target_total_yearly"]],
        ["年度节省金额", summary["saving_yearly"]],
        ["", ""],
        ["成本节省比例", f"{summary['saving_pct']:.1f}%"],
    ]
    for row in comparison_rows:
        ws4.append(row)
    for r in range(2, ws4.max_row + 1):
        for c in range(1, 3):
            ws4.cell(row=r, column=c).border = THIN_BORDER
        ws4.cell(row=r, column=1).font = Font(bold=True)
    # 金额格式
    for r in [10, 11, 12, 14, 15, 16]:
        ws4.cell(row=r, column=2).number_format = MONEY_FORMAT
    _auto_width(ws4)

    wb.save(output_path)
    print(f"✅ Excel 报告已生成: {output_path}")


# ---------------------------------------------------------------------------
# HTML 报告
# ---------------------------------------------------------------------------

def generate_html(data: PricingData, summary: Summary, output_path: str, title: str | None = None) -> None:
    """生成 HTML 可视化报告（精美版）"""
    project_name: str = title or data.get("project_name", "云迁移 TCO 分析报告")
    analysis_date: str = data.get("analysis_date", datetime.now().strftime("%Y-%m-%d"))
    currency: str = data.get("currency", "CNY")
    items: list[dict[str, Any]] = data["pricing_items"]

    # 厂商显示名映射
    vendor_display: dict[str, str] = {
        "aliyun": "阿里云", "alibaba": "阿里云",
        "huawei": "华为云", "huaweicloud": "华为云",
        "tencent": "腾讯云", "tencentcloud": "腾讯云",
        "aws": "AWS", "amazon": "AWS",
        "azure": "Azure", "microsoft": "Azure",
        "gcp": "Google Cloud", "google": "Google Cloud",
    }
    source_display: str = vendor_display.get(data.get("source_vendor", "").lower()) or data.get("source_vendor", "多云") or "多云"
    target_display: str = vendor_display.get(data.get("target_vendor", "").lower()) or data.get("target_vendor", "腾讯云") or "腾讯云"

    # 按产品汇总数据（供图表用）
    product_labels: list[str] = []
    source_values: list[float] = []
    target_values: list[float] = []
    saving_values: list[float] = []
    for product, info in sorted(summary["by_product"].items()):
        product_labels.append(product)
        source_values.append(round(info["source_monthly"], 2))
        target_values.append(round(info["target_monthly"], 2))
        saving_values.append(round(info["source_monthly"] - info["target_monthly"], 2))

    # 按厂商汇总（供饼图用）
    vendor_labels: list[str] = []
    vendor_values: list[float] = []
    for vendor, info in sorted(summary["by_vendor"].items()):
        vendor_labels.append(vendor_display.get(vendor.lower(), vendor))
        vendor_values.append(round(info["monthly"], 2))

    # 明细表行
    detail_rows_html = ""
    for i, item in enumerate(items, 1):
        side = item.get("side", "source")
        side_label = "源端" if side == "source" else "目标端"
        monthly = float(item.get("subtotal_monthly", 0) or 0)
        yearly = float(item.get("subtotal_yearly", 0) or monthly * 12)
        v = item.get("vendor", "")
        v_display = vendor_display.get(v.lower(), v)
        detail_rows_html += f"""
                        <tr class="detail-row {'source-row' if side == 'source' else 'target-row'}">
                            <td class="cell-center">{i}</td>
                            <td class="cell-center"><span class="badge {'badge-source' if side == 'source' else 'badge-target'}">{side_label}</span></td>
                            <td>{v_display}</td>
                            <td><strong>{item.get('product', '')}</strong></td>
                            <td>{item.get('resource_name', '')}</td>
                            <td class="cell-mono">{item.get('region', '')}</td>
                            <td class="cell-spec">{item.get('spec_summary', '')}</td>
                            <td class="cell-money">{monthly:,.2f}</td>
                            <td class="cell-money">{yearly:,.2f}</td>
                            <td class="cell-note">{item.get('notes', '')}</td>
                        </tr>"""

    # 按产品对比表
    product_rows_html = ""
    for _, (product, info) in enumerate(sorted(summary["by_product"].items())):
        src_m = info["source_monthly"]
        tgt_m = info["target_monthly"]
        saving = src_m - tgt_m
        pct = (saving / src_m * 100) if src_m > 0 else 0
        saving_cls = "val-positive" if saving >= 0 else "val-negative"
        arrow = "↓" if saving >= 0 else "↑"
        # 节省率进度条
        bar_pct = min(abs(pct), 100)
        bar_color = "#10b981" if saving >= 0 else "#ef4444"
        product_rows_html += f"""
                        <tr>
                            <td><strong>{product}</strong></td>
                            <td class="cell-money">{src_m:,.2f}</td>
                            <td class="cell-money">{tgt_m:,.2f}</td>
                            <td class="cell-money {saving_cls}">{arrow} {abs(saving):,.2f}</td>
                            <td>
                                <div class="pct-bar-wrap">
                                    <div class="pct-bar" style="width:{bar_pct}%;background:{bar_color};"></div>
                                    <span class="{saving_cls}">{pct:+.1f}%</span>
                                </div>
                            </td>
                        </tr>"""

    saving_class = "val-positive" if summary["saving_pct"] >= 0 else "val-negative"
    saving_direction = "节省" if summary["saving_pct"] >= 0 else "增加"
    hero_gradient = "linear-gradient(135deg, #10b981 0%, #059669 100%)" if summary["saving_pct"] >= 0 else "linear-gradient(135deg, #ef4444 0%, #dc2626 100%)"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{project_name}</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
    <style>
        /* ===== CSS Variables ===== */
        :root {{
            --c-bg: #0f172a;
            --c-bg-card: #1e293b;
            --c-bg-card-hover: #273548;
            --c-bg-table-head: #334155;
            --c-bg-table-odd: rgba(255,255,255,0.02);
            --c-bg-source: rgba(251,191,36,0.08);
            --c-bg-target: rgba(59,130,246,0.08);
            --c-border: rgba(255,255,255,0.06);
            --c-border-strong: rgba(255,255,255,0.12);
            --c-text: #e2e8f0;
            --c-text-dim: #94a3b8;
            --c-text-muted: #64748b;
            --c-accent: #3b82f6;
            --c-accent-light: #60a5fa;
            --c-source: #fbbf24;
            --c-source-dim: rgba(251,191,36,0.2);
            --c-target: #3b82f6;
            --c-target-dim: rgba(59,130,246,0.2);
            --c-positive: #10b981;
            --c-negative: #ef4444;
            --c-purple: #8b5cf6;
            --radius: 16px;
            --radius-sm: 10px;
            --radius-xs: 6px;
            --shadow: 0 4px 24px rgba(0,0,0,0.25);
            --shadow-lg: 0 8px 40px rgba(0,0,0,0.35);
            --font-sans: 'Inter', -apple-system, BlinkMacSystemFont, 'PingFang SC', 'Microsoft YaHei', sans-serif;
            --font-mono: 'JetBrains Mono', 'SF Mono', monospace;
            --transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        /* ===== Reset & Base ===== */
        *, *::before, *::after {{ margin:0; padding:0; box-sizing:border-box; }}
        html {{ scroll-behavior: smooth; }}
        body {{
            font-family: var(--font-sans);
            background: var(--c-bg);
            color: var(--c-text);
            line-height: 1.7;
            min-height: 100vh;
            -webkit-font-smoothing: antialiased;
        }}

        /* Background decoration */
        body::before {{
            content: '';
            position: fixed;
            top: -50%; left: -50%;
            width: 200%; height: 200%;
            background: radial-gradient(ellipse at 20% 50%, rgba(59,130,246,0.08) 0%, transparent 50%),
                        radial-gradient(ellipse at 80% 20%, rgba(139,92,246,0.06) 0%, transparent 50%),
                        radial-gradient(ellipse at 50% 80%, rgba(16,185,129,0.05) 0%, transparent 50%);
            pointer-events: none;
            z-index: 0;
        }}

        .page-wrap {{
            position: relative;
            z-index: 1;
            max-width: 1360px;
            margin: 0 auto;
            padding: 32px 24px 48px;
        }}

        /* ===== Header ===== */
        .hero {{
            position: relative;
            background: linear-gradient(135deg, #1e3a5f 0%, #0f172a 40%, #1e293b 100%);
            border: 1px solid var(--c-border-strong);
            border-radius: var(--radius);
            padding: 48px 48px 40px;
            margin-bottom: 28px;
            overflow: hidden;
            box-shadow: var(--shadow-lg);
        }}
        .hero::before {{
            content: '';
            position: absolute;
            top: 0; right: 0;
            width: 400px; height: 400px;
            background: radial-gradient(circle, rgba(59,130,246,0.15) 0%, transparent 70%);
            transform: translate(30%, -30%);
        }}
        .hero::after {{
            content: '';
            position: absolute;
            bottom: 0; left: 0;
            width: 300px; height: 300px;
            background: radial-gradient(circle, rgba(139,92,246,0.1) 0%, transparent 70%);
            transform: translate(-30%, 30%);
        }}
        .hero-content {{ position: relative; z-index: 1; }}
        .hero-label {{
            display: inline-flex;
            align-items: center;
            gap: 6px;
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1.5px;
            color: var(--c-accent-light);
            margin-bottom: 12px;
        }}
        .hero-label::before {{
            content: '';
            display: inline-block;
            width: 8px; height: 8px;
            border-radius: 50%;
            background: var(--c-accent);
            animation: pulse-dot 2s infinite;
        }}
        @keyframes pulse-dot {{
            0%, 100% {{ opacity: 1; transform: scale(1); }}
            50% {{ opacity: 0.5; transform: scale(1.3); }}
        }}
        .hero h1 {{
            font-size: 32px;
            font-weight: 800;
            letter-spacing: -0.5px;
            background: linear-gradient(135deg, #fff 0%, #94a3b8 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 16px;
        }}
        .hero-meta {{
            display: flex;
            flex-wrap: wrap;
            gap: 20px;
            font-size: 13px;
            color: var(--c-text-dim);
        }}
        .hero-meta-item {{
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        .hero-meta-item .dot {{
            width: 4px; height: 4px;
            border-radius: 50%;
            background: var(--c-text-muted);
        }}
        .hero-flow {{
            display: inline-flex;
            align-items: center;
            gap: 10px;
            margin-top: 20px;
            padding: 10px 20px;
            background: rgba(255,255,255,0.05);
            border-radius: var(--radius-sm);
            border: 1px solid var(--c-border);
        }}
        .hero-flow .vendor-tag {{
            padding: 4px 14px;
            border-radius: 20px;
            font-size: 13px;
            font-weight: 600;
        }}
        .vendor-tag.src {{ background: var(--c-source-dim); color: var(--c-source); }}
        .vendor-tag.tgt {{ background: var(--c-target-dim); color: var(--c-accent-light); }}
        .hero-flow .arrow {{
            color: var(--c-text-muted);
            font-size: 18px;
        }}

        /* ===== KPI ===== */
        .kpi-strip {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 16px;
            margin-bottom: 28px;
        }}
        @media (max-width: 900px) {{
            .kpi-strip {{ grid-template-columns: repeat(2, 1fr); }}
        }}
        @media (max-width: 520px) {{
            .kpi-strip {{ grid-template-columns: 1fr; }}
        }}
        .kpi {{
            position: relative;
            background: var(--c-bg-card);
            border: 1px solid var(--c-border);
            border-radius: var(--radius);
            padding: 24px 20px;
            overflow: hidden;
            transition: var(--transition);
        }}
        .kpi:hover {{
            border-color: var(--c-border-strong);
            transform: translateY(-2px);
            box-shadow: var(--shadow);
        }}
        .kpi::before {{
            content: '';
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 3px;
        }}
        .kpi.kpi-source::before {{ background: linear-gradient(90deg, var(--c-source), #f59e0b); }}
        .kpi.kpi-target::before {{ background: linear-gradient(90deg, var(--c-target), #818cf8); }}
        .kpi.kpi-saving-pos::before {{ background: linear-gradient(90deg, #10b981, #34d399); }}
        .kpi.kpi-saving-neg::before {{ background: linear-gradient(90deg, #ef4444, #f87171); }}
        .kpi-icon {{
            width: 40px; height: 40px;
            border-radius: var(--radius-xs);
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 18px;
            margin-bottom: 14px;
        }}
        .kpi.kpi-source .kpi-icon {{ background: var(--c-source-dim); }}
        .kpi.kpi-target .kpi-icon {{ background: var(--c-target-dim); }}
        .kpi.kpi-saving-pos .kpi-icon {{ background: rgba(16,185,129,0.15); }}
        .kpi.kpi-saving-neg .kpi-icon {{ background: rgba(239,68,68,0.15); }}
        .kpi-label {{
            font-size: 12px;
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            color: var(--c-text-muted);
            margin-bottom: 6px;
        }}
        .kpi-value {{
            font-size: 26px;
            font-weight: 700;
            font-family: var(--font-mono);
            letter-spacing: -0.5px;
            line-height: 1.2;
        }}
        .kpi.kpi-source .kpi-value {{ color: var(--c-source); }}
        .kpi.kpi-target .kpi-value {{ color: var(--c-accent-light); }}
        .kpi.kpi-saving-pos .kpi-value {{ color: var(--c-positive); }}
        .kpi.kpi-saving-neg .kpi-value {{ color: var(--c-negative); }}
        .kpi-sub {{
            font-size: 12px;
            color: var(--c-text-muted);
            margin-top: 6px;
            font-family: var(--font-mono);
        }}

        /* ===== Saving Hero Banner ===== */
        .saving-hero {{
            background: {hero_gradient};
            border-radius: var(--radius);
            padding: 32px 40px;
            margin-bottom: 28px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 24px;
            box-shadow: var(--shadow);
        }}
        .saving-hero-left {{ flex: 1; }}
        .saving-hero-left h3 {{
            font-size: 14px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: rgba(255,255,255,0.7);
            margin-bottom: 6px;
        }}
        .saving-hero-left .big-number {{
            font-size: 42px;
            font-weight: 800;
            color: #fff;
            font-family: var(--font-mono);
            letter-spacing: -1px;
        }}
        .saving-hero-left .desc {{
            font-size: 14px;
            color: rgba(255,255,255,0.75);
            margin-top: 4px;
        }}
        .saving-hero-right {{
            text-align: center;
            padding: 20px 32px;
            background: rgba(255,255,255,0.15);
            border-radius: var(--radius-sm);
            backdrop-filter: blur(8px);
        }}
        .saving-hero-right .pct {{
            font-size: 48px;
            font-weight: 800;
            color: #fff;
            font-family: var(--font-mono);
            line-height: 1;
        }}
        .saving-hero-right .pct-label {{
            font-size: 13px;
            color: rgba(255,255,255,0.7);
            margin-top: 4px;
        }}
        @media (max-width: 640px) {{
            .saving-hero {{ flex-direction: column; text-align: center; padding: 24px; }}
            .saving-hero-left .big-number {{ font-size: 32px; }}
            .saving-hero-right .pct {{ font-size: 36px; }}
        }}

        /* ===== Card ===== */
        .card {{
            background: var(--c-bg-card);
            border: 1px solid var(--c-border);
            border-radius: var(--radius);
            margin-bottom: 24px;
            overflow: hidden;
            transition: var(--transition);
        }}
        .card:hover {{
            border-color: var(--c-border-strong);
            box-shadow: var(--shadow);
        }}
        .card-header {{
            padding: 20px 24px 16px;
            border-bottom: 1px solid var(--c-border);
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        .card-header .icon {{
            width: 32px; height: 32px;
            border-radius: var(--radius-xs);
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 16px;
            background: rgba(59,130,246,0.1);
        }}
        .card-header h2 {{
            font-size: 16px;
            font-weight: 700;
            color: var(--c-text);
            letter-spacing: -0.2px;
        }}
        .card-body {{
            padding: 24px;
        }}

        /* ===== Charts ===== */
        .chart-row {{
            display: grid;
            grid-template-columns: 3fr 2fr;
            gap: 24px;
            margin-bottom: 24px;
        }}
        @media (max-width: 900px) {{
            .chart-row {{ grid-template-columns: 1fr; }}
        }}
        .chart-container {{
            position: relative;
            height: 380px;
        }}

        /* ===== Table ===== */
        .table-wrap {{
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        thead th {{
            background: var(--c-bg-table-head);
            color: var(--c-text-dim);
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            padding: 12px 16px;
            text-align: left;
            white-space: nowrap;
            border-bottom: 1px solid var(--c-border-strong);
            position: sticky;
            top: 0;
            z-index: 2;
        }}
        td {{
            padding: 10px 16px;
            border-bottom: 1px solid var(--c-border);
            vertical-align: middle;
        }}
        tbody tr {{ transition: background 0.15s; }}
        tbody tr:hover {{ background: rgba(255,255,255,0.03); }}
        .detail-row.source-row {{ background: var(--c-bg-source); }}
        .detail-row.target-row {{ background: var(--c-bg-target); }}
        .detail-row.source-row:hover {{ background: rgba(251,191,36,0.12); }}
        .detail-row.target-row:hover {{ background: rgba(59,130,246,0.12); }}
        .cell-center {{ text-align: center; }}
        .cell-money {{
            text-align: right;
            font-family: var(--font-mono);
            font-weight: 500;
            font-size: 13px;
            font-variant-numeric: tabular-nums;
        }}
        .cell-mono {{ font-family: var(--font-mono); font-size: 12px; color: var(--c-text-dim); }}
        .cell-spec {{ font-size: 12px; color: var(--c-text-dim); max-width: 220px; }}
        .cell-note {{ font-size: 12px; color: var(--c-text-muted); max-width: 120px; }}
        .val-positive {{ color: var(--c-positive); font-weight: 600; }}
        .val-negative {{ color: var(--c-negative); font-weight: 600; }}

        /* Badge */
        .badge {{
            display: inline-flex;
            align-items: center;
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
            letter-spacing: 0.3px;
        }}
        .badge-source {{ background: var(--c-source-dim); color: var(--c-source); }}
        .badge-target {{ background: var(--c-target-dim); color: var(--c-accent-light); }}

        /* Percentage bar */
        .pct-bar-wrap {{
            display: flex;
            align-items: center;
            gap: 10px;
            min-width: 140px;
        }}
        .pct-bar {{
            height: 6px;
            border-radius: 3px;
            transition: width 0.6s ease;
        }}
        .pct-bar-wrap span {{
            font-family: var(--font-mono);
            font-size: 12px;
            font-weight: 600;
            min-width: 52px;
            text-align: right;
        }}

        /* Totals row */
        .row-total {{
            background: rgba(255,255,255,0.04) !important;
        }}
        .row-total td {{
            font-weight: 700;
            border-top: 2px solid var(--c-border-strong);
            padding-top: 12px;
            padding-bottom: 12px;
        }}

        /* ===== Footer ===== */
        .report-footer {{
            text-align: center;
            padding: 32px 20px;
            color: var(--c-text-muted);
            font-size: 12px;
            border-top: 1px solid var(--c-border);
            margin-top: 16px;
        }}
        .report-footer p {{ margin-bottom: 4px; }}
        .report-footer .warn {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
            color: var(--c-source);
            font-weight: 500;
        }}

        /* ===== Animation on scroll ===== */
        .fade-in {{
            opacity: 0;
            transform: translateY(20px);
            transition: opacity 0.5s ease, transform 0.5s ease;
        }}
        .fade-in.visible {{
            opacity: 1;
            transform: translateY(0);
        }}

        /* ===== Print ===== */
        @media print {{
            body {{ background: #fff; color: #1e293b; }}
            body::before {{ display: none; }}
            .card, .kpi, .hero, .saving-hero {{ box-shadow: none; border: 1px solid #e2e8f0; }}
            .hero {{ background: #1e3a5f !important; }}
            .saving-hero {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
            thead th {{ background: #334155 !important; color: #fff !important; print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
            .fade-in {{ opacity: 1 !important; transform: none !important; }}
        }}
    </style>
</head>
<body>
    <div class="page-wrap">

        <!-- ===== Hero Header ===== -->
        <div class="hero fade-in">
            <div class="hero-content">
                <div class="hero-label">CMG Cloud Migration &mdash; TCO Analysis</div>
                <h1>{project_name}</h1>
                <div class="hero-meta">
                    <div class="hero-meta-item">📅 {analysis_date}</div>
                    <div class="hero-meta-item"><span class="dot"></span></div>
                    <div class="hero-meta-item">💰 {currency}</div>
                    <div class="hero-meta-item"><span class="dot"></span></div>
                    <div class="hero-meta-item">📦 {summary['total_items']} 项资源</div>
                </div>
                <div class="hero-flow">
                    <span class="vendor-tag src">{source_display}</span>
                    <span class="arrow">→</span>
                    <span class="vendor-tag tgt">{target_display}</span>
                </div>
            </div>
        </div>

        <!-- ===== Saving Hero Banner ===== -->
        <div class="saving-hero fade-in">
            <div class="saving-hero-left">
                <h3>迁移成本{saving_direction}</h3>
                <div class="big-number">{currency} {abs(summary['saving_yearly']):,.0f}<span style="font-size:20px;opacity:0.7">/年</span></div>
                <div class="desc">月度{saving_direction} {currency} {abs(summary['saving_monthly']):,.2f}，相比源端成本</div>
            </div>
            <div class="saving-hero-right">
                <div class="pct">{abs(summary['saving_pct']):.1f}%</div>
                <div class="pct-label">成本{saving_direction}比例</div>
            </div>
        </div>

        <!-- ===== KPI Strip ===== -->
        <div class="kpi-strip">
            <div class="kpi kpi-source fade-in">
                <div class="kpi-icon">☁️</div>
                <div class="kpi-label">源端月度成本</div>
                <div class="kpi-value">{summary['source_total_monthly']:,.0f}</div>
                <div class="kpi-sub">年度 {currency} {summary['source_total_yearly']:,.0f}</div>
            </div>
            <div class="kpi kpi-target fade-in">
                <div class="kpi-icon">🚀</div>
                <div class="kpi-label">目标端月度成本</div>
                <div class="kpi-value">{summary['target_total_monthly']:,.0f}</div>
                <div class="kpi-sub">年度 {currency} {summary['target_total_yearly']:,.0f}</div>
            </div>
            <div class="kpi kpi-saving-{'pos' if summary['saving_pct'] >= 0 else 'neg'} fade-in">
                <div class="kpi-icon">{'💰' if summary['saving_pct'] >= 0 else '⚠️'}</div>
                <div class="kpi-label">月度{saving_direction}</div>
                <div class="kpi-value">{abs(summary['saving_monthly']):,.0f}</div>
                <div class="kpi-sub">{currency}/月</div>
            </div>
            <div class="kpi kpi-saving-{'pos' if summary['saving_pct'] >= 0 else 'neg'} fade-in">
                <div class="kpi-icon">📊</div>
                <div class="kpi-label">{saving_direction}比例</div>
                <div class="kpi-value">{abs(summary['saving_pct']):.1f}%</div>
                <div class="kpi-sub">迁移后预计{saving_direction}</div>
            </div>
        </div>

        <!-- ===== Charts ===== -->
        <div class="chart-row">
            <div class="card fade-in">
                <div class="card-header">
                    <div class="icon">📊</div>
                    <h2>按产品成本对比（月度）</h2>
                </div>
                <div class="card-body">
                    <div class="chart-container">
                        <canvas id="productChart"></canvas>
                    </div>
                </div>
            </div>
            <div class="card fade-in">
                <div class="card-header">
                    <div class="icon">🧩</div>
                    <h2>厂商费用占比</h2>
                </div>
                <div class="card-body">
                    <div class="chart-container">
                        <canvas id="vendorChart"></canvas>
                    </div>
                </div>
            </div>
        </div>

        <!-- ===== Product Comparison Table ===== -->
        <div class="card fade-in">
            <div class="card-header">
                <div class="icon">📋</div>
                <h2>按产品成本对比</h2>
            </div>
            <div class="table-wrap">
                <table>
                    <thead>
                        <tr>
                            <th>产品类型</th>
                            <th style="text-align:right">源端月费用</th>
                            <th style="text-align:right">目标端月费用</th>
                            <th style="text-align:right">月度差异</th>
                            <th>节省比例</th>
                        </tr>
                    </thead>
                    <tbody>
                        {product_rows_html}
                        <tr class="row-total">
                            <td>合 计</td>
                            <td class="cell-money">{summary['source_total_monthly']:,.2f}</td>
                            <td class="cell-money">{summary['target_total_monthly']:,.2f}</td>
                            <td class="cell-money {saving_class}">{'↓' if summary['saving_pct'] >= 0 else '↑'} {abs(summary['saving_monthly']):,.2f}</td>
                            <td>
                                <div class="pct-bar-wrap">
                                    <div class="pct-bar" style="width:{min(abs(summary['saving_pct']), 100)}%;background:{'#10b981' if summary['saving_pct'] >= 0 else '#ef4444'};"></div>
                                    <span class="{saving_class}">{summary['saving_pct']:+.1f}%</span>
                                </div>
                            </td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>

        <!-- ===== Detail Table ===== -->
        <div class="card fade-in">
            <div class="card-header">
                <div class="icon">📑</div>
                <h2>询价明细</h2>
            </div>
            <div class="table-wrap">
                <table>
                    <thead>
                        <tr>
                            <th style="text-align:center">#</th>
                            <th style="text-align:center">类别</th>
                            <th>厂商</th>
                            <th>产品</th>
                            <th>资源名称</th>
                            <th>地域</th>
                            <th>规格</th>
                            <th style="text-align:right">月费用</th>
                            <th style="text-align:right">年费用</th>
                            <th>备注</th>
                        </tr>
                    </thead>
                    <tbody>
                        {detail_rows_html}
                    </tbody>
                </table>
            </div>
        </div>

        <!-- ===== Footer ===== -->
        <div class="report-footer fade-in">
            <p>本报告由 <strong>CMG 云迁移 Skill</strong> 自动生成 &nbsp;|&nbsp; 价格数据来源于各云厂商官方价格计算器</p>
            <p class="warn">⚠️ 实际费用可能因折扣、促销活动、使用量波动等因素与报告存在差异，请以云厂商实际账单为准</p>
            <p style="margin-top:8px;">生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </div>
    </div>

    <!-- ===== Charts Script ===== -->
    <script>
        // --- Scroll fade-in animation ---
        const observer = new IntersectionObserver((entries) => {{
            entries.forEach(e => {{ if (e.isIntersecting) {{ e.target.classList.add('visible'); }} }});
        }}, {{ threshold: 0.08 }});
        document.querySelectorAll('.fade-in').forEach(el => observer.observe(el));

        // --- Chart.js global defaults (dark theme) ---
        Chart.defaults.color = '#94a3b8';
        Chart.defaults.borderColor = 'rgba(255,255,255,0.06)';
        Chart.defaults.font.family = "'Inter', sans-serif";

        // --- Color palette ---
        const C = {{
            source:       'rgba(251, 191, 36, 0.85)',
            sourceBorder: 'rgba(251, 191, 36, 1)',
            sourceGlow:   'rgba(251, 191, 36, 0.15)',
            target:       'rgba(59, 130, 246, 0.85)',
            targetBorder: 'rgba(59, 130, 246, 1)',
            targetGlow:   'rgba(59, 130, 246, 0.15)',
            pie: [
                'rgba(59,130,246,0.85)', 'rgba(251,191,36,0.85)',
                'rgba(16,185,129,0.85)', 'rgba(139,92,246,0.85)',
                'rgba(236,72,153,0.85)', 'rgba(245,158,11,0.85)',
                'rgba(6,182,212,0.85)', 'rgba(244,63,94,0.85)',
            ],
            pieBorder: [
                '#3b82f6','#fbbf24','#10b981','#8b5cf6',
                '#ec4899','#f59e0b','#06b6d4','#f43f5e',
            ],
        }};

        // --- Product comparison bar chart ---
        const productCtx = document.getElementById('productChart');
        new Chart(productCtx, {{
            type: 'bar',
            data: {{
                labels: {json.dumps(product_labels, ensure_ascii=False)},
                datasets: [
                    {{
                        label: '{source_display}（源端）',
                        data: {json.dumps(source_values)},
                        backgroundColor: C.source,
                        borderColor: C.sourceBorder,
                        borderWidth: 1,
                        borderRadius: 6,
                        borderSkipped: false,
                    }},
                    {{
                        label: '{target_display}（目标端）',
                        data: {json.dumps(target_values)},
                        backgroundColor: C.target,
                        borderColor: C.targetBorder,
                        borderWidth: 1,
                        borderRadius: 6,
                        borderSkipped: false,
                    }},
                ],
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{
                        position: 'top',
                        labels: {{
                            usePointStyle: true,
                            pointStyle: 'rectRounded',
                            padding: 20,
                            font: {{ size: 12, weight: '500' }},
                        }},
                    }},
                    tooltip: {{
                        backgroundColor: 'rgba(30,41,59,0.95)',
                        borderColor: 'rgba(255,255,255,0.1)',
                        borderWidth: 1,
                        cornerRadius: 8,
                        padding: 12,
                        titleFont: {{ size: 13, weight: '600' }},
                        bodyFont: {{ size: 12 }},
                        callbacks: {{
                            label: function(ctx) {{
                                return ctx.dataset.label + ': {currency} ' + ctx.raw.toLocaleString();
                            }},
                        }},
                    }},
                }},
                scales: {{
                    x: {{
                        grid: {{ display: false }},
                        ticks: {{ font: {{ size: 12 }} }},
                    }},
                    y: {{
                        beginAtZero: true,
                        grid: {{ color: 'rgba(255,255,255,0.04)' }},
                        ticks: {{
                            font: {{ size: 11 }},
                            callback: function(v) {{ return '{currency} ' + v.toLocaleString(); }},
                        }},
                    }},
                }},
            }},
        }});

        // --- Vendor doughnut chart ---
        new Chart(document.getElementById('vendorChart'), {{
            type: 'doughnut',
            data: {{
                labels: {json.dumps(vendor_labels, ensure_ascii=False)},
                datasets: [{{
                    data: {json.dumps(vendor_values)},
                    backgroundColor: C.pie,
                    borderColor: C.pieBorder,
                    borderWidth: 2,
                    hoverOffset: 8,
                }}],
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                cutout: '60%',
                plugins: {{
                    legend: {{
                        position: 'bottom',
                        labels: {{
                            usePointStyle: true,
                            pointStyle: 'circle',
                            padding: 16,
                            font: {{ size: 12, weight: '500' }},
                        }},
                    }},
                    tooltip: {{
                        backgroundColor: 'rgba(30,41,59,0.95)',
                        borderColor: 'rgba(255,255,255,0.1)',
                        borderWidth: 1,
                        cornerRadius: 8,
                        padding: 12,
                        callbacks: {{
                            label: function(ctx) {{
                                const total = ctx.dataset.data.reduce((a, b) => a + b, 0);
                                const pct = ((ctx.raw / total) * 100).toFixed(1);
                                return ctx.label + ': {currency} ' + ctx.raw.toLocaleString() + ' (' + pct + '%)';
                            }},
                        }},
                    }},
                }},
            }},
        }});
    </script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ HTML 报告已生成: {output_path}")


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="CMG TCO 分析报告生成器 — 从询价 JSON 生成 Excel + HTML 报告"
    )
    parser.add_argument("json_file", help="询价数据 JSON 文件路径")
    parser.add_argument("--title", default=None, help="报告标题（默认使用 JSON 中的 project_name）")
    parser.add_argument("--output-dir", default=".", help="输出目录（默认当前目录）")
    args = parser.parse_args()

    json_file: str = args.json_file
    if not os.path.isfile(json_file):
        print(f"错误：文件不存在 {json_file}")
        sys.exit(1)

    print(f"正在加载询价数据: {json_file}")
    data: PricingData = load_pricing_data(json_file)
    items: list[dict[str, Any]] = data["pricing_items"]
    print(f"共 {len(items)} 条询价记录")

    # 计算汇总
    summary: Summary = compute_summary(data)
    print(f"源端资源: {len(summary['source_items'])} 个，月费用: {summary['source_total_monthly']:,.2f}")
    print(f"目标端资源: {len(summary['target_items'])} 个，月费用: {summary['target_total_monthly']:,.2f}")
    if summary["saving_pct"] >= 0:
        print(f"预计节省: {summary['saving_pct']:.1f}%")
    else:
        print(f"预计增加: {abs(summary['saving_pct']):.1f}%")

    # 生成报告
    date_str: str = datetime.now().strftime("%Y%m%d")
    output_dir: str = args.output_dir
    os.makedirs(output_dir, exist_ok=True)

    excel_path: str = os.path.join(output_dir, f"tco_report_{date_str}.xlsx")
    html_path: str = os.path.join(output_dir, f"tco_report_{date_str}.html")

    generate_excel(data, summary, excel_path)
    generate_html(data, summary, html_path, title=args.title)

    print("\n🎉 TCO 分析报告生成完成！")
    print(f"   Excel: {excel_path}")
    print(f"   HTML:  {html_path}")


if __name__ == "__main__":
    main()
