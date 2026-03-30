#!/usr/bin/env python3
"""
CMG 依赖拓扑分析脚本
用法: python3 dependency.py <scan_result.xlsx> [--format text|json|mermaid] [--ignore-sg SG_ID,...]

从扫描结果 Excel 中提取资源间依赖关系，生成迁移分组建议。
"""
import sys
import os
import json
import re
import argparse
from collections import defaultdict
from datetime import datetime, timezone

try:
    import openpyxl  # pyright: ignore[reportMissingModuleSource]
except ImportError:
    print("缺少依赖，请先执行: pip3 install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

# 数据库默认端口 → 类型映射
DB_PORT_MAP = {
    3306: "MySQL",
    5432: "PostgreSQL",
    1433: "SQLServer",
    6379: "Redis",
    27017: "MongoDB",
    9092: "Kafka",
    9200: "Elasticsearch",
    9300: "Elasticsearch",
}

# 关系强度
STRENGTH_STRONG = "strong"
STRENGTH_MEDIUM = "medium"
STRENGTH_WEAK = "weak"

# 关系类型
REL_VPC = "vpc_binding"
REL_SUBNET = "subnet_binding"
REL_SG = "sg_binding"
REL_LB = "lb_binding"
REL_DB = "db_binding"
REL_STORAGE = "storage_binding"

RELATION_META = {
    REL_VPC:     {"strength": STRENGTH_WEAK,   "label": "同 VPC"},
    REL_SUBNET:  {"strength": STRENGTH_WEAK,   "label": "同子网"},
    REL_SG:      {"strength": STRENGTH_MEDIUM, "label": "共享安全组"},
    REL_LB:      {"strength": STRENGTH_STRONG, "label": "负载均衡后端"},
    REL_DB:      {"strength": STRENGTH_STRONG, "label": "数据库连接"},
    REL_STORAGE: {"strength": STRENGTH_STRONG, "label": "存储挂载"},
}


# ---------------------------------------------------------------------------
# 平台自动检测
# ---------------------------------------------------------------------------

def detect_platform(sheet_names):
    """根据 Sheet 名称前缀判断源端平台"""
    for name in sheet_names:
        upper = name.upper()
        if upper.startswith("ACS_"):
            return "aliyun"
        if upper.startswith("AWS_"):
            return "aws"
        if upper.startswith("HW_"):
            return "huaweicloud"
    return "unknown"


# ---------------------------------------------------------------------------
# Excel 解析工具
# ---------------------------------------------------------------------------

def parse_sheet(ws):
    """将一个 Sheet 解析为 list[dict]，key 为表头"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        record = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else None
        result.append(record)
    return result


def find_column(record, *candidates):
    """从 record 中按候选 key 查找第一个非空值"""
    for key in candidates:
        for rk in record:
            if rk and key.lower() in rk.lower():
                val = record[rk]
                if val is not None and str(val).strip():
                    return str(val).strip()
    return None


def parse_list_field(value):
    """把可能是 JSON 数组或逗号分隔的字段解析为 list"""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    # 尝试 JSON 数组
    if s.startswith("["):
        try:
            return json.loads(s)
        except json.JSONDecodeError:
            pass
    # 逗号分隔
    return [x.strip() for x in s.split(",") if x.strip()]


# ---------------------------------------------------------------------------
# 资源提取
# ---------------------------------------------------------------------------

class Resource:
    def __init__(self, rid, rtype, name, vpc_id, subnet_id, security_groups, region, raw=None):
        self.id = rid
        self.type = rtype
        self.name = name or rid
        self.vpc_id = vpc_id or ""
        self.subnet_id = subnet_id or ""
        self.security_groups = security_groups or []
        self.region = region or "unknown"
        self.group_id = None
        self.raw = raw or {}

    def to_dict(self):
        return {
            "id": self.id,
            "type": self.type,
            "name": self.name,
            "vpc_id": self.vpc_id,
            "subnet_id": self.subnet_id,
            "security_groups": self.security_groups,
            "region": self.region,
            "group_id": self.group_id,
        }


def extract_resources(wb, platform):
    """从工作簿中提取所有资源，返回 {resource_id: Resource}"""
    resources = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        records = parse_sheet(ws)
        if not records:
            continue

        rtype = _guess_resource_type(sheet_name, platform)
        if rtype is None:
            continue

        for rec in records:
            rid = _extract_id(rec, rtype, platform)
            if not rid:
                continue

            name = find_column(rec, "InstanceName", "Name", "名称", "实例名称", "TagValue")
            vpc_id = find_column(rec, "VpcId", "Vpc")
            subnet_id = find_column(rec, "VSwitchId", "SubnetId", "Subnet")
            region = find_column(rec, "Region", "地域", "RegionId", "AvailabilityZone")
            sgs = _extract_security_groups(rec, platform)

            resources[rid] = Resource(
                rid=rid,
                rtype=rtype,
                name=name,
                vpc_id=vpc_id,
                subnet_id=subnet_id,
                security_groups=sgs,
                region=region,
                raw=rec,
            )

    return resources


# 资源类型推断表
_TYPE_KEYWORDS = {
    "ecs": "ecs", "ec2": "ecs", "instance": "ecs", "server": "ecs",
    "rds": "rds", "dbinstance": "rds", "database": "rds",
    "redis": "redis", "elasticache": "redis", "dcs": "redis",
    "mongodb": "mongodb", "dds": "mongodb",
    "slb": "slb", "elb": "slb", "loadbalancer": "slb", "nlb": "slb", "alb": "slb",
    "vpc": "vpc",
    "vswitch": "subnet", "subnet": "subnet",
    "securitygroup": "sg",
    "oss": "oss", "s3": "oss", "obs": "oss",
    "ebs": "disk", "evs": "disk", "disk": "disk",
    "nas": "nas", "efs": "nas",
    "kafka": "kafka", "msk": "kafka",
    "es": "es", "elasticsearch": "es",
    "nat": "nat",
    "eip": "eip",
    "cdn": "cdn",
    "polardb": "polardb",
}


def _guess_resource_type(sheet_name, platform):
    """根据 Sheet 名称推断资源类型"""
    lower = sheet_name.lower().replace("_", " ").replace("-", " ")
    for keyword, rtype in _TYPE_KEYWORDS.items():
        if keyword in lower:
            return rtype
    return None


def _extract_id(rec, rtype, platform):
    """提取资源 ID"""
    candidates = [
        "InstanceId", "Id", "ResourceId", "DBInstanceId",
        "LoadBalancerId", "BucketName", "DiskId", "VolumeId",
        "VpcId", "VSwitchId", "SubnetId", "SecurityGroupId",
        "FileSystemId", "实例ID", "CacheClusterId", "ClusterId",
    ]
    for c in candidates:
        val = find_column(rec, c)
        if val:
            return val
    return None


def _extract_security_groups(rec, platform):
    """提取安全组列表"""
    sg_val = find_column(rec, "SecurityGroupIds", "SecurityGroupId", "SecurityGroups", "安全组")
    return parse_list_field(sg_val)


# ---------------------------------------------------------------------------
# 依赖关系分析
# ---------------------------------------------------------------------------

class Relation:
    def __init__(self, source, target, rel_type, detail=""):
        self.source = source
        self.target = target
        self.type = rel_type
        self.strength = RELATION_META[rel_type]["strength"]
        self.detail = detail

    def to_dict(self):
        return {
            "source": self.source,
            "target": self.target,
            "type": self.type,
            "strength": self.strength,
            "detail": self.detail,
        }


def analyze_dependencies(resources, ignore_sgs=None):
    """分析资源间依赖关系，返回 [Relation]"""
    ignore_sgs = set(ignore_sgs or [])
    relations = []

    # 索引表
    vpc_resources = defaultdict(list)       # vpc_id -> [rid]
    subnet_resources = defaultdict(list)    # subnet_id -> [rid]
    sg_resources = defaultdict(list)        # sg_id -> [rid]

    for rid, res in resources.items():
        if res.vpc_id:
            vpc_resources[res.vpc_id].append(rid)
        if res.subnet_id:
            subnet_resources[res.subnet_id].append(rid)
        for sg in res.security_groups:
            if sg not in ignore_sgs:
                sg_resources[sg].append(rid)

    # --- 1. 安全组共享关系 ---
    for sg_id, rids in sg_resources.items():
        if len(rids) < 2:
            continue
        # 过滤过于宽泛的安全组（关联超过 20 个资源，大概率是默认安全组）
        if len(rids) > 20:
            continue
        for i in range(len(rids)):
            for j in range(i + 1, len(rids)):
                r1, r2 = resources[rids[i]], resources[rids[j]]
                # 同类型资源的安全组关系价值较低，跳过
                if r1.type == r2.type and r1.type not in ("ecs",):
                    continue
                relations.append(Relation(
                    rids[i], rids[j], REL_SG,
                    f"共享安全组 {sg_id}",
                ))

    # --- 2. 负载均衡 → 后端实例 ---
    for rid, res in resources.items():
        if res.type != "slb":
            continue
        backend_ids = _extract_lb_backends(res.raw)
        for bid in backend_ids:
            if bid in resources:
                relations.append(Relation(
                    rid, bid, REL_LB,
                    f"SLB {res.name} → 后端 {bid}",
                ))

    # --- 3. 数据库连接推断 ---
    ecs_list = [r for r in resources.values() if r.type == "ecs"]
    db_list = [r for r in resources.values() if r.type in ("rds", "redis", "mongodb", "polardb")]

    for ecs in ecs_list:
        for db in db_list:
            if not ecs.vpc_id or not db.vpc_id:
                continue
            if ecs.vpc_id != db.vpc_id:
                continue
            # 同 VPC 且有安全组交集 → 强推断
            sg_overlap = set(ecs.security_groups) & set(db.security_groups)
            if sg_overlap:
                port = _guess_db_port(db.type)
                relations.append(Relation(
                    ecs.id, db.id, REL_DB,
                    f"同 VPC({ecs.vpc_id}) + 共享安全组 {','.join(sg_overlap)}, 端口 {port}",
                ))
            # 同 VPC 同子网 → 弱推断（不重复添加）
            elif ecs.subnet_id and ecs.subnet_id == db.subnet_id:
                port = _guess_db_port(db.type)
                relations.append(Relation(
                    ecs.id, db.id, REL_DB,
                    f"同子网({ecs.subnet_id}), 可能连接端口 {port}",
                ))

    # --- 4. 存储挂载关系 ---
    for rid, res in resources.items():
        if res.type != "disk":
            continue
        attached_id = find_column(res.raw, "InstanceId", "ServerId", "AttachedInstanceId")
        if attached_id and attached_id in resources:
            relations.append(Relation(
                attached_id, rid, REL_STORAGE,
                f"磁盘 {rid} 挂载到 {attached_id}",
            ))

    # 去重
    relations = _deduplicate_relations(relations)
    return relations


def _extract_lb_backends(raw):
    """从 SLB/ELB 的原始数据中提取后端实例 ID"""
    backend_ids = []

    # 尝试各种可能的后端字段
    for key in ("BackendServers", "VServerGroups", "TargetGroups",
                "Instances", "Pools", "Members", "后端服务器"):
        val = None
        for rk in raw:
            if rk and key.lower() in rk.lower():
                val = raw[rk]
                break
        if val is None:
            continue

        s = str(val)
        # 从字符串中提取所有类似实例 ID 的值
        # 阿里云: i-bp..., AWS: i-0a..., 华为云: 各种 UUID
        ids = re.findall(r'i-[a-zA-Z0-9]{8,}', s)
        backend_ids.extend(ids)

        # 尝试 JSON 解析
        try:
            data = json.loads(s)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        for k in ("ServerId", "InstanceId", "Id", "server_id"):
                            if k in item:
                                backend_ids.append(str(item[k]))
        except (json.JSONDecodeError, TypeError):
            pass

    return list(set(backend_ids))


def _guess_db_port(db_type):
    """根据数据库类型返回默认端口"""
    port_map = {
        "rds": 3306, "redis": 6379, "mongodb": 27017,
        "polardb": 3306, "postgresql": 5432, "sqlserver": 1433,
    }
    return port_map.get(db_type, "unknown")


def _deduplicate_relations(relations):
    """去除重复关系（无向去重）"""
    seen = set()
    result = []
    for rel in relations:
        key = tuple(sorted([rel.source, rel.target])) + (rel.type,)
        if key not in seen:
            seen.add(key)
            result.append(rel)
    return result


# ---------------------------------------------------------------------------
# 分组（连通分量）
# ---------------------------------------------------------------------------

def compute_groups(resources, relations):
    """用并查集计算资源分组"""
    parent = {rid: rid for rid in resources}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    # 只基于中等和强依赖分组（弱依赖仅作参考）
    for rel in relations:
        if rel.source in resources and rel.target in resources:
            if rel.strength in (STRENGTH_STRONG, STRENGTH_MEDIUM):
                union(rel.source, rel.target)

    # 同 VPC 的资源也归为一组
    vpc_first = {}
    for rid, res in resources.items():
        if res.vpc_id:
            if res.vpc_id not in vpc_first:
                vpc_first[res.vpc_id] = rid
            else:
                union(rid, vpc_first[res.vpc_id])

    # 收集分组
    groups_map = defaultdict(list)
    for rid in resources:
        root = find(rid)
        groups_map[root].append(rid)

    groups = []
    for gid, (root, rids) in enumerate(sorted(groups_map.items(), key=lambda x: -len(x[1])), 1):
        vpc_id = resources[root].vpc_id
        has_strong = any(
            rel.strength == STRENGTH_STRONG
            for rel in relations
            if rel.source in rids or rel.target in rids
        )
        # 标记分组
        for rid in rids:
            resources[rid].group_id = gid

        groups.append({
            "group_id": gid,
            "name": f"业务组-{gid}" + (f" ({vpc_id})" if vpc_id else ""),
            "vpc_id": vpc_id,
            "resource_count": len(rids),
            "resource_ids": rids,
            "has_strong_dependency": has_strong,
            "suggested_wave": None,  # 后续计算
            "risk_notes": [],
        })

    # 简单 wave 排序
    _assign_waves(groups, resources, relations)
    return groups


def _assign_waves(groups, resources, relations):
    """为每个组分配迁移波次"""
    # 规则：
    # Wave 1: 只包含基础设施(vpc/subnet)和数据库的组
    # Wave 2: 包含 SLB 强依赖的组
    # Wave 3: 有中等依赖的组
    # Wave 4: 独立资源
    # Wave 5: 存储类服务

    for g in groups:
        types_in_group = set(resources[rid].type for rid in g["resource_ids"])

        if types_in_group <= {"vpc", "subnet", "rds", "redis", "mongodb", "polardb"}:
            g["suggested_wave"] = 1
            g["risk_notes"].append("基础设施 + 数据层，建议先迁移并建立数据同步")
        elif "slb" in types_in_group:
            g["suggested_wave"] = 2
            g["risk_notes"].append("包含负载均衡强依赖，需 SLB + 全部后端 ECS 同时切换")
        elif g["has_strong_dependency"]:
            g["suggested_wave"] = 2
            g["risk_notes"].append("存在强依赖关系，必须整组迁移")
        elif g["resource_count"] == 1:
            g["suggested_wave"] = 4
            g["risk_notes"].append("独立资源，可任意顺序迁移")
        elif types_in_group <= {"oss", "nas", "cdn"}:
            g["suggested_wave"] = 5
            g["risk_notes"].append("存储/CDN 服务，建议后台同步后最后切换")
        else:
            g["suggested_wave"] = 3
            g["risk_notes"].append("弱依赖组，可逐步迁移")


# ---------------------------------------------------------------------------
# 输出：文本
# ---------------------------------------------------------------------------

def output_text(scan_file, platform, resources, relations, groups):
    """输出控制台友好的文本报告"""
    print("=" * 60)
    print(f"依赖拓扑分析：{scan_file}")
    print(f"源端平台：{platform}")
    print("=" * 60)

    print(f"\n总资源数：{len(resources)}    依赖关系：{len(relations)}    分组数：{len(groups)}")

    # 按资源类型统计
    type_counts = defaultdict(int)
    for res in resources.values():
        type_counts[res.type] += 1
    print("\n【资源类型分布】")
    for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {t:<15} {c} 个")

    # 依赖关系统计
    rel_type_counts = defaultdict(int)
    for rel in relations:
        rel_type_counts[rel.type] += 1
    if rel_type_counts:
        print("\n【依赖关系统计】")
        for t, c in sorted(rel_type_counts.items(), key=lambda x: -x[1]):
            label = RELATION_META[t]["label"]
            strength = RELATION_META[t]["strength"]
            print(f"  {label:<15} {c} 条  (强度: {strength})")

    # 迁移分组建议
    wave_groups = defaultdict(list)
    for g in groups:
        wave_groups[g["suggested_wave"]].append(g)

    wave_labels = {
        1: "基础设施 + 数据层",
        2: "应用层（强依赖组）",
        3: "应用层（弱依赖组）",
        4: "独立资源",
        5: "存储服务",
    }

    print("\n" + "=" * 60)
    print("迁移分组建议")
    print("=" * 60)

    for wave_num in sorted(wave_groups.keys()):
        label = wave_labels.get(wave_num, f"Wave {wave_num}")
        print(f"\n【Wave {wave_num} — {label}】")
        for g in wave_groups[wave_num]:
            types_in_group = defaultdict(int)
            for rid in g["resource_ids"]:
                types_in_group[resources[rid].type] += 1

            type_summary = ", ".join(f"{t} ×{c}" for t, c in sorted(types_in_group.items(), key=lambda x: -x[1]))
            print(f"  {g['name']}  ({g['resource_count']} 个资源)")
            print(f"    资源: {type_summary}")
            for note in g["risk_notes"]:
                print(f"    → {note}")

            # 列出该组的强依赖关系
            group_set = set(g["resource_ids"])
            strong_rels = [r for r in relations
                          if r.strength == STRENGTH_STRONG
                          and (r.source in group_set or r.target in group_set)]
            if strong_rels:
                print(f"    强依赖:")
                for sr in strong_rels[:5]:  # 最多显示 5 条
                    print(f"      {sr.source} → {sr.target} ({sr.detail})")
                if len(strong_rels) > 5:
                    print(f"      ... 共 {len(strong_rels)} 条强依赖")

    print()


# ---------------------------------------------------------------------------
# 输出：JSON
# ---------------------------------------------------------------------------

def output_json(scan_file, platform, resources, relations, groups):
    """输出 JSON 文件"""
    data = {
        "scan_file": scan_file,
        "platform": platform,
        "analysis_time": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_resources": len(resources),
            "total_relations": len(relations),
            "total_groups": len(groups),
            "vpc_count": len(set(r.vpc_id for r in resources.values() if r.vpc_id)),
        },
        "resources": [r.to_dict() for r in resources.values()],
        "relations": [r.to_dict() for r in relations],
        "groups": groups,
    }

    out_path = os.path.splitext(scan_file)[0] + "_dependency.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已输出依赖拓扑 JSON: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# 输出：Mermaid
# ---------------------------------------------------------------------------

def output_mermaid(scan_file, platform, resources, relations, groups):
    """输出 Mermaid 格式的拓扑图"""
    lines = ["graph TB"]

    # 按 VPC 分子图
    vpc_resources = defaultdict(list)
    no_vpc = []
    for rid, res in resources.items():
        if res.vpc_id:
            vpc_resources[res.vpc_id].append(res)
        else:
            no_vpc.append(res)

    for vpc_id, res_list in vpc_resources.items():
        region = res_list[0].region if res_list else ""
        lines.append(f'    subgraph {_mermaid_id(vpc_id)}["{vpc_id} ({region})"]')
        for res in res_list:
            lines.append(f"        {_mermaid_node(res)}")
        lines.append("    end")

    # 无 VPC 的资源
    for res in no_vpc:
        lines.append(f"    {_mermaid_node(res)}")

    # 关系边
    lines.append("")
    style_map = {
        STRENGTH_STRONG: "-->",
        STRENGTH_MEDIUM: "-->",
        STRENGTH_WEAK: "-..->",
    }
    for rel in relations:
        arrow = style_map.get(rel.strength, "-->")
        label = RELATION_META[rel.type]["label"]
        src = _mermaid_id(rel.source)
        tgt = _mermaid_id(rel.target)
        lines.append(f"    {src} {arrow}|{label}| {tgt}")

    content = "\n".join(lines)
    out_path = os.path.splitext(scan_file)[0] + "_dependency.mmd"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"已输出 Mermaid 拓扑图: {out_path}")
    print(f"\n预览方式: 粘贴到 https://mermaid.live/ 或使用 VS Code Mermaid 插件")
    return out_path


def _mermaid_id(raw_id):
    """生成 Mermaid 安全的节点 ID"""
    return re.sub(r'[^a-zA-Z0-9_]', '_', str(raw_id))


def _mermaid_node(res):
    """生成 Mermaid 节点定义"""
    mid = _mermaid_id(res.id)
    label = f"{res.type.upper()}: {res.name}"
    # 数据库类型用圆柱体形状
    if res.type in ("rds", "redis", "mongodb", "polardb"):
        return f'{mid}[("{label}")]'
    # SLB 用菱形
    if res.type == "slb":
        return f'{mid}{{{{{label}}}}}'
    # 存储用圆柱
    if res.type in ("oss", "nas", "disk"):
        return f'{mid}[("{label}")]'
    # 默认矩形
    return f'{mid}["{label}"]'


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="CMG 依赖拓扑分析 — 从扫描结果中提取资源依赖关系并生成迁移分组建议"
    )
    parser.add_argument("xlsx", help="扫描结果 Excel 文件路径")
    parser.add_argument("--format", choices=["text", "json", "mermaid"], default="text",
                        help="输出格式 (默认: text)")
    parser.add_argument("--ignore-sg", default="",
                        help="忽略的安全组 ID，逗号分隔")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"错误：文件不存在 {args.xlsx}")
        sys.exit(1)

    ignore_sgs = [s.strip() for s in args.ignore_sg.split(",") if s.strip()]

    print(f"正在加载扫描结果: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx)
    platform = detect_platform(wb.sheetnames)
    print(f"检测到平台: {platform}")
    print(f"Sheet 数量: {len(wb.sheetnames)}")

    # 1. 提取资源
    print("正在提取资源...")
    resources = extract_resources(wb, platform)
    print(f"提取到 {len(resources)} 个资源")

    if not resources:
        print("未提取到有效资源，请检查 Excel 文件格式。")
        sys.exit(1)

    # 2. 分析依赖
    print("正在分析依赖关系...")
    relations = analyze_dependencies(resources, ignore_sgs)
    print(f"发现 {len(relations)} 条依赖关系")

    # 3. 计算分组
    print("正在计算迁移分组...")
    groups = compute_groups(resources, relations)
    print(f"生成 {len(groups)} 个迁移分组")

    # 4. 输出
    print()
    if args.format == "text":
        output_text(args.xlsx, platform, resources, relations, groups)
    elif args.format == "json":
        output_json(args.xlsx, platform, resources, relations, groups)
    elif args.format == "mermaid":
        output_mermaid(args.xlsx, platform, resources, relations, groups)


if __name__ == "__main__":
    main()
