#!/usr/bin/env python3
"""
Desktop Automation Library — Ultra robust
PyAutoGUI + PyGetWindow + Pillow + OpenCV (optional) + pytesseract (optional) + pyperclip (optional)
"""
import sys, json, argparse, os, time, subprocess, logging, datetime
import pyautogui
import pygetwindow as gw
from PIL import Image

# Import safety module (dynamic import to work as script or module)
try:
    # Try relative import first (if used as package)
    from .safety import get_safety, set_safe_mode
except ImportError:
    # Fallback: import by path
    safety_path = os.path.join(os.path.dirname(__file__), 'safety.py')
    if os.path.exists(safety_path):
        import importlib.util
        spec = importlib.util.spec_from_file_location('safety', safety_path)
        safety = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(safety)
        get_safety = safety.get_safety
        set_safe_mode = safety.set_safe_mode
    else:
        get_safety = None
        set_safe_mode = None
        logging.warning("safety.py not available — security features disabled")

logger = logging.getLogger(__name__)

# ============ Optional dependencies with graceful degradation ============
try:
    import pyperclip
except ImportError:
    pyperclip = None
    logger.warning("pyperclip not available — copy/paste functions disabled")

try:
    import cv2
    import numpy as np
except ImportError:
    cv2 = None
    np = None
    logger.warning("OpenCV not available — image search functions disabled")

try:
    import pytesseract
except ImportError:
    pytesseract = None
    logger.warning("pytesseract not available — OCR functions disabled")

# Optional dependencies for data/Excel
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available — Excel functions disabled")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("pandas not available — DataFrame/CSV functions disabled")

# Import advanced automation module (if present)
try:
    from . import advanced_automation as adv
    ADV_AVAILABLE = True
except ImportError:
    adv = None
    ADV_AVAILABLE = False
    logger.debug("advanced_automation.py not available — advanced features disabled")

# Import vision module (optional)
try:
    from . import vision as vis
    VISION_AVAILABLE = True
except ImportError:
    vis = None
    VISION_AVAILABLE = False
    logger.debug("vision.py not available — vision features disabled")

# ============ Basic actions ============

def click(x, y, button='left'):
    pyautogui.click(x=x, y=y, button=button)
    return {"status": "ok"}

def type_text(text, interval=0.0):
    pyautogui.typewrite(text, interval=interval)
    return {"status": "ok"}

def screenshot(path=None):
    if path is None:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        path = os.path.join(desktop, "screenshot.png")
    img = pyautogui.screenshot()
    img.save(path)
    return {"status": "ok", "path": path}

def get_active_window():
    win = gw.getActiveWindow()
    if win is None:
        return None
    return {
        "title": win.title,
        "x": win.left,
        "y": win.top,
        "width": win.width,
        "height": win.height
    }

def list_windows():
    wins = gw.getAllWindows()
    result = []
    for win in wins:
        if win.title:
            result.append({
                "title": win.title,
                "x": win.left,
                "y": win.top,
                "width": win.width,
                "height": win.height,
                "is_active": win.isActive
            })
    return result

def activate_window(title_substring):
    wins = gw.getWindowsWithTitle(title_substring)
    if not wins:
        return {"status": "error", "message": "No window found matching: " + title_substring}
    win = wins[0]
    win.activate()
    return {"status": "ok", "title": win.title}

def move_mouse(x, y):
    pyautogui.moveTo(x, y)
    return {"status": "ok"}

def press_key(key):
    pyautogui.press(key)
    return {"status": "ok"}

def scroll(amount):
    pyautogui.scroll(amount)
    return {"status": "ok"}

# Alias: 'type' → type_text (for OpenClaw compatibility)
def type(text, interval=0.0):
    return type_text(text, interval)

# ============ Advanced actions (with optional deps) ============

def find_image(template_path, confidence=0.9):
    if cv2 is None or np is None:
        return {"status": "error", "message": "OpenCV not installed. Install: pip install opencv-python"}
    try:
        template = cv2.imread(template_path, cv2.IMREAD_GRAYSCALE)
        if template is None:
            return {"status": "error", "message": f"Cannot load template image: {template_path}"}
        screen = pyautogui.screenshot()
        screen_np = np.array(screen)
        screen_gray = cv2.cvtColor(screen_np, cv2.COLOR_BGR2GRAY)
        res = cv2.matchTemplate(screen_gray, template, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(res)
        if max_val >= float(confidence):
            h, w = template.shape
            center_x = max_loc[0] + w // 2
            center_y = max_loc[1] + h // 2
            logger.debug(f"Image found at ({center_x},{center_y}) conf={max_val:.3f}")
            return {"status": "ok", "x": center_x, "y": center_y, "confidence": float(max_val)}
        else:
            logger.debug(f"Image not found (conf={max_val:.3f} < {confidence})")
            return {"status": "not_found", "confidence": float(max_val)}
    except Exception as e:
        logger.error(f"find_image error: {e}")
        return {"status": "error", "message": str(e)}

def copy_to_clipboard(text):
    if pyperclip is None:
        return {"status": "error", "message": "pyperclip not installed. Install: pip install pyperclip"}
    try:
        pyperclip.copy(text)
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def paste_from_clipboard():
    pyautogui.hotkey('ctrl', 'v')
    return {"status": "ok"}

def drag(start_x, start_y, end_x, end_y, duration=0.5, button='left'):
    pyautogui.moveTo(start_x, start_y)
    pyautogui.dragTo(end_x, end_y, duration=duration, button=button)
    return {"status": "ok"}

def wait_for_image(template_path, timeout=30, interval=1.0, confidence=0.9):
    start = time.time()
    while time.time() - start < timeout:
        res = find_image(template_path, confidence)
        if res.get("status") == "ok":
            return res
        time.sleep(interval)
    return {"status": "timeout", "message": "Image not found within timeout"}

def wait_for_text_on_screen(text, lang='fra'):
    if pytesseract is None:
        return {"status": "error", "message": "pytesseract not installed. Install: pip install pytesseract"}
    try:
        img = pyautogui.screenshot()
        data = pytesseract.image_to_data(img, lang=lang, output_type=pytesseract.Output.DICT)
        n = len(data['text'])
        for i in range(n):
            if text.lower() in data['text'][i].lower():
                x = data['left'][i] + data['width'][i] // 2
                y = data['top'][i] + data['height'][i] // 2
                logger.debug(f"Text found: '{data['text'][i]}' at ({x},{y})")
                return {"status": "ok", "x": x, "y": y, "matched_text": data['text'][i]}
        logger.debug(f"Text not found: '{text}'")
        return {"status": "not_found"}
    except Exception as e:
        logger.error(f"find_text_on_screen error: {e}")
        return {"status": "error", "message": str(e)}

# ============ Data Extraction & Excel ============

def extract_screen_data(region=None, output_format='json'):
    """
    Extrait des données structurées depuis l'écran (ou une région) via OCR.
    Retourne un tableau d'objets avec texte + bounding boxes.

    Args:
        region: dict {x, y, width, height} ou None pour tout l'écran
        output_format: 'json' (défaut) ou 'csv' (retourne JSON dans les deux cas pour l'instant)

    Returns:
        {"status":"ok", "data": [...], "format": output_format}
    """
    if pytesseract is None:
        return {"status": "error", "message": "pytesseract not installed"}

    try:
        if region:
            img = pyautogui.screenshot(region=(region['x'], region['y'], region['width'], region['height']))
        else:
            img = pyautogui.screenshot()
        data = pytesseract.image_to_data(img, lang='fra+eng', output_type=pytesseract.Output.DICT)

        rows = []
        n = len(data['text'])
        for i in range(n):
            text = data['text'][i].strip()
            if text:  # Ignorer les entrées vides
                rows.append({
                    'text': text,
                    'left': int(data['left'][i]),
                    'top': int(data['top'][i]),
                    'width': int(data['width'][i]),
                    'height': int(data['height'][i]),
                    'conf': float(data['conf'][i])
                })

        return {"status": "ok", "data": rows, "format": output_format, "count": len(rows)}
    except Exception as e:
        logger.error(f"extract_screen_data error: {e}")
        return {"status": "error", "message": str(e)}

def excel_read(filepath, sheet_name=0, range=None):
    """
    Lit un fichier Excel et retourne les données.

    Args:
        filepath: chemin vers le fichier .xlsx
        sheet_name: nom ou index (défaut 0)
        range: plage de cellules A1:XX (optionnel)

    Returns:
        {"status":"ok", "data": [...], "columns": [...], "sheet": sheet_name}
    """
    if not os.path.exists(filepath):
        return {"status": "error", "message": f"File not found: {filepath}"}

    try:
        if not OPENPYXL_AVAILABLE:
            return {"status": "error", "message": "openpyxl not installed. Install: pip install openpyxl"}

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

        data = []
        # Déterminer la plage
        if range:
            # Simple parse "A1:C10"
            start_cell, end_cell = range.split(':')
            start_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, start_cell)))
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            end_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, end_cell)))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
        else:
            # Tout le tableau utilisé
            start_row, start_col = 1, 1
            end_row = ws.max_row
            end_col = ws.max_column

        # Lire les en-têtes (première ligne)
        headers = []
        for col in range(start_col, end_col+1):
            cell_value = ws.cell(row=start_row, column=col).value
            headers.append(str(cell_value) if cell_value else f"Col{col}")

        # Lire les données
        for row in range(start_row+1, end_row+1):
            row_data = {}
            for col_idx, col in enumerate(range(start_col, end_col+1)):
                cell = ws.cell(row=row, column=col)
                row_data[headers[col_idx]] = cell.value
            data.append(row_data)

        wb.close()
        return {"status": "ok", "data": data, "columns": headers, "sheet": sheet_name, "rows": len(data)}
    except Exception as e:
        logger.error(f"excel_read error: {e}")
        return {"status": "error", "message": str(e)}

def excel_write(filepath, data, sheet_name='Sheet1', start_cell='A1'):
    """
    Écrit des données dans un fichier Excel.

    Args:
        filepath: chemin de sortie .xlsx
        data: liste de dicts (colonnes = clés) ou liste de listes
        sheet_name: nom de la feuille
        start_cell: cellule de départ (ex: 'A1')

    Returns:
        {"status":"ok", "filepath": filepath, "rows": N, "columns": M}
    """
    if not OPENPYXL_AVAILABLE:
        return {"status": "error", "message": "openpyxl not installed"}

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Déterminer la cellule de départ
        start_col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)

        # Écrire les en-têtes si data est une liste de dicts
        if data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers):
                ws.cell(row=start_row, column=start_col_idx + col_idx, value=header)
            start_row += 1
            for row_idx, row_dict in enumerate(data, start=start_row):
                for col_idx, key in enumerate(headers):
                    ws.cell(row=row_idx, column=start_col_idx + col_idx, value=row_dict.get(key))
        elif data and isinstance(data[0], (list, tuple)):
            # Liste de listes
            for col_idx, value in enumerate(data[0]):
                ws.cell(row=start_row, column=start_col_idx + col_idx, value=value)
            start_row += 1
            for row_idx, row_list in enumerate(data, start=start_row):
                for col_idx, value in enumerate(row_list):
                    ws.cell(row=row_idx, column=start_col_idx + col_idx, value=value)
        else:
            return {"status": "error", "message": "Data must be list of dicts or list of lists"}

        wb.save(filepath)
        return {"status": "ok", "filepath": filepath, "rows": len(data), "columns": len(data[0]) if data else 0}
    except Exception as e:
        logger.error(f"excel_write error: {e}")
        return {"status": "error", "message": str(e)}

def data_to_csv(data, filepath=None):
    """
    Convertit des données (liste de dicts) en CSV.

    Args:
        data: liste de dictionnaires
        filepath: chemin optionnel pour sauvegarder

    Returns:
        {"status":"ok", "csv": "...", "filepath": "..."} ou erreur
    """
    if not PANDAS_AVAILABLE:
        return {"status": "error", "message": "pandas not installed. Install: pip install pandas"}

    try:
        df = pd.DataFrame(data)
        if filepath:
            df.to_csv(filepath, index=False, encoding='utf-8')
            return {"status": "ok", "filepath": filepath, "rows": len(df), "columns": len(df.columns)}
        else:
            csv_str = df.to_csv(index=False)
            return {"status": "ok", "csv": csv_str, "rows": len(df), "columns": len(df.columns)}
    except Exception as e:
        logger.error(f"data_to_csv error: {e}")
        return {"status": "error", "message": str(e)}

# ============ Monitor & Conditional Actions ============

def monitor_screen(checks, timeout=60, interval=0.5, stop_condition=None, fallback_confidence=0.85):
    """
    Surveille l'écran jusqu'à ce qu'une condition soit remplie, puis exécute une action.

    Args:
        checks: liste de conditions, chaque condition est un dict:
            {
                "type": "image" | "text",
                "template_path": "path" (pour image),
                "text": "string" (pour texte),
                "confidence": 0.9,
                "action": "action_name",
                "action_params": { ... }
            }
        timeout: secondes max de surveillance
        interval: délai entre les vérifications (secondes)
        stop_condition: "first_match" (défaut) ou "all_matched" ou None (continue until timeout)
        fallback_confidence: si None, pas de fallback; sinon liste de confiances à essayer en descending order si échec

    Retour:
        {"status": "ok", "matches": [...], "actions_executed": [...]} ou {"status": "timeout", "matches": [...]}
    """
    start_time = time.time()
    matches = []
    actions_executed = []
    indices_checked = set()

    # Si fallback_confidence est un float, le convertir en liste [fallback, ...] décroissante
    if isinstance(fallback_confidence, (int, float)):
        # Ex: 0.85 → [0.85, 0.8, 0.75, 0.7]
        base = float(fallback_confidence)
        fallback_list = [base - 0.05*i for i in range(5) if base - 0.05*i >= 0.5]
    elif isinstance(fallback_confidence, list):
        fallback_list = sorted(fallback_confidence, reverse=True)
    else:
        fallback_list = []

    def check_condition(cond):
        ctype = cond.get('type')
        if ctype == 'image':
            # Vérifier avec confiance normale, puis fallback si besoin
            confidences = [cond.get('confidence', 0.9)] + fallback_list
            for conf in confidences:
                res = find_image(cond['template_path'], confidence=conf)
                if res.get('status') == 'ok':
                    return {'status': 'ok', 'type': 'image', 'location': res, 'confidence': conf}
            return {'status': 'not_found', 'type': 'image'}
        elif ctype == 'text':
            # OCR search
            res = find_text_on_screen(cond['text'], lang=cond.get('lang', 'fra'))
            if res.get('status') == 'ok':
                return {'status': 'ok', 'type': 'text', 'location': res}
            return {'status': 'not_found', 'type': 'text'}
        else:
            return {'status': 'error', 'message': f"Unknown check type: {ctype}"}

    while time.time() - start_time < timeout:
        # Vérifier toutes les conditions non encore satisfaites
        for idx, cond in enumerate(checks):
            if idx in indices_checked:
                continue  # Already matched
            result = check_condition(cond)
            if result.get('status') == 'ok':
                # Enregistrer le match
                match_info = {
                    'index': idx,
                    'condition': cond,
                    'result': result
                }
                matches.append(match_info)
                indices_checked.add(idx)
                # Exécuter l'action associée
                action_name = cond.get('action')
                action_params = cond.get('action_params', {}).copy()
                # Si l'action a besoin de coordonnées, fournir celles de la détection
                if result['location']:
                    if 'x' not in action_params and 'x' in result['location']:
                        action_params['x'] = result['location']['x']
                    if 'y' not in action_params and 'y' in result['location']:
                        action_params['y'] = result['location']['y']
                # Exécuter l'action (seulement les actions de base autorisées)
                safe_actions = {
                    'click', 'type', 'press_key', 'scroll', 'move_mouse',
                    'activate_window', 'drag', 'copy_to_clipboard', 'paste_from_clipboard',
                    'screenshot', 'list_windows', 'get_active_window'
                }
                if action_name in safe_actions and action_name in globals():
                    try:
                        act_res = globals()[action_name](**action_params)
                        actions_executed.append({'index': idx, 'action': action_name, 'result': act_res})
                        logger.info(f"Condition {idx} matched, executed {action_name} → {act_res}")
                    except Exception as e:
                        logger.error(f"Action {action_name} failed: {e}")
                else:
                    logger.warning(f"Action {action_name} not allowed or not found")
                # Gestion du stop_condition
                if stop_condition == 'first_match':
                    # On s'arrête après la première condition remplie
                    return {
                        "status": "ok",
                        "matches": matches,
                        "actions_executed": actions_executed,
                        "elapsed": time.time() - start_time
                    }
        # Si stop_condition == 'all_matched' et toutes les conditions sont remplies
        if stop_condition == 'all_matched' and len(indices_checked) == len(checks):
            return {
                "status": "ok",
                "matches": matches,
                "actions_executed": actions_executed,
                "elapsed": time.time() - start_time
            }
        time.sleep(interval)

    # Timeout
    return {
        "status": "timeout",
        "matches": matches,
        "actions_executed": actions_executed,
        "elapsed": time.time() - start_time
    }


# ============ Macro Recording/Playback (external scripts) ============

def record_macro(output_path=None):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    recorder_script = os.path.join(script_dir, '..', 'scripts', 'record_macro.py')
    if not os.path.exists(recorder_script):
        return {"status": "error", "message": f"Recorder script not found: {recorder_script}"}
    args = [sys.executable, recorder_script]
    if output_path:
        args.append(output_path)
    try:
        subprocess.Popen(args)
        return {"status": "started", "message": "Macro recorder GUI launched"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def play_macro(macro_path, speed=1.0):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    player_script = os.path.join(script_dir, '..', 'scripts', 'play_macro.py')
    if not os.path.exists(player_script):
        return {"status": "error", "message": f"Player script not found: {player_script}"}
    if not os.path.exists(macro_path):
        return {"status": "error", "message": f"Macro file not found: {macro_path}"}
    try:
        subprocess.run([sys.executable, player_script, macro_path, str(speed)], check=True)
        return {"status": "ok"}
    except subprocess.CalledProcessError as e:
        return {"status": "error", "message": str(e)}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ============ Entry point ============

def _setup_logging():
    """Configure logging to file and console"""
    log_dir = os.path.join(os.path.expanduser("~"), ".openclaw", "skills", "desktop-automation-logs")
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"automation_{datetime.now().strftime('%Y-%m-%d')}.log")

    logging.basicConfig(
        level=logging.INFO,
        format='[%(levelname)s] %(asctime)s — %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stderr)
        ]
    )

def main():
    _setup_logging()
    parser = argparse.ArgumentParser(description="Desktop Automation Library")
    parser.add_argument("action", choices=[
        # Basic
        "click", "type", "screenshot", "get_active_window", "list_windows",
        "activate_window", "move_mouse", "press_key", "scroll", "find_image",
        "copy_to_clipboard", "paste_from_clipboard", "drag",
        "wait_for_image", "find_text_on_screen",
        "record_macro", "play_macro",
        # Advanced monitoring
        "monitor_screen", "set_safe_mode",
        # Data
        "extract_screen_data", "excel_read", "excel_write", "data_to_csv",
        # Advanced automation
        "find_image_multiscale",
        "find_all_text_on_screen",
        "detect_ui_elements",
        "monitor_screen_with_logic",
        "play_macro_with_subroutines",
        "create_protected_macro",
        "load_and_decrypt_protected_macro",
        "generate_macro_report",
        # Vision (autonomous desktop agent)
        "screenshot_mss",
        "find_on_screen",
        "click_position",
        "click_image",
        "read_text_ocr",
        "read_text_region"
    ])
    parser.add_argument("json_args", nargs="?", help="JSON-encoded parameters")
    args = parser.parse_args()

    params = {}
    if args.json_args:
        try:
            params = json.loads(args.json_args)
        except Exception as e:
            print(json.dumps({"status": "error", "message": "Invalid JSON args: " + str(e)}))
            sys.exit(1)

    func_name = args.action

    # Handle special actions
    if func_name == "set_safe_mode":
        enabled = params.get('enabled', True)
        if set_safe_mode:
            set_safe_mode(enabled)
            print(json.dumps({"status": "ok", "message": f"Safe mode {'enabled' if enabled else 'disabled'}"}))
        else:
            print(json.dumps({"status": "error", "message": "Safety module not available"}))
        sys.exit(0)

    # Chercher la fonction dans les globals, advanced_automation, ou vision
    if func_name in globals():
        func = globals()[func_name]
    elif ADV_AVAILABLE and hasattr(adv, func_name):
        func = getattr(adv, func_name)
    elif VISION_AVAILABLE and hasattr(vis, func_name):
        func = getattr(vis, func_name)
    else:
        print(json.dumps({"status": "error", "message": f"Unknown action: {func_name}"}))
        sys.exit(1)

    # Apply safety validation if available
    if get_safety and func_name not in ('set_safe_mode',):
        validation = get_safety().validate_action(func_name, params)
        if not validation['allowed']:
            logger.warning(f"Blocked action: {func_name} — {validation['reason']}")
            print(json.dumps({"status": "error", "message": validation['reason']}))
            sys.exit(1)
        params = validation['params']

    # Check dry_run flag
    dry_run = params.pop('dry_run', False)
    if dry_run:
        logger.info(f"DRY RUN: would execute {func_name} with {params}")
        print(json.dumps({"status": "ok", "dry_run": True, "message": f"Simulated {func_name}"}))
        sys.exit(0)

    try:
        result = func(**params)
        print(json.dumps(result))
    except Exception as e:
        logger.exception(f"Error executing action {func_name}")
        print(json.dumps({"status": "error", "message": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()
