from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from src.fiscal_parser import (
    DocumentMeta,
    build_period_label,
    build_source_period_key,
    extract_metrics_from_content,
    parse_period_from_title,
    parse_yoy_value,
)
from src.fiscal_crawler import CrawlScope
from src.fiscal_exporter import FiscalExporter
from src.fiscal_transform import build_derived_metrics
from scripts.run_pipeline import _collect_output_periods


class PipelineTests(unittest.TestCase):
    def test_parse_period(self):
        self.assertEqual(parse_period_from_title("2026年1-2月财政收支情况"), ("1-2月", "2026-02"))
        self.assertEqual(parse_period_from_title("2025年一季度财政收支情况"), ("1-3月", "2025-03"))
        self.assertEqual(parse_period_from_title("2025年前三季度财政收支情况"), ("1-9月", "2025-09"))
        self.assertEqual(parse_period_from_title("2025年上半年财政收支情况"), ("1-6月", "2025-06"))
        self.assertEqual(parse_period_from_title("2025年财政收支情况"), ("1-12月", "2025-12"))
        self.assertEqual(build_period_label("1-2月", "2026-02"), "2026-01~02")
        self.assertEqual(build_source_period_key("1-2月", "2026-02"), "202601-202602")
        self.assertEqual(build_source_period_key("1-3月", "2025-03"), "202501-202503")

    def test_parse_yoy(self):
        self.assertEqual(parse_yoy_value("同比增长4.7%"), 4.7)
        self.assertEqual(parse_yoy_value("同比下降6.2%"), -6.2)
        self.assertEqual(parse_yoy_value("同比增长1.1倍"), 110.0)

    def test_extract_metrics(self):
        content = (
            "（一）一般公共预算收入情况。1-2月，全国一般公共预算收入44154亿元，同比增长0.7%。"
            "其中，全国税收收入36393亿元，同比增长0.1%；非税收入7761亿元，同比增长3.4%。"
            "主要税收收入项目情况如下：1.国内增值税15838亿元，同比增长4.7%。"
            "（二）一般公共预算支出情况。1-2月，全国一般公共预算支出46706亿元，同比增长3.6%。"
        )
        meta = DocumentMeta(
            title="2026年1-2月财政收支情况",
            url="https://example.com",
            publish_date=date(2026, 3, 19),
            source_department="国库司",
            content=content,
            period_text="1-2月",
            cutoff_month="2026-02",
        )
        extracted = extract_metrics_from_content("doc-1", meta)
        names = {item["standard_metric_name"] for item in extracted}
        self.assertIn("全国一般公共预算收入", names)
        self.assertIn("全国税收收入", names)
        self.assertIn("全国非税收入", names)
        self.assertIn("国内增值税", names)
        self.assertIn("全国一般公共预算支出", names)
        jan_feb = [item for item in extracted if item["standard_metric_name"] == "全国一般公共预算收入"][0]
        self.assertEqual(jan_feb["period_label"], "2026-01~02")
        non_tax = [item for item in extracted if item["standard_metric_name"] == "全国非税收入"][0]
        self.assertEqual(non_tax["source_metric_name"], "非税收入")
        ordered_names = [item["standard_metric_name"] for item in extracted[:4]]
        self.assertEqual(
            ordered_names,
            ["全国一般公共预算收入", "全国税收收入", "全国非税收入", "国内增值税"],
        )

    def test_extract_metrics_supports_tax_income_alias(self):
        content = (
            "（一）一般公共预算收入情况。1-7月累计，全国一般公共预算收入139334亿元，同比增长11.5%。"
            "其中，税收收入117531亿元，同比增长14.5%；非税收入21803亿元，同比下降2.3%。"
        )
        meta = DocumentMeta(
            title="2023年7月财政收支情况",
            url="https://example.com",
            publish_date=date(2023, 8, 18),
            source_department="国库司",
            content=content,
            period_text="7月",
            cutoff_month="2023-07",
        )
        extracted = extract_metrics_from_content("doc-tax-alias", meta)
        tax_income = [item for item in extracted if item["standard_metric_name"] == "全国税收收入"][0]
        self.assertEqual(tax_income["source_metric_name"], "税收收入")

    def test_export_rebate_keeps_source_text(self):
        content = "（一）一般公共预算收入情况。1-2月，出口退税5074亿元，同比增长2.0%。"
        meta = DocumentMeta(
            title="2026年1-2月财政收支情况",
            url="https://example.com",
            publish_date=date(2026, 3, 19),
            source_department="国库司",
            content=content,
            period_text="1-2月",
            cutoff_month="2026-02",
        )
        extracted = extract_metrics_from_content("doc-2", meta)
        rebate = [item for item in extracted if item["standard_metric_name"] == "出口退税"][0]
        self.assertEqual(rebate["source_metric_name"], "出口退税")

    def test_export_rebate_supports_long_source_text(self):
        content = "（一）一般公共预算收入情况。1-2月，出口货物退增值税、消费税5074亿元，同比增长2.0%。"
        meta = DocumentMeta(
            title="2026年1-2月财政收支情况",
            url="https://example.com",
            publish_date=date(2026, 3, 19),
            source_department="国库司",
            content=content,
            period_text="1-2月",
            cutoff_month="2026-02",
        )
        extracted = extract_metrics_from_content("doc-3", meta)
        rebate = [item for item in extracted if item["standard_metric_name"] == "出口退税"][0]
        self.assertEqual(rebate["source_metric_name"], "出口货物退增值税、消费税")

    def test_build_derived_metrics(self):
        records = [
            {
                "record_id": "a",
                "metric_month": "2025-09",
                "standard_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "value": 163876.0,
                "yoy_growth": 0.5,
                "source_document_id": "doc-a",
            },
            {
                "record_id": "b",
                "metric_month": "2025-10",
                "standard_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "value": 186490.0,
                "yoy_growth": 0.8,
                "source_document_id": "doc-b",
            },
        ]
        derived, issues = build_derived_metrics(records)
        oct_record = [item for item in derived if item["metric_month"] == "2025-10"][0]
        self.assertAlmostEqual(oct_record["value"], 22614.0)
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0]["metric_month"], "2025-09")

    def test_build_derived_metrics_for_combined_jan_feb(self):
        records = [
            {
                "record_id": "feb-income",
                "metric_month": "2026-02",
                "period_label": "2026-01~02",
                "period_text": "1-2月",
                "standard_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "value": 44154.0,
                "yoy_growth": 0.7,
                "source_document_id": "doc-feb",
            },
            {
                "record_id": "feb-spend",
                "metric_month": "2026-02",
                "period_label": "2026-01~02",
                "period_text": "1-2月",
                "standard_metric_name": "全国一般公共预算支出",
                "metric_category": "一般公共预算支出",
                "value": 46706.0,
                "yoy_growth": 3.6,
                "source_document_id": "doc-feb",
            },
            {
                "record_id": "feb-fund-income",
                "metric_month": "2026-02",
                "period_label": "2026-01~02",
                "period_text": "1-2月",
                "standard_metric_name": "全国政府性基金预算收入",
                "metric_category": "政府性基金预算收入",
                "value": 6381.0,
                "yoy_growth": -10.0,
                "source_document_id": "doc-feb",
            },
            {
                "record_id": "feb-fund-spend",
                "metric_month": "2026-02",
                "period_label": "2026-01~02",
                "period_text": "1-2月",
                "standard_metric_name": "全国政府性基金预算支出",
                "metric_category": "政府性基金预算支出",
                "value": 10901.0,
                "yoy_growth": 5.0,
                "source_document_id": "doc-feb",
            },
        ]
        derived, issues = build_derived_metrics(records, scope_month="2026-02")
        self.assertEqual(issues, [])
        jan_narrow = [item for item in derived if item["metric_month"] == "2026-01" and item["standard_metric_name"] == "窄口径赤字"]
        feb_narrow = [item for item in derived if item["metric_month"] == "2026-02" and item["standard_metric_name"] == "窄口径赤字"]
        jan_wide = [item for item in derived if item["metric_month"] == "2026-01" and item["standard_metric_name"] == "宽口径赤字"]
        feb_wide = [item for item in derived if item["metric_month"] == "2026-02" and item["standard_metric_name"] == "宽口径赤字"]
        self.assertTrue(jan_narrow)
        self.assertTrue(feb_narrow)
        self.assertTrue(jan_wide)
        self.assertTrue(feb_wide)
        avg_income_records = [item for item in derived if item["standard_metric_name"] == "全国一般公共预算收入"]
        self.assertEqual([item["metric_month"] for item in avg_income_records], ["2026-01", "2026-02"])
        self.assertTrue(all(item["derived_type"] == "average_from_jan_feb_combined" for item in avg_income_records))
        self.assertTrue(all(item["value"] == 22077.0 for item in avg_income_records))
        self.assertTrue(all(item["remark"] == "1-2月合并值按平均值拆分为1月和2月" for item in avg_income_records))

    def test_month_scope_fetch_and_output(self):
        scope = CrawlScope(start_month="2026-02", end_month="2026-02")
        self.assertFalse(scope.should_fetch("2026-01"))
        self.assertTrue(scope.should_fetch("2026-02"))
        self.assertFalse(scope.should_fetch("2026-03"))
        self.assertFalse(scope.should_fetch("2025-12"))
        self.assertFalse(scope.should_output("2026-01"))
        self.assertTrue(scope.should_output("2026-02", "2026-01~02"))

    def test_jan_request_maps_to_jan_feb_period(self):
        scope = CrawlScope(start_month="2026-01", end_month="2026-01")
        self.assertEqual(scope.normalized_start_month(), "2026-02")
        self.assertEqual(scope.requested_period_range_key(), "202601-202602")
        self.assertTrue(scope.should_fetch("2026-02"))
        self.assertTrue(scope.should_output("2026-02", "2026-01~02"))
        self.assertFalse(scope.should_output("2026-01", "2026-01"))

    def test_exporter_only_outputs_derived_in_summary_dir(self):
        with TemporaryDirectory() as tmpdir:
            exporter = FiscalExporter(output_dir=tmpdir)
            documents = [
                {
                    "document_id": "2025年1-2月财政收支情况",
                    "title": "2025年1-2月财政收支情况",
                    "url": "https://example.com/doc",
                    "publish_date": "2025-03-01",
                    "source_department": "国库司",
                    "content": "正文",
                    "period_text": "1-2月",
                    "cutoff_month": "2025-02",
                    "period_label": "2025-01~02",
                }
            ]
            extracted_metrics = [
                {
                    "record_id": "2025年1-2月财政收支情况:全国一般公共预算收入",
                    "metric_month": "2025-02",
                    "period_label": "2025-01~02",
                    "standard_metric_name": "全国一般公共预算收入",
                    "source_metric_name": "全国一般公共预算收入",
                    "metric_category": "一般公共预算收入",
                    "direction": "财政收入",
                    "value": 100.0,
                    "unit": "亿元",
                    "yoy_growth": 1.0,
                    "source_document_id": "2025年1-2月财政收支情况",
                    "source_title": "2025年1-2月财政收支情况",
                    "source_publish_date": "2025-03-01",
                    "period_text": "1-2月",
                    "cutoff_month": "2025-02",
                }
            ]
            derived_metrics = [
                {
                    "metric_month": "2025-01",
                    "period_label": "2025-01",
                    "standard_metric_name": "全国一般公共预算收入",
                    "metric_category": "一般公共预算收入",
                    "value": 50.0,
                    "unit": "亿元",
                    "yoy_growth": 1.0,
                    "derived_type": "average_from_jan_feb_combined",
                    "formula": "2025-01~02=100.0/2",
                    "source_record_ids": "r1",
                    "source_document_ids": "d1",
                    "is_derived_metric": True,
                    "remark": "1-2月合并值按平均值拆分为1月和2月",
                },
                {
                    "metric_month": "2025-02",
                    "period_label": "2025-02",
                    "standard_metric_name": "全国一般公共预算收入",
                    "metric_category": "一般公共预算收入",
                    "value": 50.0,
                    "unit": "亿元",
                    "yoy_growth": 1.0,
                    "derived_type": "average_from_jan_feb_combined",
                    "formula": "2025-01~02=100.0/2",
                    "source_record_ids": "r1",
                    "source_document_ids": "d1",
                    "is_derived_metric": True,
                    "remark": "1-2月合并值按平均值拆分为1月和2月",
                },
                {
                    "metric_month": "2025-01",
                    "period_label": "2025-01",
                    "standard_metric_name": "全国一般公共预算支出",
                    "metric_category": "一般公共预算支出",
                    "value": 60.0,
                    "unit": "亿元",
                    "yoy_growth": 1.5,
                    "derived_type": "average_from_jan_feb_combined",
                    "formula": "2025-01~02=120.0/2",
                    "source_record_ids": "r2",
                    "source_document_ids": "d2",
                    "is_derived_metric": True,
                    "remark": "1-2月合并值按平均值拆分为1月和2月",
                },
                {
                    "metric_month": "2025-02",
                    "period_label": "2025-02",
                    "standard_metric_name": "全国一般公共预算支出",
                    "metric_category": "一般公共预算支出",
                    "value": 60.0,
                    "unit": "亿元",
                    "yoy_growth": 1.5,
                    "derived_type": "average_from_jan_feb_combined",
                    "formula": "2025-01~02=120.0/2",
                    "source_record_ids": "r2",
                    "source_document_ids": "d2",
                    "is_derived_metric": True,
                    "remark": "1-2月合并值按平均值拆分为1月和2月",
                },
                {
                    "metric_month": "2025-01",
                    "period_label": "2025-01",
                    "standard_metric_name": "窄口径赤字",
                    "metric_category": "一般公共预算赤字",
                    "value": 20.0,
                    "unit": "亿元",
                    "yoy_growth": None,
                    "derived_type": "deficit_metric",
                    "formula": "窄口径赤字=全国一般公共预算支出-全国一般公共预算收入",
                    "source_record_ids": "r1,r2",
                    "source_document_ids": "d1,d2",
                    "is_derived_metric": True,
                    "remark": "",
                },
                {
                    "metric_month": "2025-02",
                    "period_label": "2025-02",
                    "standard_metric_name": "窄口径赤字",
                    "metric_category": "一般公共预算赤字",
                    "value": 20.0,
                    "unit": "亿元",
                    "yoy_growth": None,
                    "derived_type": "deficit_metric",
                    "formula": "窄口径赤字=全国一般公共预算支出-全国一般公共预算收入",
                    "source_record_ids": "r1,r2",
                    "source_document_ids": "d1,d2",
                    "is_derived_metric": True,
                    "remark": "",
                },
                {
                    "metric_month": "2025-01",
                    "period_label": "2025-01",
                    "standard_metric_name": "宽口径赤字",
                    "metric_category": "合并口径赤字",
                    "value": 30.0,
                    "unit": "亿元",
                    "yoy_growth": None,
                    "derived_type": "deficit_metric",
                    "formula": "宽口径赤字=(全国一般公共预算支出+全国政府性基金预算支出)-(全国一般公共预算收入+全国政府性基金预算收入)",
                    "source_record_ids": "r1,r2,r3,r4",
                    "source_document_ids": "d1,d2,d3,d4",
                    "is_derived_metric": True,
                    "remark": "",
                },
                {
                    "metric_month": "2025-02",
                    "period_label": "2025-02",
                    "standard_metric_name": "宽口径赤字",
                    "metric_category": "合并口径赤字",
                    "value": 30.0,
                    "unit": "亿元",
                    "yoy_growth": None,
                    "derived_type": "deficit_metric",
                    "formula": "宽口径赤字=(全国一般公共预算支出+全国政府性基金预算支出)-(全国一般公共预算收入+全国政府性基金预算收入)",
                    "source_record_ids": "r1,r2,r3,r4",
                    "source_document_ids": "d1,d2,d3,d4",
                    "is_derived_metric": True,
                    "remark": "",
                }
            ]
            exporter.export(
                documents=documents,
                extracted_metrics=extracted_metrics,
                derived_metrics=derived_metrics,
                allowed_periods=["202501-202502"],
                run_context={
                    "run_started_at": date(2026, 3, 26),
                    "run_finished_at": date(2026, 3, 26),
                    "run_timestamp": "20260326120000",
                    "requested_scope": "按区间 2025-01 至 2025-02",
                    "run_mode": "按区间",
                    "command_summary": "python scripts/run_pipeline.py --start-month 2025-01 --end-month 2025-02",
                    "requested_period_range": "202501-202502",
                    "reused_period_count": 0,
                    "refetched_period_count": 1,
                    "refetched_periods": ["202501-202502"],
                },
                issues=[],
            )
            period_dir = Path(tmpdir) / "202501-202502"
            summary_root_dir = Path(tmpdir) / "202501-202502_汇总"
            summary_dir = summary_root_dir / "20260326120000"
            self.assertTrue((period_dir / "raw_documents.xlsx").exists())
            self.assertTrue((period_dir / "extracted_metrics.xlsx").exists())
            self.assertFalse((period_dir / "derived_metrics.xlsx").exists())
            self.assertTrue((summary_dir / "derived_metrics_20260326120000.xlsx").exists())
            self.assertTrue((summary_dir / "monthly_summary_20260326120000.xlsx").exists())
            self.assertTrue((summary_dir / "运行汇报_20260326120000.md").exists())
            wb = load_workbook(summary_dir / "derived_metrics_20260326120000.xlsx", read_only=True)
            try:
                rows = list(wb.active.iter_rows(values_only=True))
                self.assertIn("收支方向", rows[0])
                self.assertEqual(rows[1][0], "202501")
                direction_index = rows[0].index("收支方向")
                metric_index = rows[0].index("指标（单位：亿元）")
                derived_type_index = rows[0].index("推导类型")
                remark_index = rows[0].index("备注")
                direction_map = {row[metric_index]: row[direction_index] for row in rows[1:]}
                self.assertEqual(direction_map["全国一般公共预算收入"], "财政收入")
                self.assertEqual(direction_map["全国一般公共预算支出"], "财政支出")
                self.assertEqual(direction_map["窄口径赤字"], "赤字")
                self.assertEqual(direction_map["宽口径赤字"], "赤字")
                self.assertEqual(rows[1][derived_type_index], "1-2月平均值")
                self.assertEqual(rows[1][remark_index], "1-2月合并值按平均值拆分为1月和2月")
            finally:
                wb.close()
            wb = load_workbook(summary_dir / "monthly_summary_20260326120000.xlsx")
            try:
                ws = wb.active
                self.assertEqual(ws.cell(row=1, column=1).value, "收支大类")
                self.assertEqual(ws.cell(row=1, column=3).value, "指标（单位：亿元）")
                self.assertEqual(ws.cell(row=1, column=4).value, "202501")
                self.assertEqual(ws.cell(row=1, column=5).value, "202502")
                values = [tuple(row[:3]) for row in ws.iter_rows(min_row=2, max_col=3, values_only=True)]
                self.assertIn(("财政收入", "一般公共预算收入", "全国一般公共预算收入"), values)
                self.assertIn(("财政支出", "一般公共预算支出", "全国一般公共预算支出"), values)
                self.assertIn(("赤字", "窄口径赤字", "窄口径赤字"), values)
                self.assertIn((None, "宽口径赤字", "宽口径赤字"), values)
                merged_ranges = {str(item) for item in ws.merged_cells.ranges}
                self.assertTrue(any(item.startswith("A") and ":A" in item for item in merged_ranges))
                first_data_row = [cell.value for cell in ws[2]]
                self.assertEqual(first_data_row[3], 50)
                self.assertEqual(first_data_row[4], 50)
            finally:
                wb.close()

    def test_exporter_uses_full_summary_root_dir(self):
        exporter = FiscalExporter(output_dir="output")
        self.assertEqual(
            exporter._build_run_summary_root_dirname([], None, "全量"),
            "全量汇总",
        )
        self.assertEqual(
            exporter._build_run_summary_root_dirname([], "202501-202504", "按区间"),
            "202501-202504_汇总",
        )

    def test_month_scope_collects_required_source_periods(self):
        scope = CrawlScope(start_month="2025-03", end_month="2025-03")
        documents = [
            {
                "document_id": "2025年1-2月财政收支情况",
                "title": "2025年1-2月财政收支情况",
                "url": "https://example.com/1",
                "publish_date": "2025-03-01",
                "source_department": "国库司",
                "content": "正文",
                "period_text": "1-2月",
                "cutoff_month": "2025-02",
                "period_label": "2025-01~02",
            },
            {
                "document_id": "2025年一季度财政收支情况",
                "title": "2025年一季度财政收支情况",
                "url": "https://example.com/2",
                "publish_date": "2025-04-01",
                "source_department": "国库司",
                "content": "正文",
                "period_text": "1-3月",
                "cutoff_month": "2025-03",
                "period_label": "2025-03",
            },
        ]
        extracted_metrics = [
            {
                "record_id": "r1",
                "metric_month": "2025-02",
                "period_label": "2025-01~02",
                "standard_metric_name": "全国一般公共预算收入",
                "source_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "direction": "财政收入",
                "value": 100.0,
                "unit": "亿元",
                "yoy_growth": 1.0,
                "source_document_id": "2025年1-2月财政收支情况",
                "source_title": "2025年1-2月财政收支情况",
                "source_publish_date": "2025-03-01",
                "period_text": "1-2月",
                "cutoff_month": "2025-02",
            },
            {
                "record_id": "r2",
                "metric_month": "2025-03",
                "period_label": "2025-03",
                "standard_metric_name": "全国一般公共预算收入",
                "source_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "direction": "财政收入",
                "value": 150.0,
                "unit": "亿元",
                "yoy_growth": 1.0,
                "source_document_id": "2025年一季度财政收支情况",
                "source_title": "2025年一季度财政收支情况",
                "source_publish_date": "2025-04-01",
                "period_text": "1-3月",
                "cutoff_month": "2025-03",
            },
        ]
        self.assertEqual(
            _collect_output_periods(documents, extracted_metrics, scope),
            ["202501-202502", "202501-202503"],
        )

    def test_range_summary_key_for_jan_to_mar(self):
        scope = CrawlScope(start_month="2025-01", end_month="2025-03")
        self.assertEqual(scope.requested_period_range_key(), "202501-202503")

    def test_build_derived_metrics_for_range(self):
        records = [
            {
                "record_id": "feb-income",
                "metric_month": "2025-02",
                "period_label": "2025-01~02",
                "period_text": "1-2月",
                "standard_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "value": 100.0,
                "yoy_growth": 1.0,
                "source_document_id": "doc-feb",
            },
            {
                "record_id": "mar-income",
                "metric_month": "2025-03",
                "period_label": "2025-03",
                "period_text": "1-3月",
                "standard_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "value": 150.0,
                "yoy_growth": 1.5,
                "source_document_id": "doc-mar",
            },
        ]
        derived, issues = build_derived_metrics(records, scope_start_month="2025-02", scope_end_month="2025-03")
        self.assertEqual(issues, [])
        self.assertEqual([item["metric_month"] for item in derived], ["2025-01", "2025-02", "2025-03"])

    def test_jan_feb_average_formula_display(self):
        exporter = FiscalExporter()
        formatted = exporter._format_derived_row(
            {
                "metric_month": "2025-01",
                "period_label": "2025-01",
                "standard_metric_name": "全国一般公共预算收入",
                "metric_category": "一般公共预算收入",
                "value": 50.0,
                "unit": "亿元",
                "yoy_growth": 1.0,
                "derived_type": "average_from_jan_feb_combined",
                "formula": "2025-01~02=100.0/2",
                "source_record_ids": "r1",
                "source_document_ids": "d1",
                "is_derived_metric": True,
                "remark": "1-2月合并值按平均值拆分为1月和2月",
            }
        )
        self.assertEqual(formatted["formula"], "202501 = 202501～02 / 2 = 100.0/2")
        self.assertEqual(formatted["derived_type"], "1-2月平均值")
        self.assertEqual(formatted["remark"], "1-2月合并值按平均值拆分为1月和2月")


if __name__ == "__main__":
    unittest.main()
