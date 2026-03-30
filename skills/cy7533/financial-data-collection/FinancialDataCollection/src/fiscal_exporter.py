from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .fiscal_parser import METRIC_SPECS, build_derived_period_display, build_source_period_key


RAW_DOCUMENT_HEADERS: Dict[str, str] = {
    "title": "公告标题",
    "url": "公告链接",
    "publish_date": "发布时间",
    "source_department": "来源部门",
    "content": "公告正文",
    "period": "期间",
    "remark": "备注",
}

EXTRACTED_METRIC_HEADERS: Dict[str, str] = {
    "period": "期间",
    "metric_name": "指标（单位：亿元）",
    "metric_category": "指标分类",
    "direction": "收支方向",
    "value": "指标值",
    "unit": "单位",
    "yoy_growth": "同比增速",
    "source_title": "来源公告标题",
    "source_url": "来源公告链接",
    "remark": "备注",
}

DERIVED_METRIC_HEADERS: Dict[str, str] = {
    "period": "期间",
    "metric_name": "指标（单位：亿元）",
    "direction": "收支方向",
    "metric_category": "指标分类",
    "value": "指标值",
    "unit": "单位",
    "yoy_growth": "同比增速",
    "derived_type": "推导类型",
    "formula": "计算公式",
    "is_derived_metric": "是否推导",
    "remark": "备注",
}

MONTHLY_SUMMARY_FILENAME = "monthly_summary.xlsx"

METRIC_CLASSIFICATION = {
    metric_name: {"direction": direction, "category": category}
    for metric_name, direction, category in METRIC_SPECS
}
METRIC_ORDER = {
    metric_name: index
    for index, (metric_name, _, _) in enumerate(METRIC_SPECS, start=1)
}
METRIC_ORDER["窄口径赤字"] = len(METRIC_ORDER) + 1
METRIC_ORDER["宽口径赤字"] = len(METRIC_ORDER) + 2


class FiscalExporter:
    def __init__(self, output_dir: str = "output") -> None:
        self.output_dir = Path(output_dir)

    def export(
        self,
        documents: List[dict],
        extracted_metrics: List[dict],
        derived_metrics: List[dict],
        allowed_periods: List[str],
        run_context: Dict[str, object],
        issues: List[dict],
    ) -> None:
        allowed = set(allowed_periods)
        document_rows = [self._format_raw_document_row(record) for record in documents]
        extracted_rows = sorted(
            [self._format_extracted_row(record, documents) for record in extracted_metrics],
            key=lambda item: (item["period_key"], self._metric_sort_order(item["metric_name"])),
        )
        derived_rows = sorted(
            [self._format_derived_row(record) for record in derived_metrics],
            key=lambda item: (self._period_sort_key(item["period"]), self._metric_sort_order(item["metric_name"])),
        )

        self._export_grouped_by_period(document_rows, "period_key", "raw_documents.xlsx", allowed, RAW_DOCUMENT_HEADERS)
        self._export_grouped_by_period(extracted_rows, "period_key", "extracted_metrics.xlsx", allowed, EXTRACTED_METRIC_HEADERS)
        self._export_run_summary(
            documents=document_rows,
            extracted_metrics=extracted_rows,
            derived_metrics=derived_rows,
            issues=issues,
            allowed_periods=allowed_periods,
            run_context=run_context,
        )

    def _export_grouped_by_period(
        self,
        records: List[dict],
        period_field: str,
        filename: str,
        allowed_months: set,
        header_map: Dict[str, str],
    ) -> None:
        grouped = {}
        for record in records:
            period = record.get(period_field)
            if not period or (allowed_months and period not in allowed_months):
                continue
            grouped.setdefault(period, []).append(record)
        for period, period_records in grouped.items():
            period_dir = self.output_dir / period
            period_dir.mkdir(parents=True, exist_ok=True)
            self._write_excel(period_dir / filename, period_records, header_map)

    def _export_run_summary(
        self,
        documents: List[dict],
        extracted_metrics: List[dict],
        derived_metrics: List[dict],
        issues: List[dict],
        allowed_periods: List[str],
        run_context: Dict[str, object],
    ) -> None:
        run_timestamp = str(run_context["run_timestamp"])
        summary_root_dir = self.output_dir / self._build_run_summary_root_dirname(
            allowed_periods,
            run_context.get("requested_period_range"),
            str(run_context.get("run_mode", "")),
        )
        run_dir = summary_root_dir / run_timestamp
        run_dir.mkdir(parents=True, exist_ok=True)

        self._write_excel(run_dir / self._timestamped_filename("raw_documents", "xlsx", run_timestamp), documents, RAW_DOCUMENT_HEADERS)
        self._write_excel(run_dir / self._timestamped_filename("extracted_metrics", "xlsx", run_timestamp), extracted_metrics, EXTRACTED_METRIC_HEADERS)
        self._write_excel(run_dir / self._timestamped_filename("derived_metrics", "xlsx", run_timestamp), derived_metrics, DERIVED_METRIC_HEADERS)
        self._write_monthly_summary(run_dir / self._timestamped_filename("monthly_summary", "xlsx", run_timestamp), derived_metrics)

        report_text = self._build_markdown_report(
            documents=documents,
            extracted_metrics=extracted_metrics,
            derived_metrics=derived_metrics,
            issues=issues,
            allowed_periods=allowed_periods,
            run_context=run_context,
            run_dir=run_dir,
        )
        (run_dir / self._timestamped_filename("运行汇报", "md", run_timestamp)).write_text(report_text, encoding="utf-8")

    def _write_excel(self, file_path: Path, records: Sequence[dict], header_map: Dict[str, str]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "data"
        if records:
            columns = [column for column in records[0].keys() if column in header_map]
        else:
            columns = list(header_map.keys())
        sheet.append([header_map[column] for column in columns])
        for record in records:
            sheet.append([record.get(column) for column in columns])
        self._format_sheet(sheet)
        workbook.save(file_path)

    def _build_run_summary_root_dirname(
        self,
        allowed_periods: List[str],
        requested_period_range: object,
        run_mode: str,
    ) -> str:
        if run_mode == "全量":
            return "全量汇总"
        if requested_period_range:
            return f"{requested_period_range}_汇总"
        if not allowed_periods:
            return "empty_empty_汇总"
        return f"{allowed_periods[0]}_{allowed_periods[-1]}_汇总"

    def _timestamped_filename(self, stem: str, suffix: str, run_timestamp: str) -> str:
        return f"{stem}_{run_timestamp}.{suffix}"

    def _build_markdown_report(
        self,
        documents: List[dict],
        extracted_metrics: List[dict],
        derived_metrics: List[dict],
        issues: List[dict],
        allowed_periods: List[str],
        run_context: Dict[str, object],
        run_dir: Path,
    ) -> str:
        started_at = run_context["run_started_at"]
        finished_at = run_context["run_finished_at"]
        duration_seconds = (finished_at - started_at).total_seconds()
        issue_summary = self._group_issues(issues)

        lines = [
            "# 财政数据导出运行汇报",
            "",
            "## 运行信息",
            f"- 运行开始时间：{started_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"- 运行结束时间：{finished_at.strftime('%Y-%m-%d %H:%M:%S')}",
            f"- 运行耗时：{duration_seconds:.2f} 秒",
            f"- 运行模式：{run_context['run_mode']}",
            f"- 命令参数：{run_context['command_summary']}",
            f"- 输出根目录：{self.output_dir}",
            f"- 汇总目录名称：{run_dir.parent.name}/{run_dir.name}",
            "",
            "## 导出范围",
            f"- 请求范围：{run_context['requested_scope']}",
            f"- 实际导出期间列表：{', '.join(allowed_periods) if allowed_periods else '无'}",
            f"- 起始期间：{allowed_periods[0] if allowed_periods else 'empty'}",
            f"- 结束期间：{allowed_periods[-1] if allowed_periods else 'empty'}",
            f"- 复用区间数：{run_context.get('reused_period_count', 0)}",
            f"- 重爬区间数：{run_context.get('refetched_period_count', 0)}",
            f"- 因校验失败重爬的区间：{', '.join(run_context.get('refetched_periods', [])) or '无'}",
            "",
            "## 数据结果",
            f"- 命中文档数：{len(documents)}",
            f"- 原始提取指标数：{len(extracted_metrics)}",
            f"- 推导结果数：{len(derived_metrics)}",
            f"- 异常数：{len(issues)}",
            "",
            "## 文件清单",
            "- 汇总目录文件：",
            f"  - {run_dir.parent.name}/{run_dir.name}/{self._timestamped_filename('raw_documents', 'xlsx', str(run_context['run_timestamp']))}",
            f"  - {run_dir.parent.name}/{run_dir.name}/{self._timestamped_filename('extracted_metrics', 'xlsx', str(run_context['run_timestamp']))}",
            f"  - {run_dir.parent.name}/{run_dir.name}/{self._timestamped_filename('derived_metrics', 'xlsx', str(run_context['run_timestamp']))}",
            f"  - {run_dir.parent.name}/{run_dir.name}/{self._timestamped_filename('monthly_summary', 'xlsx', str(run_context['run_timestamp']))}",
            f"  - {run_dir.parent.name}/{run_dir.name}/{self._timestamped_filename('运行汇报', 'md', str(run_context['run_timestamp']))}",
            f"- 本次生成的按期间目录：{', '.join(allowed_periods) if allowed_periods else '无'}",
            "",
            "## 异常摘要",
        ]

        if not issues:
            lines.append("无异常")
        else:
            for month in sorted(issue_summary):
                month_issues = issue_summary[month]
                lines.append(f"- {month}：{len(month_issues)} 条")
                for issue in month_issues[:5]:
                    lines.append(f"  - {issue['standard_metric_name']}：{issue['detail']}")

        lines.extend(
            [
                "",
                "## 备注",
                "- 1-2月作为合并期处理，原始层期间统一使用 YYYYMM-YYYYMM，推导层期间使用 YYYYMM～MM。",
                "- 一季度、上半年、前三季度、全年标题会分别标准化为 1-3月、1-6月、1-9月、1-12月口径。",
                "- monthly_summary.xlsx 基于推导层生成，按指标和月份展开。",
                "",
            ]
        )
        return "\n".join(lines)

    def _group_issues(self, issues: List[dict]) -> Dict[str, List[dict]]:
        grouped: Dict[str, List[dict]] = defaultdict(list)
        for issue in issues:
            grouped[issue["metric_month"]].append(issue)
        return grouped

    def _format_raw_document_row(self, record: dict) -> dict:
        return {
            "period_key": build_source_period_key(record["period_text"], record["cutoff_month"]),
            "title": record["title"],
            "url": record["url"],
            "publish_date": record["publish_date"],
            "source_department": record["source_department"],
            "content": record["content"],
            "period": build_source_period_key(record["period_text"], record["cutoff_month"]),
            "remark": "",
        }

    def _format_extracted_row(self, record: dict, documents: List[dict]) -> dict:
        source_url = ""
        for doc in documents:
            if doc["document_id"] == record["source_document_id"]:
                source_url = doc["url"]
                break
        remark = ""
        if record["source_metric_name"] != record["standard_metric_name"]:
            remark = f"原文指标：{record['source_metric_name']}"
        return {
            "period_key": build_source_period_key(record["period_text"], record["cutoff_month"]),
            "period": build_source_period_key(record["period_text"], record["cutoff_month"]),
            "metric_name": record["standard_metric_name"],
            "metric_category": record["metric_category"],
            "direction": record["direction"],
            "value": record["value"],
            "unit": record["unit"],
            "yoy_growth": record["yoy_growth"],
            "source_title": record["source_title"],
            "source_url": source_url,
            "remark": remark,
        }

    def _format_derived_row(self, record: dict) -> dict:
        direction, _ = self._classify_metric(record["standard_metric_name"])
        return {
            "period_key": record["period_label"],
            "period": build_derived_period_display(record["period_label"]),
            "metric_name": record["standard_metric_name"],
            "direction": direction,
            "metric_category": record["metric_category"],
            "value": record["value"],
            "unit": record["unit"],
            "yoy_growth": record["yoy_growth"],
            "derived_type": self._display_derived_type(record["derived_type"]),
            "formula": self._display_formula(record),
            "is_derived_metric": "是" if record["is_derived_metric"] else "否",
            "remark": record.get("remark", ""),
        }

    def _display_derived_type(self, derived_type: str) -> str:
        mapping = {
            "combined_jan_feb_cumulative": "1-2月合并值",
            "average_from_jan_feb_combined": "1-2月平均值",
            "single_month_from_january_cumulative": "1月累计值",
            "monthly_from_cumulative_difference": "累计差值推导",
            "deficit_metric": "赤字派生指标",
        }
        return mapping.get(derived_type, derived_type)

    def _display_formula(self, record: dict) -> str:
        period = build_derived_period_display(record["period_label"])
        formula = str(record.get("formula", ""))
        value = record.get("value")
        if record["derived_type"] == "average_from_jan_feb_combined":
            year, month = record["metric_month"].split("-")
            return f"{year}{month} = {year}01～02 / 2 = {round(float(value) * 2, 6)}/2"
        if record["derived_type"] == "combined_jan_feb_cumulative":
            return f"{period} = {value}"
        if record["derived_type"] == "single_month_from_january_cumulative":
            return f"{period} = {value}"
        if record["derived_type"] == "monthly_from_cumulative_difference":
            match = formula.split("=")[-1]
            if "-" in match:
                left, right = match.split("-", 1)
                year, month = record["metric_month"].split("-")
                prev_month = f"{int(month)-1:02d}"
                return f"{year}_{month} = ({year}_01-{year}_{month}) - ({year}_01-{year}_{prev_month}) = {left}-{right}"
        if record["derived_type"] == "deficit_metric":
            if record["standard_metric_name"] == "窄口径赤字":
                return f"{period} = 全国一般公共预算支出 - 全国一般公共预算收入 = {value}"
            if record["standard_metric_name"] == "宽口径赤字":
                return f"{period} = (全国一般公共预算支出 + 全国政府性基金预算支出) - (全国一般公共预算收入 + 全国政府性基金预算收入) = {value}"
            return f"{period} = {value}"
        return formula

    def _write_monthly_summary(self, file_path: Path, derived_metrics: Sequence[dict]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "summary"

        periods = self._sort_periods(
            expanded
            for record in derived_metrics
            if record.get("period")
            for expanded in self._expand_summary_periods(str(record["period"]))
        )
        header = ["收支大类", "指标归类", "指标（单位：亿元）", *periods]
        sheet.append(header)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        rows = self._build_monthly_summary_rows(derived_metrics, periods)
        for row in rows:
            sheet.append(row)

        if rows:
            self._merge_monthly_summary_groups(sheet, rows)
        sheet.freeze_panes = "D2"
        self._format_sheet(sheet)
        workbook.save(file_path)

    def _format_sheet(self, sheet) -> None:
        center_alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is None:
                    continue
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    def _build_monthly_summary_rows(self, derived_metrics: Sequence[dict], periods: Sequence[str]) -> List[List[object]]:
        values_by_metric = {}
        for record in derived_metrics:
            metric_name = str(record["metric_name"])
            direction, category = self._classify_metric(metric_name)
            key = (direction, category, metric_name)
            period_values = values_by_metric.setdefault(key, {})
            for expanded_period, expanded_value in self._expand_summary_period_values(str(record["period"]), record["value"]):
                period_values[expanded_period] = expanded_value

        direction_order = {"财政收入": 0, "财政支出": 1, "赤字": 2}
        sorted_keys = sorted(values_by_metric, key=lambda item: (direction_order.get(item[0], 99), self._metric_sort_order(item[2])))
        rows = []
        for key in sorted_keys:
            direction, category, metric_name = key
            period_values = values_by_metric[key]
            rows.append([direction, category, metric_name, *[period_values.get(period) for period in periods]])
        return rows

    def _classify_metric(self, metric_name: str) -> tuple[str, str]:
        if metric_name == "窄口径赤字":
            return "赤字", "窄口径赤字"
        if metric_name == "宽口径赤字":
            return "赤字", "宽口径赤字"
        info = METRIC_CLASSIFICATION.get(metric_name)
        if not info:
            return "其他", "其他"
        return str(info["direction"]), str(info["category"])

    def _sort_periods(self, periods: Sequence[str]) -> List[str]:
        unique_periods = sorted(set(periods), key=self._period_sort_key)
        return list(unique_periods)

    def _metric_sort_order(self, metric_name: str) -> int:
        return METRIC_ORDER.get(metric_name, 999999)

    def _expand_summary_periods(self, period: str) -> List[str]:
        if "～" not in period:
            return [period]
        start, end = period.split("～", 1)
        year = start[:4]
        start_month = int(start[4:6])
        end_month = int(end)
        return [f"{year}{month:02d}" for month in range(start_month, end_month + 1)]

    def _expand_summary_period_values(self, period: str, value: object) -> List[tuple[str, object]]:
        expanded_periods = self._expand_summary_periods(period)
        if len(expanded_periods) <= 1:
            return [(expanded_periods[0], value)]
        if value is None:
            return [(expanded_period, None) for expanded_period in expanded_periods]
        avg_value = round(float(value) / len(expanded_periods), 6)
        return [(expanded_period, avg_value) for expanded_period in expanded_periods]

    def _period_sort_key(self, period: str) -> tuple[int, int]:
        if "～" in period:
            left, right = period.split("～", 1)
            return int(left[:4]), int(right)
        return int(period[:4]), int(period[4:6])

    def _merge_monthly_summary_groups(self, sheet, rows: Sequence[Sequence[object]]) -> None:
        self._merge_ranges(sheet, self._build_merge_ranges(rows, value_index=0))
        self._merge_ranges(sheet, self._build_merge_ranges(rows, value_index=1, group_index=0))

    def _build_merge_ranges(
        self,
        rows: Sequence[Sequence[object]],
        value_index: int,
        group_index: int | None = None,
    ) -> List[tuple[int, int, int]]:
        ranges: List[tuple[int, int, int]] = []
        start = 2
        prev_value = rows[0][value_index]
        prev_group = rows[0][group_index] if group_index is not None else None

        for idx in range(1, len(rows) + 1):
            value = rows[idx][value_index] if idx < len(rows) else None
            group = rows[idx][group_index] if group_index is not None and idx < len(rows) else None
            if value == prev_value and (group_index is None or group == prev_group):
                continue
            end_row = idx + 1
            if prev_value not in (None, "") and end_row - start > 0:
                ranges.append((start, end_row, value_index + 1))
            start = idx + 2
            prev_value = value
            prev_group = group
        return ranges

    def _merge_ranges(self, sheet, ranges: Sequence[tuple[int, int, int]]) -> None:
        for start_row, end_row, column in ranges:
            sheet.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)
