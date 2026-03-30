from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="财政部财政收支数据爬取与导出")
    parser.add_argument("--start-month", help="起始月份，例如 2025-01")
    parser.add_argument("--end-month", help="结束月份，例如 2025-03")
    parser.add_argument("--output-dir", default="output", help="输出目录")
    args = parser.parse_args()
    if args.start_month and len(args.start_month) != 7:
        parser.error("--start-month 格式必须为 YYYY-MM")
    if args.end_month and len(args.end_month) != 7:
        parser.error("--end-month 格式必须为 YYYY-MM")
    if args.start_month and not args.end_month:
        args.end_month = args.start_month
    if args.end_month and not args.start_month:
        args.start_month = args.end_month
    if args.start_month and args.end_month and args.start_month > args.end_month:
        parser.error("--start-month 不能晚于 --end-month")
    return args


def main() -> int:
    run_started_at = datetime.now()
    args = parse_args()

    from src.fiscal_crawler import CrawlScope, FiscalCrawler
    from src.fiscal_exporter import FiscalExporter
    from src.fiscal_parser import (
        DocumentMeta,
        build_period_label,
        build_source_period_key,
        extract_metrics_from_content,
        parse_period_from_title,
    )
    from src.fiscal_transform import build_derived_metrics

    scope = CrawlScope(start_month=args.start_month, end_month=args.end_month)
    crawler = FiscalCrawler()
    exporter = FiscalExporter(output_dir=args.output_dir)

    print("开始检查缓存并抓取公告...")
    documents = []
    extracted_metrics = []
    reused_periods = []
    refetched_periods = []
    for entry in crawler.iter_document_entries():
        try:
            period_text, cutoff_month = parse_period_from_title(entry["title"])
        except Exception:
            continue
        if not scope.should_fetch(cutoff_month):
            continue
        period_key = build_source_period_key(period_text, cutoff_month)
        cache_dir = Path(args.output_dir) / period_key
        if _is_period_cache_valid(cache_dir):
            cached_documents = _load_cached_documents(cache_dir)
            cached_metrics = _load_cached_extracted_metrics(cache_dir)
            documents.extend(cached_documents)
            extracted_metrics.extend(cached_metrics)
            reused_periods.append(period_key)
            continue

        meta = crawler.fetch_document(entry["url"])
        document = _meta_to_record(meta)
        documents.append(document)
        extracted_metrics.extend(extract_metrics_from_content(document["document_id"], meta))
        refetched_periods.append(period_key)

    documents = _dedupe_documents(documents)
    extracted_metrics = _dedupe_metrics(extracted_metrics)
    print(f"抓取完成，命中文档 {len(documents)} 篇")
    print(f"指标提取完成，共 {len(extracted_metrics)} 条")

    derived_metrics, issues = build_derived_metrics(
        extracted_metrics,
        scope_year=None,
        scope_month=None,
        scope_start_month=scope.normalized_start_month(),
        scope_end_month=scope.normalized_end_month(),
    )
    print(f"推导完成，共 {len(derived_metrics)} 条，异常 {len(issues)} 条")

    output_periods = _collect_output_periods(documents, extracted_metrics, scope)
    run_finished_at = datetime.now()
    run_context = {
        "run_started_at": run_started_at,
        "run_finished_at": run_finished_at,
        "run_timestamp": run_finished_at.strftime("%Y%m%d%H%M%S"),
        "requested_scope": _build_requested_scope(args),
        "run_mode": _build_run_mode(args),
        "command_summary": _build_command_summary(args),
        "requested_period_range": scope.requested_period_range_key(),
        "reused_period_count": len(sorted(set(reused_periods))),
        "refetched_period_count": len(sorted(set(refetched_periods))),
        "refetched_periods": sorted(set(refetched_periods)),
    }
    exporter.export(
        documents,
        extracted_metrics,
        derived_metrics,
        allowed_periods=output_periods,
        run_context=run_context,
        issues=issues,
    )
    print(f"导出完成，输出目录: {args.output_dir}")

    if issues:
        for issue in issues[:20]:
            print(f"[WARN] {issue['metric_month']} {issue['standard_metric_name']} - {issue['detail']}")
    return 0


def _parse_date(value: str):
    from datetime import date

    return date.fromisoformat(value)


def _collect_output_periods(documents, extracted_metrics, scope) -> list:
    from src.fiscal_parser import build_source_period_key

    periods = set()
    fetch_start = scope.fetch_start_month()
    fetch_end = scope.normalized_end_month()
    for record in documents:
        if fetch_start and fetch_end:
            if fetch_start <= record["cutoff_month"] <= fetch_end:
                periods.add(build_source_period_key(record["period_text"], record["cutoff_month"]))
            continue
        if scope.should_output(record["cutoff_month"], record.get("period_label")):
            periods.add(build_source_period_key(record["period_text"], record["cutoff_month"]))
    for record in extracted_metrics:
        if fetch_start and fetch_end:
            if fetch_start <= record["metric_month"] <= fetch_end:
                periods.add(build_source_period_key(record["period_text"], record["cutoff_month"]))
            continue
        if scope.should_output(record["metric_month"], record.get("period_label")):
            periods.add(build_source_period_key(record["period_text"], record["cutoff_month"]))
    return sorted(periods)


def _build_requested_scope(args: argparse.Namespace) -> str:
    if args.start_month and args.end_month:
        return f"按区间 {args.start_month} 至 {args.end_month}"
    return "全量"


def _build_run_mode(args: argparse.Namespace) -> str:
    if args.start_month and args.end_month:
        return "按区间"
    return "全量"


def _build_command_summary(args: argparse.Namespace) -> str:
    parts = ["python scripts/run_pipeline.py"]
    if args.start_month:
        parts.extend(["--start-month", args.start_month])
    if args.end_month:
        parts.extend(["--end-month", args.end_month])
    if args.output_dir != "output":
        parts.extend(["--output-dir", args.output_dir])
    return " ".join(parts)


def _is_period_cache_valid(period_dir: Path) -> bool:
    raw_file = period_dir / "raw_documents.xlsx"
    metrics_file = period_dir / "extracted_metrics.xlsx"
    if not period_dir.exists() or not raw_file.exists() or not metrics_file.exists():
        return False
    return _count_data_rows(metrics_file) >= 40


def _count_data_rows(file_path: Path) -> int:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    try:
        rows = list(sheet.iter_rows(values_only=True))
        return max(len(rows) - 1, 0)
    finally:
        workbook.close()


def _load_cached_documents(period_dir: Path) -> list:
    from src.fiscal_parser import build_period_label, parse_period_from_title

    workbook = load_workbook(period_dir / "raw_documents.xlsx", read_only=True, data_only=True)
    sheet = workbook.active
    try:
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header = rows[0]
    index = {name: idx for idx, name in enumerate(header)}
    documents = []
    for row in rows[1:]:
        title = row[index["公告标题"]]
        if not title:
            continue
        period_text, cutoff_month = parse_period_from_title(str(title))
        documents.append(
            {
                "document_id": str(title),
                "title": str(title),
                "url": row[index["公告链接"]] or "",
                "publish_date": str(row[index["发布时间"]]),
                "source_department": row[index["来源部门"]] or "",
                "content": row[index["公告正文"]] or "",
                "period_text": period_text,
                "cutoff_month": cutoff_month,
                "period_label": build_period_label(period_text, cutoff_month),
            }
        )
    return documents


def _load_cached_extracted_metrics(period_dir: Path) -> list:
    from src.fiscal_parser import build_period_label, parse_period_from_title

    workbook = load_workbook(period_dir / "extracted_metrics.xlsx", read_only=True, data_only=True)
    sheet = workbook.active
    try:
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header = rows[0]
    index = {name: idx for idx, name in enumerate(header)}
    metric_header = "指标（单位：亿元）" if "指标（单位：亿元）" in index else "指标"
    metrics = []
    for row in rows[1:]:
        source_title = row[index["来源公告标题"]]
        metric_name = row[index[metric_header]]
        if not source_title or not metric_name:
            continue
        period_text, cutoff_month = parse_period_from_title(str(source_title))
        remark = row[index.get("备注", -1)] if "备注" in index else ""
        source_metric_name = str(metric_name)
        if remark and str(remark).startswith("原文指标："):
            source_metric_name = str(remark).split("：", 1)[1]
        metrics.append(
            {
                "record_id": f"{source_title}:{metric_name}",
                "metric_month": cutoff_month,
                "period_label": build_period_label(period_text, cutoff_month),
                "standard_metric_name": str(metric_name),
                "source_metric_name": source_metric_name,
                "metric_category": row[index["指标分类"]] or "",
                "direction": row[index["收支方向"]] or "",
                "value": float(row[index["指标值"]]),
                "unit": row[index["单位"]] or "亿元",
                "yoy_growth": row[index["同比增速"]],
                "source_document_id": str(source_title),
                "source_title": str(source_title),
                "source_publish_date": "",
                "period_text": period_text,
                "cutoff_month": cutoff_month,
                "source_order": len(metrics) + 1,
            }
        )
    return metrics


def _meta_to_record(meta) -> dict:
    from src.fiscal_parser import build_period_label

    return {
        "document_id": meta.title,
        "title": meta.title,
        "url": meta.url,
        "publish_date": meta.publish_date.isoformat(),
        "source_department": meta.source_department,
        "content": meta.content,
        "period_text": meta.period_text,
        "cutoff_month": meta.cutoff_month,
        "period_label": build_period_label(meta.period_text, meta.cutoff_month),
    }


def _dedupe_documents(documents: list) -> list:
    by_title = {}
    for record in documents:
        by_title[record["title"]] = record
    return sorted(by_title.values(), key=lambda item: item["publish_date"])


def _dedupe_metrics(metrics: list) -> list:
    by_key = {}
    for record in metrics:
        by_key[record["record_id"]] = record
    return sorted(by_key.values(), key=lambda item: (item["metric_month"], item.get("source_order", 999999), item["standard_metric_name"]))


if __name__ == "__main__":
    raise SystemExit(main())
