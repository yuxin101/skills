"""
CMB Salary Batch Import Script
Imports net salary data from salary Excel file into CMB AgencyPayment bank template.
Preserves all original formats, formulas, data validations, and styles.
"""
import shutil
import subprocess
import os
import openpyxl
from lxml import etree

def import_salary(salary_file, template_file, output_file,
                   salary_sheet="工资表",
                   name_col=2, card_col=4, bank_col=5, salary_col=23,
                   data_start_row=4, data_end_row=12,
                   template_data_start_row=3):
    """
    Import salary data into CMB bank template.

    Args:
        salary_file: Path to salary Excel file
        template_file: Path to CMB AgencyPayment template
        output_file: Path to output file
        salary_sheet: Sheet name in salary file (default: 工资表)
        name_col: Column number for employee name (B=2, default)
        card_col: Column number for bank card (D=4, default)
        bank_col: Column number for bank name (E=5, default)
        salary_col: Column number for net salary (W=23, default)
        data_start_row: First data row in salary sheet (default: 4)
        data_end_row: Last data row in salary sheet (default: 12)
        template_data_start_row: First data row in template (default: 3)
    """
    # === 1. Read salary data ===
    salary_wb = openpyxl.load_workbook(salary_file, data_only=True)
    salary_ws = salary_wb[salary_sheet]

    employees = []
    for row in range(data_start_row, data_end_row + 1):
        name = salary_ws.cell(row=row, column=name_col).value
        bank_card = salary_ws.cell(row=row, column=card_col).value
        bank_name = salary_ws.cell(row=row, column=bank_col).value
        net_salary = salary_ws.cell(row=row, column=salary_col).value

        if name and net_salary is not None:
            try:
                net_salary = float(net_salary)
                if net_salary == 0:
                    continue
            except (ValueError, TypeError):
                continue

            employees.append({
                'name': str(name).strip(),
                'bank_card': str(bank_card).strip() if bank_card else '',
                'bank_name': str(bank_name).strip() if bank_name else '',
                'net_salary': net_salary
            })
    salary_wb.close()

    if not employees:
        print("No employee data found.")
        return False

    print(f"Read {len(employees)} employees")
    for e in employees:
        print(f"  {e['name']}: {e['net_salary']:.2f} | {e['bank_card']}")

    # === 2. Copy template to output ===
    shutil.copy2(template_file, output_file)

    # === 3. Unpack xlsx ===
    unpack_dir = "/tmp/cmb_import_work/"
    os.makedirs(unpack_dir, exist_ok=True)
    result = subprocess.run(
        ["python3", "/mnt/c/Users/70426/.openclaw/skills/minimax-xlsx/scripts/xlsx_unpack.py",
         output_file, unpack_dir],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"Unpack failed: {result.stderr}")
        return False

    # === 4. Update sharedStrings ===
    ns_main = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ss_path = unpack_dir + "xl/sharedStrings.xml"

    parser = etree.XMLParser(remove_blank_text=False)
    ss_tree = etree.parse(ss_path, parser)
    ss_root = ss_tree.getroot()

    existing_strings = []
    for si in ss_root.findall(f'{{{ns_main}}}si'):
        t = si.find(f'{{{ns_main}}}t')
        if t is not None and t.text:
            existing_strings.append(t.text)
        else:
            parts = [r.text for r in si.findall(f'.//{{{ns_main}}}t') if r.text]
            existing_strings.append(''.join(parts))

    def get_or_add_string(s):
        if s in existing_strings:
            return existing_strings.index(s)
        idx = len(existing_strings)
        existing_strings.append(s)
        si = etree.SubElement(ss_root, f'{{{ns_main}}}si')
        t = etree.SubElement(si, f'{{{ns_main}}}t')
        t.text = s
        ss_root.set('count', str(len(existing_strings)))
        ss_root.set('uniqueCount', str(len(existing_strings)))
        return idx

    for emp in employees:
        get_or_add_string(emp['name'])
        get_or_add_string(emp['bank_card'])
        if emp['bank_name']:
            get_or_add_string(emp['bank_name'])

    ss_tree.write(ss_path, encoding='utf-8', xml_declaration=True)

    # === 5. Edit sheet1.xml ===
    sheet_path = unpack_dir + "xl/worksheets/sheet1.xml"
    tree = etree.parse(sheet_path, parser)
    root = tree.getroot()

    sheet_data = root.find(f'{{{ns_main}}}sheetData')

    rows_by_num = {}
    for row_el in sheet_data.findall(f'{{{ns_main}}}row'):
        r = int(row_el.get('r'))
        rows_by_num[r] = row_el

    def get_or_create_cell(row_el, col_letter, row_num):
        addr = f"{col_letter}{row_num}"
        for c in row_el.findall(f'{{{ns_main}}}c'):
            if c.get('r') == addr:
                return c
        col_order = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        new_c = etree.Element(f'{{{ns_main}}}c')
        new_c.set('r', addr)
        insert_pos = len(row_el)
        for i, c in enumerate(row_el.findall(f'{{{ns_main}}}c')):
            c_col = ''.join(filter(str.isalpha, c.get('r')))
            if col_order.index(c_col) > col_order.index(col_letter):
                insert_pos = list(row_el).index(c)
                break
        row_el.insert(insert_pos, new_c)
        return new_c

    def set_string_cell(cell_el, value_idx):
        cell_el.set('t', 's')
        for v in cell_el.findall(f'{{{ns_main}}}v'):
            cell_el.remove(v)
        v = etree.SubElement(cell_el, f'{{{ns_main}}}v')
        v.text = str(value_idx)

    def set_numeric_cell(cell_el, value):
        if 't' in cell_el.attrib:
            del cell_el.attrib['t']
        for v in cell_el.findall(f'{{{ns_main}}}v'):
            cell_el.remove(v)
        v = etree.SubElement(cell_el, f'{{{ns_main}}}v')
        v.text = str(value)

    for i, emp in enumerate(employees):
        excel_row = template_data_start_row + i

        row_el = rows_by_num.get(excel_row)
        if row_el is None:
            row_el = etree.SubElement(sheet_data, f'{{{ns_main}}}row')
            row_el.set('r', str(excel_row))
            rows_by_num[excel_row] = row_el

        # A: bank card
        c = get_or_create_cell(row_el, 'A', excel_row)
        set_string_cell(c, existing_strings.index(emp['bank_card']))

        # B: name
        c = get_or_create_cell(row_el, 'B', excel_row)
        set_string_cell(c, existing_strings.index(emp['name']))

        # C: net salary (numeric)
        c = get_or_create_cell(row_el, 'C', excel_row)
        set_numeric_cell(c, emp['net_salary'])

        # D: bank name
        if emp['bank_name']:
            c = get_or_create_cell(row_el, 'D', excel_row)
            set_string_cell(c, existing_strings.index(emp['bank_name']))

        print(f"Row {excel_row}: {emp['name']} | {emp['bank_card']} | {emp['net_salary']:.2f}")

    tree.write(sheet_path, encoding='utf-8', xml_declaration=True)

    # === 6. Pack ===
    result = subprocess.run(
        ["python3", "/mnt/c/Users/70426/.openclaw/skills/minimax-xlsx/scripts/xlsx_pack.py",
         unpack_dir, output_file],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"Pack failed: {result.stderr}")
        return False

    print(f"\nOutput: {output_file}")
    return True


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 4:
        print("Usage: python3 import_salary.py <salary_file> <template_file> <output_file>")
        sys.exit(1)

    salary_file = sys.argv[1]
    template_file = sys.argv[2]
    output_file = sys.argv[3]

    ok = import_salary(salary_file, template_file, output_file)
    sys.exit(0 if ok else 1)
