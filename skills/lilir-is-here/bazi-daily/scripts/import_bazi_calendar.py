#!/usr/bin/env python3
"""Convert Bazi calendar xlsx into upsert SQL for bazi_daily_calendar."""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


HEADER_ALIASES = {
    "date": ("公历年月日", "日期", "date"),
    "flow_year": ("流年", "flow_year"),
    "flow_month": ("流月", "flow_month"),
    "flow_day": ("流日", "flow_day"),
}


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def find_header_row(sheet) -> Tuple[int, Dict[str, int]]:
    for row_idx in range(1, 51):
        values = [normalize_header(c.value) for c in sheet[row_idx]]
        mapping: Dict[str, int] = {}
        for idx, val in enumerate(values, start=1):
            for field, aliases in HEADER_ALIASES.items():
                if val in {a.lower() for a in aliases}:
                    mapping[field] = idx
        if all(k in mapping for k in HEADER_ALIASES):
            return row_idx, mapping
    raise ValueError("Cannot find header row with date/flow_year/flow_month/flow_day")


def parse_date(value: object) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    # Accept values like 2026/03/03 or 2026-03-03
    text = text.replace("/", "-")
    return dt.date.fromisoformat(text).isoformat()


def sql_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def extract_records(sheet, header_row: int, mapping: Dict[str, int]) -> List[Dict[str, str]]:
    records: List[Dict[str, str]] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        raw_date = sheet.cell(row_idx, mapping["date"]).value
        raw_year = sheet.cell(row_idx, mapping["flow_year"]).value
        raw_month = sheet.cell(row_idx, mapping["flow_month"]).value
        raw_day = sheet.cell(row_idx, mapping["flow_day"]).value

        date_text = parse_date(raw_date)
        year_text = "" if raw_year is None else str(raw_year).strip()
        month_text = "" if raw_month is None else str(raw_month).strip()
        day_text = "" if raw_day is None else str(raw_day).strip()

        if not any((date_text, year_text, month_text, day_text)):
            continue
        if not all((date_text, year_text, month_text, day_text)):
            raise ValueError(f"Incomplete row at line {row_idx}")

        records.append(
            {
                "date": date_text,
                "flow_year": year_text,
                "flow_month": month_text,
                "flow_day": day_text,
            }
        )
    return records


def validate_records(records: List[Dict[str, str]]) -> None:
    seen = set()
    for r in records:
        d = r["date"]
        if d in seen:
            raise ValueError(f"Duplicate date found: {d}")
        seen.add(d)


def build_sql(records: List[Dict[str, str]], table: str) -> str:
    today = dt.datetime.now(dt.timezone.utc).isoformat()
    lines = [
        "BEGIN;",
        f"CREATE TABLE IF NOT EXISTS {table} (",
        "  date TEXT PRIMARY KEY,",
        "  flow_year TEXT NOT NULL,",
        "  flow_month TEXT NOT NULL,",
        "  flow_day TEXT NOT NULL,",
        "  source TEXT,",
        "  updated_at TEXT",
        ");",
    ]
    for r in records:
        lines.extend(
            [
                f"INSERT INTO {table} (date, flow_year, flow_month, flow_day, source, updated_at)",
                "VALUES ("
                + ", ".join(
                    [
                        sql_quote(r["date"]),
                        sql_quote(r["flow_year"]),
                        sql_quote(r["flow_month"]),
                        sql_quote(r["flow_day"]),
                        sql_quote("xlsx_2026"),
                        sql_quote(today),
                    ]
                )
                + ")",
                "ON CONFLICT(date) DO UPDATE SET",
                "  flow_year=excluded.flow_year,",
                "  flow_month=excluded.flow_month,",
                "  flow_day=excluded.flow_day,",
                "  source=excluded.source,",
                "  updated_at=excluded.updated_at;",
            ]
        )
    lines.append("COMMIT;")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to xlsx file")
    parser.add_argument("--output", required=True, help="Path to output .sql")
    parser.add_argument("--sheet", default=None, help="Sheet name, default first sheet")
    parser.add_argument("--table", default="bazi_daily_calendar", help="Target table name")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    wb = load_workbook(input_path, data_only=True, read_only=True)
    sheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    header_row, mapping = find_header_row(sheet)
    records = extract_records(sheet, header_row, mapping)
    if not records:
        raise ValueError("No valid data rows found in xlsx")
    validate_records(records)

    sql = build_sql(records, args.table)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8")

    print(f"sheet={sheet.title}")
    print(f"header_row={header_row}")
    print(f"records={len(records)}")
    print(f"output={output_path}")


if __name__ == "__main__":
    main()
